"""RFP Project API routes."""
import io
import os
import re
import shutil
import uuid
import asyncio
import logging
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from ..database import get_db
from ..models.user import User
from ..models.workspace import WorkspaceMember
from ..models.project import RFPProject, AIConfig, AnonymizationMapping, ProjectStatus, EntityType, ComplianceResult, GapAnalysisResult, ProjectMember, ContentReuseResult
from ..models.document import Document, DocumentChunk, DocumentCategory, ProcessingStatus
from ..models.chapter import Chapter, ChapterType, ChapterStatus
from ..models.response_document import ResponseDocument, DocumentFormat, ContentType
from ..schemas.project import (
    ProjectCreate, ProjectUpdate, ProjectOut,
    ImprovementAxisRequest, ImprovementAxisUpdate, GapAnalysisRequest,
    GenerateStructureRequest, PrefillRequest, ComplianceAnalysisRequest,
)
from ..schemas.document import (
    StatisticsOut, AnonymizationMappingOut, AnonymizationReportOut, AnonymizationEntityGroup,
    AnonymizationMappingCreate, AnonymizationMappingUpdate,
    FieldsToCompleteOut, FieldToComplete, FieldReplaceRequest,
)
from ..schemas.response_document import ResponseDocumentOut, ResponseDocumentUpdate, BulkUpdateSelectionRequest
from ..services.ai_service import MistralAIService, create_ai_service, log_ai_usage_from_service
from ..services.vector_service import VectorService
from ..services.anonymization_service import AnonymizationService
from ..services.progress_service import set_progress, get_or_idle
from ..services.moderation_service import moderate_prompt
from .deps import get_current_user, require_project_owner_or_admin, get_workspace_membership, get_project_membership

router = APIRouter(prefix="/projects", tags=["Projects"])
logger = logging.getLogger(__name__)

# Redis progress namespaces (replace in-memory dicts)
_NS_GEN = "structure_gen"
_NS_PREFILL = "prefill"
_NS_GAP = "gap_analysis"
_NS_COMPLIANCE = "compliance"
_NS_REC = "rec_gen"
_NS_DETECT = "detect_deliverables"
_NS_FILL = "fill_deliverables"
_NS_FILL_EXCEL = "fill_excel"
_NS_FILL_PDF = "fill_pdf"
_NS_REANON = "reanon"


async def _get_ai_service(workspace_id: uuid.UUID, db: AsyncSession) -> MistralAIService:
    """Helper to get AI service from workspace config."""
    result = await db.execute(
        select(AIConfig).where(AIConfig.workspace_id == workspace_id)
    )
    config = result.scalar_one_or_none()
    if not config:
        raise HTTPException(
            status_code=400,
            detail="Configuration IA non définie. Configurez le fournisseur IA dans l'administration.",
        )
    provider = getattr(config, "provider", "mistral") or "mistral"
    if provider == "mistral" and not config.mistral_api_key_encrypted:
        raise HTTPException(
            status_code=400,
            detail="Clé API Mistral non configurée. Configurez-la dans l'administration.",
        )
    return create_ai_service(config)


@router.get("/workspace/{workspace_id}", response_model=list[ProjectOut])
async def list_projects(
    workspace_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List projects in a workspace that the current user has access to.

    A user can see a project if:
    - They are a system admin, OR
    - They are an explicit project member (ProjectMember)
    """
    from ..models.user import UserRole

    # Verify workspace membership first
    is_admin = current_user.role == UserRole.ADMIN
    if not is_admin:
        ws_member_result = await db.execute(
            select(WorkspaceMember)
            .where(WorkspaceMember.workspace_id == workspace_id)
            .where(WorkspaceMember.user_id == current_user.id)
        )
        if not ws_member_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Accès non autorisé à cet espace de travail")

    if is_admin:
        result = await db.execute(
            select(RFPProject)
            .where(RFPProject.workspace_id == workspace_id)
            .order_by(RFPProject.updated_at.desc())
        )
        projects = list(result.scalars().all())
    else:
        # Only show projects where user is an explicit member
        result = await db.execute(
            select(RFPProject)
            .join(ProjectMember, ProjectMember.project_id == RFPProject.id)
            .where(RFPProject.workspace_id == workspace_id)
            .where(ProjectMember.user_id == current_user.id)
            .order_by(RFPProject.updated_at.desc())
        )
        projects = list(result.scalars().all())

    project_list = []
    for p in projects:
        doc_count = (await db.execute(
            select(func.count()).where(Document.project_id == p.id)
        )).scalar() or 0
        ch_count = (await db.execute(
            select(func.count()).where(Chapter.project_id == p.id)
        )).scalar() or 0

        # Get current user's role in this project
        pm_result = await db.execute(
            select(ProjectMember)
            .where(ProjectMember.project_id == p.id)
            .where(ProjectMember.user_id == current_user.id)
        )
        pm = pm_result.scalar_one_or_none()
        user_role = pm.role if pm else None

        project_list.append(ProjectOut(
            id=str(p.id),
            workspace_id=str(p.workspace_id),
            name=p.name,
            description=p.description,
            client_name=p.client_name,
            company_name=getattr(p, 'company_name', '') or '',
            rfp_reference=p.rfp_reference,
            deadline=p.deadline,
            status=p.status.value,
            improvement_axes=p.improvement_axes,
            ai_context=p.ai_context or "",
            enabled_categories=p.enabled_categories or ["old_rfp", "old_response", "new_rfp"],
            context_mode=p.context_mode or "rag",
            created_by=str(p.created_by),
            created_at=p.created_at,
            updated_at=p.updated_at,
            document_count=doc_count,
            chapter_count=ch_count,
            current_user_role=user_role,
        ))
    return project_list


@router.post("/workspace/{workspace_id}", response_model=ProjectOut, status_code=201)
async def create_project(
    workspace_id: uuid.UUID,
    request: ProjectCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a new RFP project. Any workspace member can create a project."""
    # Moderate user-provided text fields
    if request.ai_context:
        moderation = moderate_prompt(request.ai_context, "ai_context")
        if not moderation:
            raise HTTPException(status_code=422, detail=moderation.message)

    # Verify workspace membership
    await get_workspace_membership(workspace_id, current_user, db)

    project = RFPProject(
        workspace_id=workspace_id,
        name=request.name,
        description=request.description,
        client_name=request.client_name,
        company_name=request.company_name,
        rfp_reference=request.rfp_reference,
        deadline=request.deadline,
        ai_context=request.ai_context,
        enabled_categories=request.enabled_categories,
        context_mode=request.context_mode,
        created_by=current_user.id,
    )
    db.add(project)
    await db.flush()

    # Auto-add creator as project owner
    project_member = ProjectMember(
        project_id=project.id,
        user_id=current_user.id,
        role="owner",
    )
    db.add(project_member)
    await db.commit()
    await db.refresh(project)

    return ProjectOut(
        id=str(project.id),
        workspace_id=str(project.workspace_id),
        name=project.name,
        description=project.description,
        client_name=project.client_name,
        company_name=project.company_name or '',
        rfp_reference=project.rfp_reference,
        deadline=project.deadline,
        status=project.status.value,
        improvement_axes=project.improvement_axes,
        ai_context=project.ai_context or "",
        enabled_categories=project.enabled_categories or ["old_rfp", "old_response", "new_rfp"],
        context_mode=project.context_mode or "rag",
        created_by=str(project.created_by),
        created_at=project.created_at,
        updated_at=project.updated_at,
        document_count=0,
        chapter_count=0,
    )


@router.get("/{project_id}", response_model=ProjectOut)
async def get_project(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get project details."""
    from ..models.user import UserRole

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Check access: admin or explicit project member
    membership = await get_project_membership(project_id, current_user, db)
    user_role = membership.role if membership else None

    doc_count = (await db.execute(
        select(func.count()).where(Document.project_id == project_id)
    )).scalar() or 0
    ch_count = (await db.execute(
        select(func.count()).where(Chapter.project_id == project_id)
    )).scalar() or 0

    return ProjectOut(
        id=str(project.id),
        workspace_id=str(project.workspace_id),
        name=project.name,
        description=project.description,
        client_name=project.client_name,
        company_name=project.company_name or '',
        rfp_reference=project.rfp_reference,
        deadline=project.deadline,
        status=project.status.value,
        improvement_axes=project.improvement_axes,
        ai_context=project.ai_context or "",
        enabled_categories=project.enabled_categories or ["old_rfp", "old_response", "new_rfp"],
        context_mode=project.context_mode or "rag",
        created_by=str(project.created_by),
        created_at=project.created_at,
        updated_at=project.updated_at,
        document_count=doc_count,
        chapter_count=ch_count,
        current_user_role=user_role,
    )


@router.put("/{project_id}", response_model=ProjectOut)
async def update_project(
    project_id: uuid.UUID,
    request: ProjectUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update project details."""
    from ..models.user import UserRole

    # Moderate user-provided text fields
    for field_name in ("ai_context", "improvement_axes"):
        value = getattr(request, field_name, None)
        if value:
            moderation = moderate_prompt(value, field_name)
            if not moderation:
                raise HTTPException(status_code=422, detail=moderation.message)

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Check project membership
    membership = await get_project_membership(project_id, current_user, db)
    is_admin = current_user.role == UserRole.ADMIN
    is_owner = membership and membership.role == "owner"

    # AI config fields are restricted to project owner or admin
    ai_config_fields = {"ai_context", "enabled_categories", "context_mode"}
    general_fields = {"name", "description", "client_name", "company_name", "rfp_reference", "deadline", "improvement_axes"}

    # Non-owner/non-admin editors can only update general fields
    if not is_admin and not is_owner:
        # Viewers cannot update at all
        if membership and membership.role == "viewer":
            raise HTTPException(status_code=403, detail="Les lecteurs ne peuvent pas modifier le projet")

    for field in general_fields:
        value = getattr(request, field, None)
        if value is not None:
            setattr(project, field, value)

    for field in ai_config_fields:
        value = getattr(request, field, None)
        if value is not None:
            if not is_admin and not is_owner:
                raise HTTPException(status_code=403, detail="Seul un administrateur ou le propriétaire du projet peut modifier la configuration IA")
            setattr(project, field, value)

    if request.status is not None:
        project.status = request.status

    await db.commit()
    await db.refresh(project)

    return await get_project(project_id, current_user, db)


@router.delete("/{project_id}", status_code=204)
async def delete_project(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a project and all its data (owner or admin only)."""
    await require_project_owner_or_admin(project_id, current_user, db)

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    VectorService.delete_project_data(str(project_id))

    # Clean up stale Redis progress keys for this project's documents
    from ..services.progress_service import delete_many
    doc_result = await db.execute(
        select(Document.id).where(Document.project_id == project_id)
    )
    doc_ids = [str(row[0]) for row in doc_result.all()]
    if doc_ids:
        delete_many("document", doc_ids)

    await db.delete(project)
    await db.commit()


# ── AI-powered features ──

@router.get("/{project_id}/gap-analysis")
async def get_gap_analysis(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get the latest persisted gap analysis for a project."""
    result = await db.execute(
        select(GapAnalysisResult)
        .where(GapAnalysisResult.project_id == project_id)
        .order_by(GapAnalysisResult.created_at.desc())
        .limit(1)
    )
    gr = result.scalar_one_or_none()
    if not gr:
        return {"analysis": None}

    return {
        "analysis": {
            "id": str(gr.id),
            "summary": gr.summary,
            "new_requirements": gr.new_requirements or [],
            "removed_requirements": gr.removed_requirements or [],
            "modified_requirements": gr.modified_requirements or [],
            "unchanged_requirements": gr.unchanged_requirements or [],
            "created_at": gr.created_at.isoformat() if gr.created_at else None,
        }
    }


@router.post("/{project_id}/gap-analysis")
async def analyze_gap(
    project_id: uuid.UUID,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch gap analysis as a background task (returns immediately)."""
    pid = str(project_id)

    existing = get_or_idle(_NS_GAP, pid)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Analyse des ecarts deja en cours")

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    set_progress(_NS_GAP, pid, {
        "status": "running", "step": "starting", "progress": 0,
        "message": "Demarrage de l'analyse des ecarts...",
    })

    from ..tasks.project_tasks import gap_analysis_task
    gap_analysis_task.apply_async(
        args=(str(project_id), str(project.workspace_id)), priority=3,
    )

    return {"success": True, "message": "Analyse des ecarts lancee en arriere-plan"}


@router.get("/{project_id}/gap-analysis-status")
async def get_gap_analysis_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of gap analysis."""
    pid = str(project_id)
    return get_or_idle(_NS_GAP, pid)


@router.get("/{project_id}/gap-analysis/export-pdf")
async def export_gap_analysis_pdf(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export the latest gap analysis as a structured PDF document."""
    import fitz  # PyMuPDF

    result = await db.execute(
        select(GapAnalysisResult)
        .where(GapAnalysisResult.project_id == project_id)
        .order_by(GapAnalysisResult.created_at.desc())
        .limit(1)
    )
    gr = result.scalar_one_or_none()
    if not gr:
        raise HTTPException(status_code=404, detail="Aucune analyse des ecarts disponible")

    proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = proj_result.scalar_one_or_none()
    project_name = project.name if project else "Projet"

    # Build PDF with PyMuPDF
    doc = fitz.open()

    MARGIN = 50
    PAGE_W, PAGE_H = fitz.paper_size("a4")
    TEXT_W = PAGE_W - 2 * MARGIN
    Y_BOTTOM = PAGE_H - MARGIN

    # Colors
    COL_TITLE = (0.106, 0.227, 0.361)   # #1B3A5C – dark blue
    COL_BLUE = (0.082, 0.396, 0.753)     # #1565C0 – new requirements
    COL_RED = (0.776, 0.157, 0.157)      # #C62828 – removed
    COL_ORANGE = (0.937, 0.424, 0.0)     # #EF6C00 – modified
    COL_GREEN = (0.180, 0.490, 0.196)    # #2E7D32 – unchanged
    COL_BLACK = (0.0, 0.0, 0.0)
    COL_GRAY = (0.4, 0.4, 0.4)
    COL_LIGHTGRAY = (0.92, 0.92, 0.92)
    COL_WHITE = (1.0, 1.0, 1.0)

    PRIORITY_COLORS = {
        "high": (0.898, 0.224, 0.208),    # red
        "medium": (0.937, 0.424, 0.0),    # orange
        "low": (0.180, 0.490, 0.196),     # green
    }
    PRIORITY_LABELS = {"high": "Haute", "medium": "Moyenne", "low": "Basse"}

    page = doc.new_page(width=PAGE_W, height=PAGE_H)
    y = MARGIN

    def _new_page():
        nonlocal page, y
        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        y = MARGIN

    def _check_space(needed: float):
        nonlocal y
        if y + needed > Y_BOTTOM:
            _new_page()

    def _write(text: str, fontsize: float = 10, color=COL_BLACK, bold: bool = False, indent: float = 0, max_width: float = 0):
        nonlocal y
        fontname = "hebo" if bold else "helv"
        w = max_width or (TEXT_W - indent)
        lines = []
        for paragraph in text.split("\n"):
            if not paragraph.strip():
                lines.append("")
                continue
            words = paragraph.split()
            current_line = ""
            for word in words:
                test = f"{current_line} {word}".strip()
                tw = fitz.get_text_length(test, fontname=fontname, fontsize=fontsize)
                if tw > w and current_line:
                    lines.append(current_line)
                    current_line = word
                else:
                    current_line = test
            if current_line:
                lines.append(current_line)

        line_h = fontsize * 1.4
        for line in lines:
            _check_space(line_h)
            page.insert_text(
                fitz.Point(MARGIN + indent, y + fontsize),
                line, fontsize=fontsize, fontname=fontname, color=color,
            )
            y += line_h

    def _draw_separator():
        nonlocal y
        _check_space(10)
        page.draw_line(
            fitz.Point(MARGIN, y), fitz.Point(MARGIN + TEXT_W, y),
            color=COL_LIGHTGRAY, width=0.5,
        )
        y += 10

    def _draw_section_header(icon_text: str, title: str, count: int, color):
        nonlocal y
        _check_space(35)
        # Section background bar
        bar_rect = fitz.Rect(MARGIN, y, MARGIN + TEXT_W, y + 28)
        page.draw_rect(bar_rect, color=None, fill=(*color, 0.08) if hasattr(color, '__len__') else color)
        page.draw_rect(fitz.Rect(MARGIN, y, MARGIN + 4, y + 28), color=None, fill=color)
        page.insert_text(
            fitz.Point(MARGIN + 12, y + 18),
            f"{icon_text}  {title} ({count})", fontsize=12, fontname="hebo", color=color,
        )
        y += 35

    # ── Cover / Title ──
    _write("Analyse des Ecarts", fontsize=22, color=COL_TITLE, bold=True)
    _write(project_name, fontsize=13, color=COL_GRAY)
    if gr.created_at:
        _write(f"Date : {gr.created_at.strftime('%d/%m/%Y %H:%M')}", fontsize=9, color=COL_GRAY)
    y += 8

    # ── Stats summary bar ──
    new_count = len(gr.new_requirements or [])
    removed_count = len(gr.removed_requirements or [])
    modified_count = len(gr.modified_requirements or [])
    unchanged_count = len(gr.unchanged_requirements or [])
    total = new_count + removed_count + modified_count + unchanged_count

    _check_space(50)
    stats = [
        (f"{new_count} nouvelles", COL_BLUE),
        (f"{removed_count} supprimees", COL_RED),
        (f"{modified_count} modifiees", COL_ORANGE),
        (f"{unchanged_count} inchangees", COL_GREEN),
    ]
    stat_x = MARGIN
    for label, color in stats:
        page.insert_text(fitz.Point(stat_x, y + 12), label, fontsize=10, fontname="hebo", color=color)
        stat_x += fitz.get_text_length(label, fontname="hebo", fontsize=10) + 20
    y += 22

    # Total bar
    if total > 0:
        bar_y = y
        bar_h = 6
        page.draw_rect(fitz.Rect(MARGIN, bar_y, MARGIN + TEXT_W, bar_y + bar_h), color=None, fill=COL_LIGHTGRAY)
        x_off = MARGIN
        for count, color in [(new_count, COL_BLUE), (removed_count, COL_RED), (modified_count, COL_ORANGE), (unchanged_count, COL_GREEN)]:
            seg_w = TEXT_W * count / total
            if seg_w > 0:
                page.draw_rect(fitz.Rect(x_off, bar_y, x_off + seg_w, bar_y + bar_h), color=None, fill=color)
                x_off += seg_w
        y += bar_h + 10

    _draw_separator()

    # ── Summary ──
    if gr.summary:
        _write("Resume", fontsize=13, color=COL_TITLE, bold=True)
        y += 4
        _write(gr.summary, fontsize=10, color=COL_GRAY)
        y += 10
        _draw_separator()

    # ── New Requirements ──
    new_reqs = gr.new_requirements or []
    if new_reqs:
        _draw_section_header("\u25b6", "Nouvelles exigences", len(new_reqs), COL_BLUE)
        for req in new_reqs:
            _check_space(40)
            title = req.get("title", "")
            priority = req.get("priority", "medium")
            p_label = PRIORITY_LABELS.get(priority, priority)
            p_color = PRIORITY_COLORS.get(priority, COL_GRAY)
            _write(f"[{p_label}]  {title}", fontsize=10, bold=True, indent=10, color=p_color)
            desc = req.get("description", "")
            if desc:
                _write(desc, fontsize=9, color=COL_GRAY, indent=20)
            y += 4
        _draw_separator()

    # ── Removed Requirements ──
    removed_reqs = gr.removed_requirements or []
    if removed_reqs:
        _draw_section_header("\u2716", "Exigences supprimees", len(removed_reqs), COL_RED)
        for req in removed_reqs:
            _check_space(35)
            _write(req.get("title", ""), fontsize=10, bold=True, indent=10, color=COL_RED)
            desc = req.get("description", "")
            if desc:
                _write(desc, fontsize=9, color=COL_GRAY, indent=20)
            y += 4
        _draw_separator()

    # ── Modified Requirements ──
    modified_reqs = gr.modified_requirements or []
    if modified_reqs:
        _draw_section_header("\u270E", "Exigences modifiees", len(modified_reqs), COL_ORANGE)
        for req in modified_reqs:
            _check_space(70)
            _write(req.get("title", ""), fontsize=10, bold=True, indent=10, color=COL_ORANGE)
            old_desc = req.get("old_description", "")
            new_desc = req.get("new_description", "")
            impact = req.get("impact", "")
            if old_desc:
                _write(f"Avant : {old_desc}", fontsize=9, color=COL_RED, indent=20)
            if new_desc:
                _write(f"Apres : {new_desc}", fontsize=9, color=COL_GREEN, indent=20)
            if impact:
                _write(f"Impact : {impact}", fontsize=9, color=COL_GRAY, indent=20)
            y += 4
        _draw_separator()

    # ── Unchanged Requirements ──
    unchanged_reqs = gr.unchanged_requirements or []
    if unchanged_reqs:
        _draw_section_header("\u2714", "Exigences inchangees", len(unchanged_reqs), COL_GREEN)
        for req in unchanged_reqs:
            _check_space(20)
            _write(f"- {req.get('title', '')}", fontsize=9, indent=10, color=COL_GREEN)
            desc = req.get("description", "")
            if desc:
                _write(desc, fontsize=8, color=COL_GRAY, indent=20)
        _draw_separator()

    # ── Footer on each page ──
    for i in range(len(doc)):
        p = doc[i]
        footer_text = f"Analyse des ecarts - {project_name} | Page {i + 1}/{len(doc)}"
        tw = fitz.get_text_length(footer_text, fontname="helv", fontsize=8)
        p.insert_text(
            fitz.Point(PAGE_W / 2 - tw / 2, PAGE_H - 25),
            footer_text, fontsize=8, fontname="helv", color=COL_GRAY,
        )

    pdf_bytes = doc.tobytes()
    doc.close()

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="analyse_ecarts_{project_name}.pdf"'},
    )


async def _run_gap_analysis(project_id: uuid.UUID, workspace_id: uuid.UUID):
    """Background task for gap analysis.

    DB connections are released during the slow AI call to minimize pool pressure.
    """
    from ..database import task_session
    pid = str(project_id)

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_GAP, pid, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
        })

    try:
        # ── Phase 1: Load config + anonymize (short DB session) ──
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            _update("searching", 10, "Chargement intégral des documents d'appel d'offres...")

            # Load ALL chunks from DB (pre-anonymized) — no vector search limits
            anon_old = await _get_all_chunks_anonymized_by_category(
                db, project_id, DocumentCategory.OLD_RFP
            )
            anon_new = await _get_all_chunks_anonymized_by_category(
                db, project_id, DocumentCategory.NEW_RFP
            )

            if not anon_old.strip() or not anon_new.strip():
                set_progress(_NS_GAP, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Documents d'ancien et/ou de nouvel appel d'offres manquants",
                })
                return

            _update("anonymizing", 20, "Preparation de l'analyse...")
        # DB released

        # ── Phase 2: AI analysis (NO DB connection held) ──
        _update("analyzing", 40, "Analyse IA des ecarts en cours...")
        analysis = await ai_service.analyze_gap(anon_old, anon_new)

        # Log AI usage for gap analysis
        async with task_session() as usage_db:
            await log_ai_usage_from_service(usage_db, project_id, "gap_analysis", ai_service)

        # ── Phase 3: Deanonymize + save (short DB session) ──
        _update("deanonymizing", 75, "Deanonymisation des resultats...")
        async with task_session() as db:
            for key in ["summary"]:
                if key in analysis and isinstance(analysis[key], str):
                    analysis[key] = await AnonymizationService.deanonymize_text(analysis[key], project_id, db)
            for req_list_key in ["new_requirements", "removed_requirements", "modified_requirements", "unchanged_requirements"]:
                for req in analysis.get(req_list_key, []):
                    for field in ["title", "description", "old_description", "new_description", "impact", "source_old", "source_new"]:
                        if field in req and req[field]:
                            req[field] = await AnonymizationService.deanonymize_text(req[field], project_id, db)

            _update("saving", 90, "Enregistrement des resultats...")
            gr = GapAnalysisResult(
                project_id=project_id,
                summary=analysis.get("summary", ""),
                new_requirements=analysis.get("new_requirements", []),
                removed_requirements=analysis.get("removed_requirements", []),
                modified_requirements=analysis.get("modified_requirements", []),
                unchanged_requirements=analysis.get("unchanged_requirements", []),
            )
            db.add(gr)
            await db.commit()

        set_progress(_NS_GAP, pid, {
            "status": "completed", "step": "done", "progress": 100,
            "message": "Analyse des ecarts terminee",
        })

    except Exception as e:
        logger.exception("Gap analysis failed for project %s", project_id)
        set_progress(_NS_GAP, pid, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


async def _get_all_chunks_by_category(
    db: AsyncSession, project_id: uuid.UUID, category: DocumentCategory
) -> str:
    """Get ALL document chunks for a category, ordered sequentially.
    Returns the full concatenated text."""
    result = await db.execute(
        select(DocumentChunk)
        .join(Document, Document.id == DocumentChunk.document_id)
        .where(Document.project_id == project_id)
        .where(Document.category == category)
        .order_by(Document.original_filename, DocumentChunk.page_number, DocumentChunk.chunk_index)
    )
    chunks = result.scalars().all()
    return "\n\n".join([c.content for c in chunks if c.content.strip()])


async def _get_all_chunks_anonymized_by_category(
    db: AsyncSession, project_id: uuid.UUID, category: DocumentCategory
) -> str:
    """Get ALL pre-anonymized document chunks for a category.
    Uses anonymized_content already computed at upload time, avoiding
    redundant NER inference."""
    result = await db.execute(
        select(DocumentChunk, Document.original_filename)
        .join(Document, Document.id == DocumentChunk.document_id)
        .where(Document.project_id == project_id)
        .where(Document.category == category)
        .order_by(Document.original_filename, DocumentChunk.page_number, DocumentChunk.chunk_index)
    )
    rows = result.all()
    parts = []
    current_doc = None
    current_section = None
    for chunk, doc_name in rows:
        text = (chunk.anonymized_content or chunk.content or "").strip()
        if not text:
            continue
        # Add document header when switching to a new document
        if doc_name != current_doc:
            current_doc = doc_name
            current_section = None
            parts.append(f"\n\n=== DOCUMENT: {doc_name} ===\n")
        # Add section header when switching to a new section
        section = chunk.section_title or ""
        if section and section != current_section:
            current_section = section
            parts.append(f"\n--- {section} ---\n")
        parts.append(text)
    return "\n\n".join(parts)


async def _get_chunks_anonymized_by_document_ids(
    db: AsyncSession, project_id: uuid.UUID, document_ids: list[str]
) -> str:
    """Get pre-anonymized chunks for a specific set of document IDs."""
    doc_uuids = [uuid.UUID(did) for did in document_ids]
    result = await db.execute(
        select(DocumentChunk, Document.original_filename)
        .join(Document, Document.id == DocumentChunk.document_id)
        .where(Document.project_id == project_id)
        .where(Document.id.in_(doc_uuids))
        .order_by(Document.original_filename, DocumentChunk.page_number, DocumentChunk.chunk_index)
    )
    rows = result.all()
    parts = []
    current_doc = None
    current_section = None
    for chunk, doc_name in rows:
        text = (chunk.anonymized_content or chunk.content or "").strip()
        if not text:
            continue
        if doc_name != current_doc:
            current_doc = doc_name
            current_section = None
            parts.append(f"\n\n=== DOCUMENT: {doc_name} ===\n")
        section = chunk.section_title or ""
        if section and section != current_section:
            current_section = section
            parts.append(f"\n--- {section} ---\n")
        parts.append(text)
    return "\n\n".join(parts)


async def _get_chunks_anonymized_by_categories(
    db: AsyncSession, project_id: uuid.UUID, categories: list[str]
) -> str:
    """Get pre-anonymized chunks for multiple document categories."""
    from ..models.document import DocumentCategory
    cat_enums = []
    for c in categories:
        try:
            cat_enums.append(DocumentCategory(c))
        except ValueError:
            continue
    if not cat_enums:
        return ""
    result = await db.execute(
        select(DocumentChunk, Document.original_filename)
        .join(Document, Document.id == DocumentChunk.document_id)
        .where(Document.project_id == project_id)
        .where(Document.category.in_(cat_enums))
        .order_by(Document.original_filename, DocumentChunk.page_number, DocumentChunk.chunk_index)
    )
    rows = result.all()
    parts = []
    current_doc = None
    current_section = None
    for chunk, doc_name in rows:
        text = (chunk.anonymized_content or chunk.content or "").strip()
        if not text:
            continue
        if doc_name != current_doc:
            current_doc = doc_name
            current_section = None
            parts.append(f"\n\n=== DOCUMENT: {doc_name} ===\n")
        section = chunk.section_title or ""
        if section and section != current_section:
            current_section = section
            parts.append(f"\n--- {section} ---\n")
        parts.append(text)
    return "\n\n".join(parts)


async def _get_generated_chapters_context(
    db: AsyncSession, project_id: uuid.UUID
) -> str:
    """Get anonymized generated chapter content for use as context."""
    _anon = AnonymizationService.apply_existing_mappings
    result = await db.execute(
        select(Chapter)
        .where(Chapter.project_id == project_id)
        .where(Chapter.content != "")
        .order_by(Chapter.order)
    )
    chapters = result.scalars().all()
    parts = ["\n\n=== CONTENU GÉNÉRÉ (CHAPITRES RÉDIGÉS) ===\n"]
    for ch in chapters:
        anon_title = await _anon(ch.title, project_id, db)
        anon_content = await _anon(ch.content, project_id, db)
        parts.append(f"## {anon_title}\n{anon_content}")
    return "\n\n".join(parts) if len(parts) > 1 else ""


async def _get_image_analyses_by_category(
    db: AsyncSession, project_id: uuid.UUID, category: DocumentCategory
) -> str:
    """Get anonymized image analysis descriptions for documents in a category.

    Returns a formatted text block with image descriptions, OCR text, and key
    information extracted by the vision model during document processing.
    Only includes images that were successfully analyzed (not decorative icons/logos).
    """
    from ..models.document import DocumentImage, ImageAnalysisStatus

    result = await db.execute(
        select(DocumentImage, Document.original_filename)
        .join(Document, Document.id == DocumentImage.document_id)
        .where(Document.project_id == project_id)
        .where(Document.category == category)
        .where(DocumentImage.analysis_status == ImageAnalysisStatus.COMPLETED.value)
        .where(DocumentImage.image_type.notin_(["logo", "icone"]))
        .order_by(Document.original_filename, DocumentImage.page_number)
    )
    rows = result.all()
    if not rows:
        return ""

    parts = []
    current_doc = None
    for img, doc_name in rows:
        # Skip non-informative images
        desc = (img.anonymized_description or "").strip()
        ocr = (img.anonymized_ocr_text or img.ocr_text or "").strip()
        key_info = img.key_information or []
        if not desc and not ocr and not key_info:
            continue

        if doc_name != current_doc:
            current_doc = doc_name
            parts.append(f"\n=== IMAGES DU DOCUMENT: {doc_name} ===")

        img_parts = []
        img_type = img.image_type or img.image_category or "autre"
        img_parts.append(f"[Image page {img.page_number} — type: {img_type}]")
        if desc:
            img_parts.append(f"  Description: {desc}")
        if key_info:
            img_parts.append(f"  Informations clés: {', '.join(str(k) for k in key_info)}")
        if ocr:
            img_parts.append(f"  Texte extrait: {ocr[:500]}")
        if img.suggested_usage:
            img_parts.append(f"  Usage suggéré: {img.suggested_usage}")

        parts.append("\n".join(img_parts))

    return "\n\n".join(parts)


async def _get_full_text_anonymized_by_category(
    db: AsyncSession, project_id: uuid.UUID, category: DocumentCategory,
) -> str:
    """Get full anonymized text for documents in a category (full context mode).

    Uses Document.anonymized_full_text stored at upload time — the raw extracted
    text anonymized as a single block, exactly like pasting into a chat.
    Falls back to reassembled chunks for documents uploaded before this feature.
    """
    result = await db.execute(
        select(Document)
        .where(Document.project_id == project_id)
        .where(Document.category == category)
        .where(Document.processing_status == ProcessingStatus.COMPLETED)
        .order_by(Document.original_filename)
    )
    docs: list[Document] = result.scalars().all()
    parts: list[str] = []
    fallback_doc_ids: list[uuid.UUID] = []
    for doc in docs:
        anon = (doc.anonymized_full_text or "").strip()
        if anon:
            parts.append(f"\n\n=== DOCUMENT: {doc.original_filename} ===\n")
            parts.append(anon)
        else:
            fallback_doc_ids.append(doc.id)

    # Fallback: reassemble from chunks for older documents without full_text
    if fallback_doc_ids:
        chunk_result = await db.execute(
            select(DocumentChunk, Document.original_filename)
            .join(Document, Document.id == DocumentChunk.document_id)
            .where(DocumentChunk.document_id.in_(fallback_doc_ids))
            .order_by(Document.original_filename, DocumentChunk.page_number, DocumentChunk.chunk_index)
        )
        current_doc = None
        for chunk, doc_name in chunk_result.all():
            text = (chunk.anonymized_content or chunk.content or "").strip()
            if not text:
                continue
            if doc_name != current_doc:
                current_doc = doc_name
                parts.append(f"\n\n=== DOCUMENT: {doc_name} ===\n")
            parts.append(text)

    return "\n\n".join(parts)


@router.post("/{project_id}/generate-structure")
async def generate_structure(
    project_id: uuid.UUID,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch structure generation as a background task (returns immediately)."""
    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Quick check: has new RFP docs?
    doc_count = (await db.execute(
        select(func.count()).select_from(Document)
        .where(Document.project_id == project_id)
        .where(Document.category == DocumentCategory.NEW_RFP)
    )).scalar() or 0
    if doc_count == 0:
        raise HTTPException(status_code=400, detail="Aucun document de nouvel appel d'offres indexe")

    pid = str(project_id)

    # Check if already running
    existing = get_or_idle(_NS_GEN, pid)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Generation deja en cours")

    workspace_id = project.workspace_id
    set_progress(_NS_GEN, pid, {
        "status": "running",
        "step": "starting",
        "progress": 0,
        "message": "Demarrage de la generation...",
    })

    from ..tasks.project_tasks import generate_structure_task
    generate_structure_task.apply_async(
        args=(str(project_id), str(workspace_id)), priority=0,
    )

    return {"success": True, "message": "Generation lancee en arriere-plan"}


@router.get("/{project_id}/generation-status")
async def get_generation_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of the structure generation task."""
    pid = str(project_id)
    return get_or_idle(_NS_GEN, pid)


def _make_stream_progress_callback(pid: str, step: str, start_pct: int, end_pct: int, label: str, max_tokens: int):
    """Create a streaming progress callback for use with generate_streaming().

    Returns an async callback(token_count, char_count) that updates
    _generation_progress based on *real* tokens received from Mistral.
    """
    import time
    t0 = time.monotonic()

    async def _on_progress(token_count: int, char_count: int):
        elapsed = int(time.monotonic() - t0)
        # Real progress based on actual tokens received vs expected max
        ratio = min(token_count / max_tokens, 0.95)
        pct = start_pct + int((end_pct - start_pct) * ratio)
        set_progress(_NS_GEN, pid, {
            "status": "running",
            "step": step,
            "progress": pct,
            "message": f"{label} — {token_count} tokens recus ({char_count:,} car.) — {elapsed}s",
        })

    return _on_progress


async def _run_structure_generation(project_id: uuid.UUID, workspace_id: uuid.UUID):
    """Background task for the full structure generation pipeline."""
    from ..database import task_session
    pid = str(project_id)

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_GEN, pid, {
            "status": "running",
            "step": step,
            "progress": progress,
            "message": message,
        })

    try:
        # ── Phase 1: Load data from DB (short-lived session) ──
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            # Load project AI context and anonymize it (may contain client/company names)
            proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
            proj = proj_result.scalar_one()
            raw_ai_context = proj.ai_context or ""
            proj_ai_context = await AnonymizationService.apply_existing_mappings(
                raw_ai_context, project_id, db
            ) if raw_ai_context else ""
            proj_company_name = getattr(proj, 'company_name', '') or ''
            proj_client_name = proj.client_name or ''

            _update("loading", 5, "Chargement des documents (AO, ancien AO, ancienne reponse)...")
            # Load all 3 categories in parallel for speed
            anon_new_rfp, anon_old_rfp, anon_old_response = await asyncio.gather(
                _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.NEW_RFP),
                _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.OLD_RFP),
                _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.OLD_RESPONSE),
            )
            if not anon_new_rfp:
                set_progress(_NS_GEN, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Aucun document de nouvel appel d'offres indexe",
                })
                return

        # DB session released — all data is now in memory

        new_rfp_k = len(anon_new_rfp) // 1000
        old_rfp_k = len(anon_old_rfp) // 1000
        old_resp_k = len(anon_old_response) // 1000
        _update("loading", 15,
                f"Documents charges: nouvel AO ({new_rfp_k}K car.), "
                f"ancien AO ({old_rfp_k}K car.), "
                f"ancienne reponse ({old_resp_k}K car.)")

        # ── Phase 2: Gap analysis (streamed, real token progress) ──
        gap_analysis = None
        if anon_old_rfp:
            _update("gap_analysis", 15, "Analyse des ecarts ancien/nouveau AO...")
            gap_cb = _make_stream_progress_callback(
                pid, "gap_analysis", 15, 40, "Analyse des ecarts", max_tokens=12000,
            )
            gap_analysis = await ai_service.analyze_gap(
                anon_old_rfp, anon_new_rfp, on_progress=gap_cb,
            )
            # Log AI usage for gap analysis within structure generation
            async with task_session() as usage_db:
                await log_ai_usage_from_service(usage_db, project_id, "structure_gap_analysis", ai_service)
            gap_new = len(gap_analysis.get("new_requirements", []))
            gap_mod = len(gap_analysis.get("modified_requirements", []))
            gap_del = len(gap_analysis.get("removed_requirements", []))
            _update("gap_analysis", 40,
                    f"Ecarts identifies: {gap_new} nouvelles, {gap_mod} modifiees, {gap_del} supprimees")

        # ── Phase 3: Check for response documents ──
        # Only generate chapters for "redaction" type documents (text to write)
        # "completion" type documents (Excel/PDF to fill in) are handled separately
        resp_docs = []
        async with task_session() as db:
            result = await db.execute(
                select(ResponseDocument)
                .where(ResponseDocument.project_id == project_id)
                .where(ResponseDocument.is_selected == True)
                .order_by(ResponseDocument.order)
            )
            all_docs = result.scalars().all()

            # Safety net: skip docs that are clearly completion-type even if
            # content_type was set to REDACTION (e.g. xlsx BPU misclassified)
            _completion_kw = [
                "bpu", "bordereau", "dqe", "dpgf",
                "dc1", "dc2", "dc3", "dc4",
                "attri", "noti", "acte d'engagement",
                "formulaire", "cerfa",
                "a completer", "à compléter",
                "a remplir", "à remplir",
                "grille", "questionnaire",
                "annexe conformit", "annexe rgpd",
            ]

            def _is_truly_redaction(rd) -> bool:
                """Return True only if the document really needs chapter generation."""
                title_lower = (rd.title or "").lower()
                if rd.expected_format == DocumentFormat.XLSX:
                    return False
                if any(kw in title_lower for kw in _completion_kw):
                    return False
                return (
                    rd.content_type == ContentType.REDACTION
                    or rd.content_type in ("redaction", "REDACTION")
                )

            # Anonymize document titles/descriptions before sending to AI
            _anon = AnonymizationService.apply_existing_mappings
            resp_docs = []
            for rd in all_docs:
                if _is_truly_redaction(rd):
                    anon_t = await _anon(rd.title or "", project_id, db)
                    anon_d = await _anon(rd.description or "", project_id, db)
                    resp_docs.append((str(rd.id), anon_t, anon_d))
            completion_docs_count = sum(
                1 for rd in all_docs
                if not _is_truly_redaction(rd)
            )

        # ── Phase 3: Generate structure ──
        order = 0
        created_count = 0
        delta_stats = {"new": 0, "modified": 0, "unchanged": 0}
        failed_doc_ids: set = set()

        # If deliverables were detected but all are completion-type, no chapters to generate
        has_deliverables = (len(resp_docs) + completion_docs_count) > 0
        if has_deliverables and not resp_docs:
            set_progress(_NS_GEN, pid, {
                "status": "completed",
                "step": "done",
                "progress": 100,
                "chapters_created": 0,
                "delta_stats": delta_stats,
                "has_gap_analysis": gap_analysis is not None,
                "completion_docs_count": completion_docs_count,
                "message": f"Aucun document a rediger detecte — {completion_docs_count} document(s) "
                           f"a completer (Excel/PDF) a traiter dans l'onglet Livrables",
            })
            return

        if resp_docs:
            import time

            # ── Phase 3a: Summarize RFP once (saves re-sending 60K per doc) ──
            total_docs = len(resp_docs)
            rfp_summary = ""
            if total_docs > 1:
                _update("summarizing", 40, "Resume de l'AO pour generation parallele...")
                sum_cb = _make_stream_progress_callback(
                    pid, "summarizing", 40, 50, "Resume AO", max_tokens=6000,
                )
                rfp_summary = await ai_service.summarize_rfp_for_structure(
                    anon_new_rfp, on_progress=sum_cb,
                )
                logger.info("RFP summary: %d chars (vs %d full)", len(rfp_summary), len(anon_new_rfp))

            # ── Phase 3b: Generate structure for all docs IN PARALLEL ──
            all_doc_structures: list[tuple[str, list]] = []
            _doc_done_count = 0

            sem = asyncio.Semaphore(10)  # Mistral/Scaleway handle 24 req/s

            async def _gen_one_doc(doc_idx, doc_id, doc_title, doc_desc):
                nonlocal _doc_done_count
                t0_doc = time.monotonic()

                async def _doc_progress(token_count: int, char_count: int,
                                        _idx=doc_idx, _title=doc_title, _t0=t0_doc):
                    elapsed = int(time.monotonic() - _t0)
                    ratio = min(token_count / 8000, 0.95)
                    # Progress between 50-88% for parallel generation
                    base = 50 if rfp_summary else 40
                    pct = base + int(38 * (_doc_done_count + ratio) / total_docs)
                    set_progress(_NS_GEN, pid, {
                        "status": "running", "step": "generating", "progress": pct,
                        "message": f"Documents en parallele ({_doc_done_count + 1}/{total_docs}): "
                                   f"{_title} — {token_count} tokens — {elapsed}s",
                    })

                try:
                    async with sem:
                        structure = await ai_service.generate_response_structure_for_document(
                            document_title=doc_title,
                            document_description=doc_desc,
                            new_rfp_content=anon_new_rfp,
                            old_rfp_content=anon_old_rfp,
                            old_response_content=anon_old_response,
                            rfp_summary=rfp_summary,
                            on_progress=_doc_progress,
                            ai_context=proj_ai_context,
                            company_name=proj_company_name,
                            client_name=proj_client_name,
                        )
                        _doc_done_count += 1
                        return (doc_id, structure)
                except Exception as exc:
                    logger.error("Structure generation error for doc '%s': %s", doc_title, exc)
                    _doc_done_count += 1
                    return (doc_id, [])

            _update("generating", 50 if rfp_summary else 40,
                    f"Generation parallele de {total_docs} document(s)...")

            results = await asyncio.gather(*[
                _gen_one_doc(idx, doc_id, doc_title, doc_desc)
                for idx, (doc_id, doc_title, doc_desc) in enumerate(resp_docs)
            ])

            all_doc_structures = [(did, struct) for did, struct in results if struct]
            failed_doc_ids = {did for did, struct in results if not struct}

            # ── Retry failed documents sequentially ──
            if failed_doc_ids:
                failed_titles = [t for (did, t, _d) in resp_docs if did in failed_doc_ids]
                logger.warning(
                    "Structure generation failed for %d/%d docs (first attempt): %s",
                    len(failed_doc_ids), total_docs, failed_titles,
                )
                _update("generating", 82,
                        f"Nouvelle tentative pour {len(failed_doc_ids)} document(s) en echec...")

                for idx, (doc_id, doc_title, doc_desc) in enumerate(resp_docs):
                    if doc_id not in failed_doc_ids:
                        continue
                    await asyncio.sleep(2)  # small delay before retry
                    # Retry without summary (use full RFP content) for better results
                    retry_structure = await ai_service.generate_response_structure_for_document(
                        document_title=doc_title,
                        document_description=doc_desc,
                        new_rfp_content=anon_new_rfp,
                        old_rfp_content=anon_old_rfp,
                        old_response_content=anon_old_response,
                        rfp_summary="",
                        ai_context=proj_ai_context,
                        company_name=proj_company_name,
                        client_name=proj_client_name,
                    )
                    if retry_structure:
                        all_doc_structures.append((doc_id, retry_structure))
                        failed_doc_ids.discard(doc_id)
                        logger.info("Retry succeeded for doc '%s'", doc_title)
                    else:
                        logger.error("Retry also failed for doc '%s'", doc_title)

            # Log AI usage for structure generation (multi-doc)
            async with task_session() as usage_db:
                await log_ai_usage_from_service(usage_db, project_id, "generate_structure", ai_service)

            if not all_doc_structures:
                set_progress(_NS_GEN, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "L'IA n'a genere aucune structure pour les documents selectionnes.",
                })
                return

            # ── Phase 3c: Batch insert chapters (single commit) ──
            _update("saving", 88, "Enregistrement des chapitres en base...")

            async with task_session() as db:
                valid_types = {t.value for t in ChapterType}

                def _build_chapters_flat(items, parent_id, resp_doc_id):
                    """Build Chapter objects with pre-assigned UUIDs (no flush needed)."""
                    nonlocal order, created_count
                    chapters = []
                    for item in items:
                        order += 1
                        created_count += 1
                        ch_id = uuid.uuid4()
                        raw_type = item.get("chapter_type", "chapter")
                        ch_type = raw_type if raw_type in valid_types else ("sub_chapter" if parent_id else "chapter")
                        chapters.append(Chapter(
                            id=ch_id,
                            project_id=project_id,
                            parent_id=parent_id,
                            response_document_id=resp_doc_id,
                            title=item.get("title", ""),
                            description=item.get("description", ""),
                            order=order,
                            chapter_type=ch_type,
                            rfp_requirement=item.get("rfp_requirement", ""),
                        ))
                        children = item.get("children", [])
                        if children:
                            chapters.extend(_build_chapters_flat(children, ch_id, resp_doc_id))
                    return chapters

                all_chapters = []
                for doc_id, structure in all_doc_structures:
                    all_chapters.extend(_build_chapters_flat(structure, None, uuid.UUID(doc_id)))

                db.add_all(all_chapters)
                await db.commit()

        else:
            # ── Legacy: single-document generation ──
            _update("generating", 40, "Generation IA de la structure en cours...")
            gen_cb = _make_stream_progress_callback(
                pid, "generating", 40, 85, "Generation structure", max_tokens=12000,
            )
            structure = await ai_service.generate_response_structure(
                new_rfp_content=anon_new_rfp,
                old_rfp_content=anon_old_rfp,
                old_response_content=anon_old_response,
                gap_analysis=gap_analysis,
                on_progress=gen_cb,
                ai_context=proj_ai_context,
                company_name=proj_company_name,
                client_name=proj_client_name,
            )

            # Log AI usage for legacy structure generation
            async with task_session() as usage_db:
                await log_ai_usage_from_service(usage_db, project_id, "generate_structure", ai_service)

            if not structure:
                set_progress(_NS_GEN, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "L'IA n'a pas retourne de JSON valide apres 2 tentatives. "
                               "Verifiez les logs serveur pour le diagnostic. Reessayez.",
                })
                return

            _update("saving", 88, "Enregistrement des chapitres en base...")

            async with task_session() as db:
                valid_types = {t.value for t in ChapterType}

                def _build_legacy_chapters(items, parent_id=None):
                    """Build Chapter objects with pre-assigned UUIDs (batch insert)."""
                    nonlocal order, created_count
                    chapters = []
                    for item in items:
                        order += 1
                        created_count += 1
                        ch_id = uuid.uuid4()
                        delta = item.get("delta", "unchanged")
                        if delta in delta_stats:
                            delta_stats[delta] += 1
                        notes = []
                        if delta and delta != "unchanged":
                            notes.append({"type": "delta", "value": delta})
                        raw_type = item.get("chapter_type", "chapter")
                        ch_type = raw_type if raw_type in valid_types else ("sub_chapter" if parent_id else "chapter")
                        chapters.append(Chapter(
                            id=ch_id,
                            project_id=project_id,
                            parent_id=parent_id,
                            title=item.get("title", ""),
                            description=item.get("description", ""),
                            order=order,
                            chapter_type=ch_type,
                            rfp_requirement=item.get("rfp_requirement", ""),
                            notes=notes,
                        ))
                        children = item.get("children", [])
                        if children:
                            chapters.extend(_build_legacy_chapters(children, parent_id=ch_id))
                    return chapters

                all_chapters = _build_legacy_chapters(structure)
                db.add_all(all_chapters)
                await db.commit()

        _update("generating", 85,
                f"Structure generee: {created_count} chapitres")

        completion_msg = ""
        if completion_docs_count > 0:
            completion_msg = f" — {completion_docs_count} document(s) a completer (Excel/PDF) a traiter dans l'onglet Livrables"

        # Build failure message for documents that failed even after retry
        failed_msg = ""
        if resp_docs and failed_doc_ids:
            failed_titles = [t for (did, t, _d) in resp_docs if did in failed_doc_ids]
            failed_msg = f" — ATTENTION: echec pour {len(failed_titles)} document(s): {', '.join(failed_titles)}"

        success_doc_count = len(resp_docs) - len(failed_doc_ids) if resp_docs else 0

        set_progress(_NS_GEN, pid, {
            "status": "completed",
            "step": "done",
            "progress": 100,
            "chapters_created": created_count,
            "delta_stats": delta_stats,
            "has_gap_analysis": gap_analysis is not None,
            "completion_docs_count": completion_docs_count,
            "message": f"{created_count} chapitres crees"
                       + (f" pour {success_doc_count}/{len(resp_docs)} document(s) redactionnels" if resp_docs else
                          f" ({delta_stats['new']} nouveaux, {delta_stats['modified']} modifies, {delta_stats['unchanged']} inchanges)")
                       + failed_msg
                       + completion_msg,
        })

    except Exception as e:
        logger.exception("Structure generation failed for project %s", project_id)
        set_progress(_NS_GEN, pid, {
            "status": "error",
            "step": "error",
            "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


@router.post("/{project_id}/prefill")
async def prefill_chapters(
    project_id: uuid.UUID,
    request: PrefillRequest,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch chapter pre-filling as a background task (returns immediately)."""
    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    pid = str(project_id)

    # Check if already running
    existing = get_or_idle(_NS_PREFILL, pid)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Pre-remplissage deja en cours")

    # Verify old response documents exist
    old_resp_count = (await db.execute(
        select(func.count()).select_from(Document)
        .where(Document.project_id == project_id)
        .where(Document.category == DocumentCategory.OLD_RESPONSE)
    )).scalar() or 0
    if old_resp_count == 0:
        raise HTTPException(status_code=400, detail="Aucun document d'ancienne reponse indexe")

    workspace_id = project.workspace_id
    chapter_ids = request.chapter_ids or []

    set_progress(_NS_PREFILL, pid, {
        "status": "running",
        "step": "starting",
        "progress": 0,
        "message": "Demarrage du pre-remplissage...",
    })

    from ..tasks.project_tasks import prefill_chapters_task
    prefill_chapters_task.apply_async(
        args=(str(project_id), str(workspace_id), chapter_ids), priority=1,
    )

    return {"success": True, "message": "Pre-remplissage lance en arriere-plan"}


@router.get("/{project_id}/prefill-status")
async def get_prefill_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of the chapter pre-filling task."""
    pid = str(project_id)
    return get_or_idle(_NS_PREFILL, pid)


async def _run_prefill(project_id: uuid.UUID, workspace_id: uuid.UUID, chapter_ids: list[str]):
    """Background task for pre-filling chapters from old response.

    Processes chapters sequentially and saves each one immediately
    so partial results are preserved if something fails.
    DB connections are released during slow AI calls to minimize pool pressure.
    """
    from ..database import task_session
    pid = str(project_id)

    def _update(step: str, progress: int, message: str, prefilled_count: int = 0):
        set_progress(_NS_PREFILL, pid, {
            "status": "running",
            "step": step,
            "progress": progress,
            "message": message,
            "prefilled_count": prefilled_count,
        })

    try:
        # Short DB session to load config + chapter list
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            # Load project settings + anonymize AI context
            proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
            proj = proj_result.scalar_one()
            raw_ai_context = proj.ai_context or ""
            proj_ai_context = await AnonymizationService.apply_existing_mappings(
                raw_ai_context, project_id, db
            ) if raw_ai_context else ""
            proj_context_mode = proj.context_mode or "rag"
            proj_company_name = getattr(proj, 'company_name', '') or ''
            proj_client_name = proj.client_name or ''

            _update("loading", 5, "Chargement des chapitres...")

            query = select(Chapter).where(Chapter.project_id == project_id)
            if chapter_ids:
                chapter_uuids = [uuid.UUID(cid) for cid in chapter_ids]
                query = query.where(Chapter.id.in_(chapter_uuids))
            result = await db.execute(query.order_by(Chapter.order))
            chapters = result.scalars().all()

            # Capture plain data we need (detach from session)
            # Anonymize chapter metadata (titles may contain client/company names)
            _anon = AnonymizationService.apply_existing_mappings
            to_prefill = []
            for ch in chapters:
                if not ch.content:
                    to_prefill.append({
                        "id": ch.id,
                        "title": ch.title,
                        "anon_title": await _anon(ch.title, project_id, db),
                        "description": ch.description,
                        "anon_description": await _anon(ch.description, project_id, db),
                        "rfp_requirement": ch.rfp_requirement,
                        "anon_rfp_requirement": await _anon(ch.rfp_requirement, project_id, db),
                    })
            total_chapters = len(chapters)
        # DB connection released here

        total = len(to_prefill)

        if total == 0:
            set_progress(_NS_PREFILL, pid, {
                "status": "completed",
                "step": "done",
                "progress": 100,
                "prefilled_count": 0,
                "message": "Aucun chapitre vide a pre-remplir",
            })
            return

        _update("loading", 10,
                f"{total} chapitre(s) vide(s) a pre-remplir sur {total_chapters} total")

        # Full context mode: load ALL old response content once (already anonymized)
        full_old_response = ""
        if proj_context_mode == "full":
            async with task_session() as db:
                full_old_response = await _get_full_text_anonymized_by_category(
                    db, project_id, DocumentCategory.OLD_RESPONSE
                )
            if not full_old_response.strip():
                set_progress(_NS_PREFILL, pid, {
                    "status": "completed", "step": "done", "progress": 100,
                    "prefilled_count": 0,
                    "message": "Aucun contenu d'ancienne reponse trouve",
                })
                return

        prefilled = 0
        skipped = 0

        for idx, ch_data in enumerate(to_prefill):
            try:
                progress = 10 + int(85 * (idx + 1) / total)
                _update("prefilling", progress,
                        f"Chapitre {idx + 1}/{total}: {ch_data['title'][:60]}...",
                        prefilled_count=prefilled)

                old_response_chunks = []
                if proj_context_mode == "full":
                    # Full context: reuse pre-loaded content (already anonymized)
                    anon_content = full_old_response
                else:
                    # RAG: vector search for relevant chunks
                    search_query = f"{ch_data['title']} {ch_data['description']}"
                    old_response_chunks = VectorService.search(
                        str(project_id), search_query, top_k=5, category_filter="old_response"
                    )

                    if not old_response_chunks:
                        skipped += 1
                        continue

                    old_content = "\n\n".join([c["content"] for c in old_response_chunks])

                    # Short DB session for anonymization
                    async with task_session() as db:
                        anon_content = await AnonymizationService.anonymize_text(old_content, project_id, db)

                # AI call (no DB connection held) — all fields anonymized
                content = await ai_service.generate_chapter_content(
                    chapter_title=ch_data["anon_title"],
                    chapter_description=ch_data["anon_description"],
                    rfp_requirement=ch_data["anon_rfp_requirement"],
                    old_response_content=anon_content,
                    ai_context=proj_ai_context,
                    company_name=proj_company_name,
                    client_name=proj_client_name,
                )

                # Log AI usage for prefill chapter generation
                async with task_session() as usage_db:
                    await log_ai_usage_from_service(usage_db, project_id, "prefill_chapter", ai_service)

                # Short DB session for deanonymization + save
                refs = [
                    {"document": c["document_name"], "page": c["page_number"], "score": c["score"]}
                    for c in old_response_chunks[:3]
                ] if old_response_chunks else []
                async with task_session() as db:
                    deanon = await AnonymizationService.deanonymize_text(content, project_id, db)
                    chap_result = await db.execute(select(Chapter).where(Chapter.id == ch_data["id"]))
                    chapter = chap_result.scalar_one()
                    chapter.content = deanon
                    chapter.is_prefilled = True
                    chapter.status = ChapterStatus.IN_PROGRESS
                    chapter.source_references = refs
                    await db.commit()
                prefilled += 1

            except Exception as ch_err:
                logger.warning("Prefill failed for chapter %s: %s", ch_data["title"], str(ch_err)[:200])
                skipped += 1
                continue

        set_progress(_NS_PREFILL, pid, {
            "status": "completed",
            "step": "done",
            "progress": 100,
            "prefilled_count": prefilled,
            "message": f"{prefilled} chapitre(s) pre-rempli(s)"
                       + (f" ({skipped} sans contenu pertinent)" if skipped else ""),
        })

    except Exception as e:
        logger.exception("Prefill failed for project %s", project_id)
        set_progress(_NS_PREFILL, pid, {
            "status": "error",
            "step": "error",
            "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


# ── Response Documents (Deliverables) ──


@router.post("/{project_id}/detect-deliverables")
async def detect_deliverables(
    project_id: uuid.UUID,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Detect expected deliverables from the RFP (background task)."""
    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    doc_count = (await db.execute(
        select(func.count()).select_from(Document)
        .where(Document.project_id == project_id)
        .where(Document.category == DocumentCategory.NEW_RFP)
    )).scalar() or 0
    if doc_count == 0:
        raise HTTPException(status_code=400, detail="Aucun document de nouvel AO indexe")

    pid = str(project_id)
    existing = get_or_idle(_NS_DETECT, pid)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Detection deja en cours")

    set_progress(_NS_DETECT, pid, {
        "status": "running", "step": "starting", "progress": 0,
        "message": "Demarrage de la detection des livrables...",
    })

    from ..tasks.project_tasks import detect_deliverables_task
    detect_deliverables_task.apply_async(
        args=(str(project_id), str(project.workspace_id)), priority=5,
    )
    return {"success": True, "message": "Detection lancee en arriere-plan"}


@router.get("/{project_id}/detect-deliverables-status")
async def get_detect_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of deliverable detection."""
    pid = str(project_id)
    return get_or_idle(_NS_DETECT, pid)


async def _run_detect_deliverables(project_id: uuid.UUID, workspace_id: uuid.UUID):
    """Background task: analyze RFP to detect expected deliverables.

    DB connections are released during the slow AI call to minimize pool pressure.
    """
    from ..database import task_session
    pid = str(project_id)

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_DETECT, pid, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
        })

    try:
        # ── Phase 1: Load anonymized chunks (short DB session) ──
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            _update("loading", 10, "Chargement du contenu de l'AO...")
            anon_new_rfp = await _get_all_chunks_anonymized_by_category(
                db, project_id, DocumentCategory.NEW_RFP
            )
            if not anon_new_rfp:
                set_progress(_NS_DETECT, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Aucun document de nouvel AO indexe",
                })
                return

            anon_old_response = await _get_all_chunks_anonymized_by_category(
                db, project_id, DocumentCategory.OLD_RESPONSE
            )
        # DB released

        # ── Phase 2: AI detection (NO DB connection held) ──
        _update("analyzing", 20, "Analyse IA des livrables attendus...")
        import time
        t0 = time.monotonic()

        async def _on_detect_progress(token_count: int, char_count: int):
            elapsed = int(time.monotonic() - t0)
            ratio = min(token_count / 8000, 0.95)
            pct = 20 + int(60 * ratio)
            set_progress(_NS_DETECT, pid, {
                "status": "running", "step": "analyzing", "progress": pct,
                "message": f"Analyse IA — {token_count} tokens ({char_count:,} car.) — {elapsed}s",
            })

        deliverables = await ai_service.detect_deliverables(
            new_rfp_content=anon_new_rfp,
            old_response_content=anon_old_response,
            on_progress=_on_detect_progress,
        )

        if not deliverables:
            set_progress(_NS_DETECT, pid, {
                "status": "error", "step": "error", "progress": 0,
                "message": "L'IA n'a pas detecte de livrables. Reessayez.",
            })
            return

        # ── Phase 3: Save results (short DB session) ──
        _update("saving", 85, f"{len(deliverables)} livrable(s) detecte(s), enregistrement...")

        async with task_session() as db:
            # Remove old response documents (re-detection replaces previous)
            old_docs = await db.execute(
                select(ResponseDocument).where(ResponseDocument.project_id == project_id)
            )
            for old_doc in old_docs.scalars().all():
                await db.delete(old_doc)

            # Create new response documents
            for idx, d in enumerate(deliverables):
                fmt = d.get("expected_format", "docx")
                try:
                    doc_format = DocumentFormat(fmt)
                except ValueError:
                    doc_format = DocumentFormat.OTHER

                # Determine content_type from AI response or infer from format
                raw_content_type = d.get("content_type", "")
                title_lower = d.get("title", "").lower()

                # Keywords that ALWAYS indicate a completion document,
                # regardless of what the AI classified
                completion_keywords = [
                    "bpu", "bordereau", "dqe", "dpgf",
                    "dc1", "dc2", "dc3", "dc4",
                    "attri", "noti", "acte d'engagement",
                    "formulaire", "cerfa",
                    "a completer", "à compléter",
                    "a remplir", "à remplir",
                    "cadre de réponse", "cadre de reponse",
                    "grille", "questionnaire",
                    "annexe conformit", "annexe rgpd",
                ]
                is_clearly_completion = (
                    doc_format == DocumentFormat.XLSX
                    or any(kw in title_lower for kw in completion_keywords)
                )

                if is_clearly_completion:
                    ct = ContentType.COMPLETION
                    if raw_content_type == "redaction":
                        logger.warning(
                            "AI classified '%s' (format=%s) as redaction but "
                            "heuristic overrides to COMPLETION",
                            d.get("title"), doc_format,
                        )
                elif raw_content_type == "completion":
                    ct = ContentType.COMPLETION
                elif raw_content_type == "redaction":
                    ct = ContentType.REDACTION
                else:
                    ct = ContentType.REDACTION

                rd = ResponseDocument(
                    project_id=project_id,
                    title=d.get("title", f"Document {idx + 1}"),
                    description=d.get("description", ""),
                    expected_format=doc_format,
                    content_type=ct,
                    is_selected=d.get("suggested", True),
                    order=idx + 1,
                    rfp_source=d.get("rfp_source", ""),
                )
                db.add(rd)

            await db.commit()

        set_progress(_NS_DETECT, pid, {
            "status": "completed", "step": "done", "progress": 100,
            "deliverables_count": len(deliverables),
            "message": f"{len(deliverables)} livrable(s) detecte(s) dans l'AO",
        })

    except Exception as e:
        logger.exception("Deliverable detection failed for project %s", project_id)
        set_progress(_NS_DETECT, pid, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


@router.get("/{project_id}/response-documents", response_model=list[ResponseDocumentOut])
async def list_response_documents(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all detected response documents for a project."""
    result = await db.execute(
        select(ResponseDocument)
        .where(ResponseDocument.project_id == project_id)
        .order_by(ResponseDocument.order)
    )
    docs = result.scalars().all()

    out = []
    for d in docs:
        ch_count = (await db.execute(
            select(func.count()).select_from(Chapter)
            .where(Chapter.response_document_id == d.id)
        )).scalar() or 0
        out.append(ResponseDocumentOut(
            id=str(d.id),
            project_id=str(d.project_id),
            title=d.title,
            description=d.description,
            expected_format=d.expected_format.value,
            content_type=d.content_type.value if hasattr(d.content_type, 'value') else (d.content_type or "REDACTION").lower(),
            is_selected=d.is_selected,
            order=d.order,
            rfp_source=d.rfp_source,
            fill_content=d.fill_content or "",
            fill_status=d.fill_status or "pending",
            source_document_ids=[str(x) for x in (d.source_document_ids or [])],
            source_categories=d.source_categories or [],
            include_generated_content=d.include_generated_content or False,
            custom_notes=d.custom_notes or "",
            created_at=d.created_at,
            updated_at=d.updated_at,
            chapter_count=ch_count,
        ))
    return out


@router.put("/{project_id}/response-documents/{doc_id}", response_model=ResponseDocumentOut)
async def update_response_document(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    request: ResponseDocumentUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update a response document (toggle selection, rename, etc.)."""
    result = await db.execute(
        select(ResponseDocument)
        .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")

    for field in ["title", "description", "expected_format", "content_type", "is_selected", "order",
                   "source_document_ids", "source_categories", "include_generated_content", "custom_notes"]:
        value = getattr(request, field, None)
        if value is not None:
            # Convert string values to proper enums for SQLAlchemy
            if field == "content_type" and isinstance(value, str):
                try:
                    value = ContentType(value)
                except ValueError:
                    value = ContentType.REDACTION
            elif field == "expected_format" and isinstance(value, str):
                try:
                    value = DocumentFormat(value)
                except ValueError:
                    value = DocumentFormat.OTHER
            setattr(doc, field, value)

    await db.commit()
    await db.refresh(doc)

    ch_count = (await db.execute(
        select(func.count()).select_from(Chapter)
        .where(Chapter.response_document_id == doc.id)
    )).scalar() or 0

    return ResponseDocumentOut(
        id=str(doc.id), project_id=str(doc.project_id),
        title=doc.title, description=doc.description,
        expected_format=doc.expected_format.value,
        content_type=doc.content_type.value if hasattr(doc.content_type, 'value') else (doc.content_type or "REDACTION").lower(),
        is_selected=doc.is_selected, order=doc.order,
        rfp_source=doc.rfp_source,
        fill_content=doc.fill_content or "",
        fill_status=doc.fill_status or "pending",
        source_document_ids=[str(x) for x in (doc.source_document_ids or [])],
        source_categories=doc.source_categories or [],
        include_generated_content=doc.include_generated_content or False,
        custom_notes=doc.custom_notes or "",
        created_at=doc.created_at, updated_at=doc.updated_at,
        chapter_count=ch_count,
    )


@router.post("/{project_id}/response-documents/{doc_id}/reset-fill")
async def reset_fill_content(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Reset fill_content and fill_status for a response document so it can be regenerated."""
    result = await db.execute(
        select(ResponseDocument)
        .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")

    doc.fill_content = ""
    doc.fill_status = "pending"
    await db.commit()
    await db.refresh(doc)

    ch_count = (await db.execute(
        select(func.count()).select_from(Chapter)
        .where(Chapter.response_document_id == doc.id)
    )).scalar() or 0

    return ResponseDocumentOut(
        id=str(doc.id), project_id=str(doc.project_id),
        title=doc.title, description=doc.description,
        expected_format=doc.expected_format.value,
        content_type=doc.content_type.value if hasattr(doc.content_type, 'value') else (doc.content_type or "REDACTION").lower(),
        is_selected=doc.is_selected, order=doc.order,
        rfp_source=doc.rfp_source,
        fill_content="",
        fill_status="pending",
        source_document_ids=[str(x) for x in (doc.source_document_ids or [])],
        source_categories=doc.source_categories or [],
        include_generated_content=doc.include_generated_content or False,
        custom_notes=doc.custom_notes or "",
        created_at=doc.created_at, updated_at=doc.updated_at,
        chapter_count=ch_count,
    )


@router.post("/{project_id}/response-documents/confirm-selection")
async def confirm_document_selection(
    project_id: uuid.UUID,
    request: BulkUpdateSelectionRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Bulk-update which response documents are selected."""
    for item in request.selections:
        doc_id = uuid.UUID(item["id"])
        result = await db.execute(
            select(ResponseDocument)
            .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
        )
        doc = result.scalar_one_or_none()
        if doc:
            doc.is_selected = item.get("is_selected", True)

    await db.commit()
    selected_count = (await db.execute(
        select(func.count()).select_from(ResponseDocument)
        .where(ResponseDocument.project_id == project_id)
        .where(ResponseDocument.is_selected == True)
    )).scalar() or 0

    return {"success": True, "selected_count": selected_count}


# ── Auto-fill completion documents (Excel/PDF) ──


@router.post("/{project_id}/fill-deliverables")
async def fill_deliverables(
    project_id: uuid.UUID,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch auto-fill for completion-type deliverables (background task)."""
    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Check for completion-type documents
    comp_result = await db.execute(
        select(ResponseDocument)
        .where(ResponseDocument.project_id == project_id)
        .where(ResponseDocument.is_selected == True)
        .where(ResponseDocument.content_type == ContentType.COMPLETION)
    )
    comp_docs = comp_result.scalars().all()
    if not comp_docs:
        raise HTTPException(status_code=400, detail="Aucun document à compléter détecté")

    pid = str(project_id)
    existing = get_or_idle(_NS_FILL, pid)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Auto-remplissage déjà en cours")

    set_progress(_NS_FILL, pid, {
        "status": "running", "step": "starting", "progress": 0,
        "message": "Démarrage de l'auto-remplissage...",
    })

    from ..tasks.project_tasks import fill_deliverables_task
    fill_deliverables_task.apply_async(
        args=(str(project_id), str(project.workspace_id)), priority=5,
    )
    return {"success": True, "message": "Auto-remplissage lancé en arrière-plan"}


@router.get("/{project_id}/fill-deliverables-status")
async def get_fill_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of auto-fill for completion documents."""
    pid = str(project_id)
    return get_or_idle(_NS_FILL, pid)


async def _run_fill_deliverables(project_id: uuid.UUID, workspace_id: uuid.UUID):
    """Background task: auto-fill completion-type documents (BPU, DQE, forms, etc.).

    DB connections are released during the slow AI calls to minimize pool pressure.
    """
    from ..database import task_session
    pid = str(project_id)

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_FILL, pid, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
        })

    try:
        # ── Phase 1: Load all data (short DB session) ──
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            # Load project AI context and anonymize it
            proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
            proj = proj_result.scalar_one()
            raw_ai_context = proj.ai_context or ""
            proj_ai_context = await AnonymizationService.apply_existing_mappings(
                raw_ai_context, project_id, db
            ) if raw_ai_context else ""
            proj_company_name = getattr(proj, 'company_name', '') or ''
            proj_client_name = proj.client_name or ''

            _update("loading", 5, "Chargement des documents à compléter...")

            result = await db.execute(
                select(ResponseDocument)
                .where(ResponseDocument.project_id == project_id)
                .where(ResponseDocument.is_selected == True)
                .where(ResponseDocument.content_type == ContentType.COMPLETION)
                .order_by(ResponseDocument.order)
            )
            comp_docs = result.scalars().all()
            total = len(comp_docs)

            if total == 0:
                set_progress(_NS_FILL, pid, {
                    "status": "completed", "step": "done", "progress": 100,
                    "filled_count": 0,
                    "message": "Aucun document à compléter",
                })
                return

            # Capture plain data (detach from session) — anonymize titles/descriptions
            _anon = AnonymizationService.apply_existing_mappings
            docs_data = []
            for doc in comp_docs:
                docs_data.append({
                    "id": doc.id,
                    "title": await _anon(doc.title or "", project_id, db),
                    "description": await _anon(doc.description or "", project_id, db),
                    "expected_format": doc.expected_format.value,
                    "source_document_ids": doc.source_document_ids or [],
                    "source_categories": doc.source_categories or [],
                    "include_generated_content": doc.include_generated_content or False,
                    "custom_notes": doc.custom_notes or "",
                })

            _update("loading", 10, f"{total} document(s) à compléter...")

            # Default context: load all NEW_RFP + OLD_RESPONSE for docs without specific source selection
            anon_new_rfp, anon_old_response = await asyncio.gather(
                _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.NEW_RFP),
                _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.OLD_RESPONSE),
            )

            chapters_result = await db.execute(
                select(Chapter)
                .where(Chapter.project_id == project_id)
                .where(Chapter.content != "")
                .order_by(Chapter.order)
            )
            existing_chapters = chapters_result.scalars().all()
            # Anonymize chapter context before sending to AI
            chapter_parts_list = []
            for ch in existing_chapters[:10]:
                anon_ch_t = await _anon(ch.title, project_id, db)
                anon_ch_c = await _anon(ch.content[:2000], project_id, db)
                chapter_parts_list.append(f"## {anon_ch_t}\n{anon_ch_c}")
            chapter_context = "\n\n".join(chapter_parts_list)

            # Pre-load generated chapters context if any doc needs it
            generated_context = ""
            needs_generated = any(d["include_generated_content"] for d in docs_data)
            if needs_generated:
                generated_context = await _get_generated_chapters_context(db, project_id)

            # Pre-load custom context for docs that have specific source selections
            per_doc_context = {}
            for doc_data in docs_data:
                src_ids = doc_data["source_document_ids"]
                src_cats = doc_data["source_categories"]
                inc_gen = doc_data["include_generated_content"]
                has_custom_sources = bool(src_ids or src_cats or inc_gen)
                if has_custom_sources:
                    parts = []
                    if src_cats:
                        parts.append(await _get_chunks_anonymized_by_categories(
                            db, project_id, src_cats
                        ))
                    if src_ids:
                        parts.append(await _get_chunks_anonymized_by_document_ids(
                            db, project_id, src_ids
                        ))
                    if inc_gen and generated_context:
                        parts.append(generated_context)
                    per_doc_context[doc_data["id"]] = "\n\n".join(p for p in parts if p)
        # DB released

        import time
        # Build default combined context (for docs without specific source selection)
        default_combined_context = anon_old_response
        if chapter_context:
            default_combined_context += "\n\n--- CONTENU DÉJÀ RÉDIGÉ ---\n\n" + chapter_context

        # ── Phase 2: Parallel AI generation (NO DB connection held) ──
        sem = asyncio.Semaphore(10)
        _fill_done = 0

        async def _fill_one_doc(idx, doc_data):
            nonlocal _fill_done
            t0_doc = time.monotonic()

            async def _on_fill_progress(token_count: int, char_count: int,
                                        _t0=t0_doc, _title=doc_data["title"]):
                elapsed = int(time.monotonic() - _t0)
                ratio = min(token_count / 8000, 0.95)
                pct = 10 + int(85 * (_fill_done + ratio) / total)
                set_progress(_NS_FILL, pid, {
                    "status": "running", "step": "filling", "progress": pct,
                    "message": f"Documents en parallele ({_fill_done + 1}/{total}): "
                               f"{_title} — {token_count} tokens — {elapsed}s",
                })

            async with sem:
                # Use per-doc source context if user selected specific sources,
                # otherwise fall back to the default (all OLD_RESPONSE + chapters)
                doc_id = doc_data["id"]
                if doc_id in per_doc_context:
                    doc_context = per_doc_context[doc_id]
                else:
                    doc_context = default_combined_context

                fill_content = await ai_service.generate_fill_content(
                    document_title=doc_data["title"],
                    document_description=doc_data["description"],
                    expected_format=doc_data["expected_format"],
                    new_rfp_content=anon_new_rfp,
                    old_response_content=doc_context,
                    on_progress=_on_fill_progress,
                    ai_context=proj_ai_context,
                    company_name=proj_company_name,
                    client_name=proj_client_name,
                    custom_notes=doc_data.get("custom_notes", ""),
                )
                _fill_done += 1
                return (doc_data["id"], fill_content)

        _update("filling", 10, f"Remplissage parallele de {total} document(s)...")

        results = await asyncio.gather(*[
            _fill_one_doc(idx, doc_data) for idx, doc_data in enumerate(docs_data)
        ])

        # Log AI usage for fill content generation
        async with task_session() as usage_db:
            await log_ai_usage_from_service(usage_db, project_id, "fill_content", ai_service)

        # ── Phase 3: Deanonymize + save (short DB session) ──
        _update("saving", 96, "Deanonymisation et enregistrement...")
        filled_count = 0
        async with task_session() as db:
            for doc_id, raw_content in results:
                deanon = await AnonymizationService.deanonymize_text(raw_content, project_id, db)
                doc_result = await db.execute(
                    select(ResponseDocument).where(ResponseDocument.id == doc_id)
                )
                doc = doc_result.scalar_one()
                doc.fill_content = deanon
                doc.fill_status = "completed"
                filled_count += 1
            await db.commit()

        set_progress(_NS_FILL, pid, {
            "status": "completed", "step": "done", "progress": 100,
            "filled_count": filled_count,
            "message": f"{filled_count} document(s) à compléter traité(s)",
        })

    except Exception as e:
        logger.exception("Fill deliverables failed for project %s", project_id)
        set_progress(_NS_FILL, pid, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


@router.get("/{project_id}/compliance-analysis")
async def get_compliance_analysis(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get the latest persisted compliance analysis for a project."""
    result = await db.execute(
        select(ComplianceResult)
        .where(ComplianceResult.project_id == project_id)
        .order_by(ComplianceResult.created_at.desc())
        .limit(1)
    )
    cr = result.scalar_one_or_none()
    if not cr:
        return {"analysis": None}

    return {
        "analysis": {
            "id": str(cr.id),
            "score": cr.score,
            "summary": cr.summary,
            "covered_requirements": cr.covered_requirements or [],
            "missing_elements": cr.missing_elements or [],
            "recommendations": cr.recommendations or [],
            "created_at": cr.created_at.isoformat() if cr.created_at else None,
        }
    }


@router.post("/{project_id}/compliance-analysis")
async def analyze_compliance(
    project_id: uuid.UUID,
    request: dict = {},

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch compliance analysis as a background task (returns immediately).

    Optional body parameter:
        target_scope: "all" (default) | "memoire_only" | "documents_only"
            - "all": analyze both mémoire technique chapters AND uploaded response documents
            - "memoire_only": analyze only the mémoire technique (generated chapters)
            - "documents_only": analyze only uploaded 'Notre réponse' documents
    """
    pid = str(project_id)
    target_scope = request.get("target_scope", "all") if isinstance(request, dict) else "all"
    if target_scope not in ("all", "memoire_only", "documents_only"):
        target_scope = "all"

    existing = get_or_idle(_NS_COMPLIANCE, pid)
    if existing and existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Analyse de conformite deja en cours")

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Quick pre-checks before launching background task
    # Check for chapters (mémoire technique) OR uploaded NEW_RESPONSE documents
    chapters_result = await db.execute(
        select(func.count()).select_from(Chapter).where(
            Chapter.project_id == project_id,
            Chapter.content != None,
            Chapter.content != "",
        )
    )
    has_chapters = (chapters_result.scalar() or 0) > 0

    nr_result = await db.execute(
        select(Document.id).where(
            Document.project_id == project_id,
            Document.category == DocumentCategory.NEW_RESPONSE,
            Document.processing_status == ProcessingStatus.COMPLETED,
        ).limit(1)
    )
    has_new_response_docs = nr_result.scalar_one_or_none() is not None

    # Validate scope against available content
    if target_scope == "memoire_only" and not has_chapters:
        raise HTTPException(
            status_code=400,
            detail="Aucun chapitre rédigé dans le mémoire technique.",
        )
    if target_scope == "documents_only" and not has_new_response_docs:
        raise HTTPException(
            status_code=400,
            detail="Aucun document 'Notre réponse' chargé.",
        )
    if target_scope == "all" and not has_chapters and not has_new_response_docs:
        raise HTTPException(
            status_code=400,
            detail="Aucun contenu à analyser. Rédigez les chapitres du mémoire technique ou chargez des documents 'Notre réponse'.",
        )

    set_progress(_NS_COMPLIANCE, pid, {
        "status": "running", "step": "starting", "progress": 0,
        "message": "Demarrage de l'analyse de conformite...",
    })

    from ..tasks.project_tasks import compliance_analysis_task
    compliance_analysis_task.apply_async(
        args=(str(project_id), str(project.workspace_id), target_scope), priority=3,
    )

    return {"success": True, "message": "Analyse de conformite lancee en arriere-plan"}


@router.get("/{project_id}/compliance-analysis-status")
async def get_compliance_analysis_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of compliance analysis."""
    pid = str(project_id)
    return get_or_idle(_NS_COMPLIANCE, pid)


async def _run_compliance_analysis(project_id: uuid.UUID, workspace_id: uuid.UUID, target_scope: str = "all"):
    """Background task for compliance analysis.

    DB connections are released during the slow AI call to minimize pool pressure.

    Args:
        target_scope: "all" | "memoire_only" | "documents_only"
            Controls which response content to analyze.
    """
    from ..database import task_session
    pid = str(project_id)
    include_chapters = target_scope in ("all", "memoire_only")
    include_uploaded = target_scope in ("all", "documents_only")

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_COMPLIANCE, pid, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
        })

    try:
        # ── Phase 1: Load data + anonymize (short DB session) ──
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            anon_chapters = ""
            anon_uploaded_response = ""

            # ── 1a. Load the mémoire technique (generated chapters) ──
            if include_chapters:
                _update("loading", 5, "Chargement du memoire technique (chapitres)...")
                chapters_result = await db.execute(
                    select(Chapter).where(Chapter.project_id == project_id).order_by(Chapter.order)
                )
                chapters = chapters_result.scalars().all()
                chapters_with_content = [c for c in chapters if (c.content or "").strip()]

                chapter_parts = []
                if chapters_with_content:
                    chapter_parts.append("\n\n=== DOCUMENT: Memoire Technique (redige avec l'outil) ===\n")
                    for c in chapters_with_content:
                        text = (c.anonymized_content or c.content or "").strip()
                        anon_ch_title = await AnonymizationService.apply_existing_mappings(
                            c.title, project_id, db
                        )
                        chapter_parts.append(f"\n--- {anon_ch_title} ---\n")
                        chapter_parts.append(text)
                anon_chapters = "\n\n".join(chapter_parts)

            # ── 1b. Load uploaded NEW_RESPONSE documents (if any) ──
            if include_uploaded:
                _update("loading", 10, "Chargement des documents reponse uploades...")
                anon_uploaded_response = await _get_all_chunks_anonymized_by_category(
                    db, project_id, DocumentCategory.NEW_RESPONSE
                )

            # Combine based on scope
            response_parts = []
            if anon_chapters.strip():
                response_parts.append(anon_chapters)
            if anon_uploaded_response.strip():
                response_parts.append(anon_uploaded_response)
            anon_response = "\n\n".join(response_parts)

            scope_label = {
                "memoire_only": "le memoire technique",
                "documents_only": "les documents reponse uploades",
                "all": "le memoire technique et les documents reponse",
            }.get(target_scope, "la reponse")

            if not anon_response.strip():
                set_progress(_NS_COMPLIANCE, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": f"Aucun contenu a analyser dans {scope_label}.",
                })
                return

            # ── 1c. Load ALL RFP content (CCAP, CCTP, RC, etc.) ──
            _update("searching", 15, "Chargement integral du cahier des charges (CCAP, CCTP, RC...)...")
            anon_rfp = await _get_all_chunks_anonymized_by_category(
                db, project_id, DocumentCategory.NEW_RFP
            )

            if not anon_rfp.strip():
                set_progress(_NS_COMPLIANCE, pid, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Aucun document d'appel d'offres indexe",
                })
                return

            # ── 1d. Load image analyses from documents (schemas, illustrations, etc.) ──
            _update("loading_images", 20, "Chargement des analyses d'images des documents...")
            rfp_images_text = await _get_image_analyses_by_category(
                db, project_id, DocumentCategory.NEW_RFP
            )
            if rfp_images_text.strip():
                anon_rfp += "\n\n--- CONTENU EXTRAIT DES IMAGES/SCHEMAS DES DOCUMENTS AO ---\n" + rfp_images_text

            response_images_text = ""
            if include_uploaded:
                response_images_text = await _get_image_analyses_by_category(
                    db, project_id, DocumentCategory.NEW_RESPONSE
                )
            if response_images_text.strip():
                anon_response += "\n\n--- CONTENU EXTRAIT DES IMAGES/SCHEMAS DES DOCUMENTS DE REPONSE ---\n" + response_images_text

            _update("anonymizing", 25, "Preparation de l'analyse...")
        # DB released

        # ── Phase 2: AI analysis (NO DB connection held) ──
        _update("analyzing", 40, "Analyse IA de la conformite en cours...")

        async def _compliance_progress_cb(tokens: int, chars: int):
            # Map streaming progress to 40-70% range
            pct = min(70, 40 + int(tokens / 20))
            _update("analyzing", pct, f"Analyse IA en cours... ({tokens} tokens)")

        analysis = await ai_service.analyze_compliance(
            anon_response, anon_rfp, on_progress=_compliance_progress_cb,
            target_scope=target_scope,
        )

        # Log AI usage for compliance analysis
        async with task_session() as usage_db:
            await log_ai_usage_from_service(usage_db, project_id, "compliance_analysis", ai_service)

        # ── Phase 3: Deanonymize + save (short DB session) ──
        _update("deanonymizing", 75, "Deanonymisation des resultats...")
        async with task_session() as db:
            for req in analysis.get("covered_requirements", []):
                for key in ("requirement", "comment", "source_rfp", "source_response"):
                    if key in req and req[key]:
                        req[key] = await AnonymizationService.deanonymize_text(req[key], project_id, db)
            for elem in analysis.get("missing_elements", []):
                for key in ("requirement", "description", "source_rfp"):
                    if key in elem and elem[key]:
                        elem[key] = await AnonymizationService.deanonymize_text(elem[key], project_id, db)
            for i, rec in enumerate(analysis.get("recommendations", [])):
                if rec:
                    analysis["recommendations"][i] = await AnonymizationService.deanonymize_text(rec, project_id, db)
            if analysis.get("summary"):
                analysis["summary"] = await AnonymizationService.deanonymize_text(analysis["summary"], project_id, db)

            _update("saving", 90, "Enregistrement des resultats...")
            cr = ComplianceResult(
                project_id=project_id,
                score=analysis.get("score", 0),
                summary=analysis.get("summary", ""),
                covered_requirements=analysis.get("covered_requirements", []),
                missing_elements=analysis.get("missing_elements", []),
                recommendations=analysis.get("recommendations", []),
            )
            db.add(cr)
            await db.commit()

        set_progress(_NS_COMPLIANCE, pid, {
            "status": "completed", "step": "done", "progress": 100,
            "message": "Analyse de conformite terminee",
        })

    except Exception as e:
        logger.exception("Compliance analysis failed for project %s", project_id)
        set_progress(_NS_COMPLIANCE, pid, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


async def _find_best_chapter(
    db: AsyncSession, project_id: uuid.UUID, search_text: str,
) -> tuple:
    """Find the best matching chapter for a given compliance gap text.

    Uses a scoring approach: compares the gap text against each chapter's
    title, description, and rfp_requirement to find the most relevant one.
    Returns (chapter, score) or (None, 0) if no chapters exist.
    """
    result = await db.execute(
        select(Chapter)
        .where(Chapter.project_id == project_id)
        .order_by(Chapter.order)
    )
    all_chapters = result.scalars().all()
    if not all_chapters:
        return None, 0

    search_lower = search_text.lower()
    search_words = set(w for w in search_lower.split() if len(w) > 3)

    best_chapter = None
    best_score = -1

    for chapter in all_chapters:
        score = 0
        ch_title = (chapter.title or "").lower()
        ch_desc = (chapter.description or "").lower()
        ch_rfp = (chapter.rfp_requirement or "").lower()
        ch_combined = f"{ch_title} {ch_desc} {ch_rfp}"

        # Word overlap scoring
        ch_words = set(w for w in ch_combined.split() if len(w) > 3)
        overlap = search_words & ch_words
        score += len(overlap) * 3

        # Substring matching on title (strong signal)
        for word in search_words:
            if word in ch_title:
                score += 5

        # Prefer chapters that already have content (more context for generation)
        if (chapter.content or "").strip():
            score += 1

        # Prefer leaf chapters (sub-chapters) over root chapters
        if chapter.parent_id:
            score += 1

        if score > best_score:
            best_score = score
            best_chapter = chapter

    return best_chapter, best_score


@router.post("/{project_id}/compliance-analysis/generate-recommendation")
async def generate_recommendation_content(
    project_id: uuid.UUID,
    request: dict,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch recommendation content generation as a background task (returns immediately).

    Body: {
        "recommendation": "the recommendation or missing element text",
        "task_id": "unique id from frontend to track this specific generation",
        "chapter_id": "optional override – if omitted the best chapter is auto-detected",
        "missing_description": "optional description of what is missing (for missing elements)",
        "inject": true/false (default true) – whether to inject into the chapter
    }

    Returns immediately with {task_id}. Poll status via GET .../generate-recommendation-status/{task_id}.
    """
    recommendation = request.get("recommendation", "").strip()
    task_id = request.get("task_id", str(uuid.uuid4()))
    chapter_id = request.get("chapter_id")
    missing_description = request.get("missing_description", "").strip()
    inject = request.get("inject", True)

    if not recommendation:
        raise HTTPException(status_code=400, detail="Recommendation manquante")

    # Don't relaunch if already running
    existing = get_or_idle(_NS_REC, task_id)
    if existing and existing.get("status") in ("running", "queued"):
        return {"task_id": task_id}

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Quick config check
    await _get_ai_service(project.workspace_id, db)

    set_progress(_NS_REC, task_id, {
        "status": "queued", "step": "queued", "progress": 0,
        "message": "En file d'attente...",
        "chapter_id": None, "chapter_title": None, "content": None,
    })

    from ..tasks.project_tasks import generate_recommendation_task
    generate_recommendation_task.apply_async(
        args=(
            task_id, str(project_id), str(project.workspace_id),
            recommendation, missing_description, chapter_id, inject,
        ),
        priority=5,
    )

    return {"task_id": task_id}


@router.get("/{project_id}/compliance-analysis/generate-recommendation-status/{task_id}")
async def get_rec_gen_status(
    project_id: uuid.UUID,
    task_id: str,
    current_user: User = Depends(get_current_user),
):
    """Poll the progress of a recommendation content generation task."""
    result = get_or_idle(_NS_REC, task_id)
    # Ensure frontend-expected keys are always present
    result.setdefault("chapter_id", None)
    result.setdefault("chapter_title", None)
    result.setdefault("content", None)
    return result


async def _run_rec_generation(
    task_id: str, project_id: uuid.UUID, workspace_id: uuid.UUID,
    recommendation: str, missing_description: str,
    chapter_id_override: str | None, inject: bool,
):
    """Background task for recommendation/missing-element content generation.

    Concurrency is controlled by Celery worker --concurrency setting.
    DB connections are released during the slow AI call.
    """
    from ..database import task_session
    from ..tasks.chapter_tasks import _load_project_images
    from ..models.document import DocumentImage

    def _update(step: str, progress: int, message: str, **extra):
        set_progress(_NS_REC, task_id, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
            "chapter_id": extra.get("chapter_id"),
            "chapter_title": extra.get("chapter_title"),
            "content": extra.get("content"),
        })

    try:
        _update("starting", 5, "Demarrage...")

        # ── Phase 1: Load data + anonymize (short DB session) ──
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            project_result = await db.execute(
                select(RFPProject).where(RFPProject.id == project_id)
            )
            project = project_result.scalar_one()
            raw_ai_context = project.ai_context or ""
            ai_context = await AnonymizationService.apply_existing_mappings(
                raw_ai_context, project_id, db
            ) if raw_ai_context else ""
            proj_company_name = getattr(project, 'company_name', '') or ''
            proj_client_name = project.client_name or ''

            _update("searching", 10, "Recherche de contexte...")

            # Get RFP context via vector search — use both recommendation and description
            search_query = f"{recommendation} {missing_description}" if missing_description else recommendation
            rfp_chunks = VectorService.search(
                str(project_id), search_query, top_k=5, category_filter="new_rfp"
            )
            rfp_context = "\n\n".join([c["content"] for c in rfp_chunks]) if rfp_chunks else ""

            # Search old response documents for relevant content
            old_response_chunks = VectorService.search(
                str(project_id), search_query, top_k=5, category_filter="old_response"
            )
            old_response_context = "\n\n".join([c["content"] for c in old_response_chunks]) if old_response_chunks else ""

            # Also search inspiration documents
            inspiration_chunks = VectorService.search(
                str(project_id), search_query, top_k=3, category_filter="inspiration"
            )
            inspiration_context = "\n\n".join([c["content"] for c in inspiration_chunks]) if inspiration_chunks else ""

            # Auto-detect or load target chapter
            _update("matching", 20, "Identification du meilleur chapitre...")
            target_chapter = None
            if chapter_id_override:
                ch_result = await db.execute(
                    select(Chapter)
                    .where(Chapter.id == uuid.UUID(chapter_id_override))
                    .where(Chapter.project_id == project_id)
                )
                target_chapter = ch_result.scalar_one_or_none()
            elif inject:
                target_chapter, _score = await _find_best_chapter(db, project_id, search_query)

            existing_chapter_content = ""
            chapter_title = ""
            chapter_description = ""
            anon_chapter_title = ""
            resolved_chapter_id = None
            if target_chapter:
                existing_chapter_content = target_chapter.content or ""
                chapter_title = target_chapter.title or ""
                chapter_description = target_chapter.description or ""
                anon_chapter_title = await AnonymizationService.apply_existing_mappings(
                    chapter_title, project_id, db
                )
                resolved_chapter_id = str(target_chapter.id)

            _update("anonymizing", 28, f"Preparation (chapitre: {chapter_title or 'auto'})...",
                    chapter_id=resolved_chapter_id, chapter_title=chapter_title)

            # Load analyzed images for potential insertion
            available_images = await _load_project_images(
                db, project_id, chapter_title or recommendation, chapter_description or missing_description,
            )

            # Anonymize all texts
            anon_rec = await AnonymizationService.anonymize_text(recommendation, project_id, db)
            anon_rfp = await AnonymizationService.anonymize_text(rfp_context, project_id, db) if rfp_context else ""
            anon_old_response = await AnonymizationService.anonymize_text(old_response_context, project_id, db) if old_response_context else ""
            anon_inspiration = await AnonymizationService.anonymize_text(inspiration_context, project_id, db) if inspiration_context else ""
            anon_existing = await AnonymizationService.anonymize_text(existing_chapter_content, project_id, db) if existing_chapter_content else ""
            anon_missing = await AnonymizationService.anonymize_text(missing_description, project_id, db) if missing_description else ""
        # DB released

        # ── Phase 2: AI generation (NO DB connection held) ──
        _update("generating", 40, "Generation IA en cours...",
                chapter_id=resolved_chapter_id, chapter_title=chapter_title)

        system_prompt = """Tu es un expert senior en réponse aux appels d'offres.
À partir d'une lacune ou recommandation identifiée lors d'une analyse de conformité,
tu dois générer un contenu structuré qui comble COMPLÈTEMENT et EXPLICITEMENT cette lacune.

OBJECTIF CRITIQUE:
Le contenu généré doit répondre DIRECTEMENT et DE MANIÈRE EXHAUSTIVE à l'exigence identifiée
comme manquante. Lorsqu'une re-analyse de conformité sera effectuée, ce contenu doit permettre
de passer l'exigence de "manquant" à "complet". Sois donc TRÈS SPÉCIFIQUE et CONCRET.

Règles de rédaction:
- Rédige en français, de manière professionnelle, argumentée et convaincante.
- Utilise du markdown (sous-titres ##, listes à puces -, **gras**, tableaux si pertinent).
- Le contenu doit être directement intégrable dans un mémoire technique de réponse.
- COMMENCE par un sous-titre ## qui reprend clairement le sujet de l'exigence.
- DÉTAILLE chaque point de l'exigence avec des éléments concrets: procédures, méthodologies,
  engagements chiffrés, délais, moyens, outils, et responsabilités.
- Si une ancienne réponse est fournie, EXPLOITE ces informations pour enrichir le contenu
  (méthodologies, références, expériences, chiffres clés). Adapte-les au contexte actuel.
- Si l'ancienne réponse ne contient pas l'information nécessaire, complète avec un contenu
  pertinent et cohérent par rapport au contexte de l'appel d'offres et aux compétences attendues.
- Ne répète PAS le contenu déjà présent dans le chapitre cible.
- Si des images sont disponibles et pertinentes pour illustrer un point, insère le marqueur
  [INSERT_IMAGE:identifiant] sur sa propre ligne à l'endroit approprié.

Anonymisation:
- Le texte peut contenir des marqueurs anonymisés comme [ENTREPRISE_1], [SOLUTION_1], etc.
- Réutilise EXACTEMENT les mêmes marqueurs. N'en invente JAMAIS de nouveaux."""

        # Add identity and anti-hallucination guardrails
        from ..services.ai_service import _build_identity_block
        system_prompt += _build_identity_block(proj_company_name, proj_client_name)

        if ai_context:
            system_prompt += f"""

Contexte de rédaction (informations sur notre société et notre approche):
{ai_context}"""

        user_parts = []
        if missing_description:
            user_parts.append(
                f"ÉLÉMENT MANQUANT IDENTIFIÉ:\n"
                f"Exigence: {anon_rec}\n"
                f"Ce qui manque dans le mémoire actuel: {anon_missing}\n\n"
                f"Tu DOIS produire un contenu qui couvre EXPLICITEMENT et COMPLÈTEMENT cette exigence, "
                f"de sorte qu'un relecteur puisse valider que le mémoire répond désormais à ce point."
            )
        else:
            user_parts.append(f"RECOMMANDATION À TRAITER:\n{anon_rec}")

        if anon_rfp:
            user_parts.append(f"CONTEXTE DU CAHIER DES CHARGES (extraits pertinents):\n{anon_rfp[:5000]}")
        if anon_old_response:
            user_parts.append(f"ÉLÉMENTS DE L'ANCIENNE RÉPONSE (à exploiter et adapter):\n{anon_old_response[:5000]}")
        if anon_inspiration:
            user_parts.append(f"DOCUMENTS D'INSPIRATION (exemples et bonnes pratiques):\n{anon_inspiration[:3000]}")
        if anon_existing and anon_chapter_title:
            user_parts.append(
                f"CONTENU ACTUEL DU CHAPITRE \"{anon_chapter_title}\" (ne pas répéter, compléter):\n{anon_existing[:3000]}"
            )

        # Add available images catalog
        if available_images:
            img_lines = ["IMAGES DISPONIBLES pour illustration (insère [INSERT_IMAGE:id] si pertinent) :"]
            for img in available_images:
                img_id = img.get("id", "")
                img_desc = img.get("anonymized_description", img.get("description", ""))
                img_type = img.get("image_type", img.get("type", ""))
                img_usage = img.get("suggested_usage", "")
                line = f"- `{img_id}` [{img_type}] : {img_desc}"
                if img_usage:
                    line += f" (usage suggéré: {img_usage})"
                img_lines.append(line)
            user_parts.append("\n".join(img_lines))

        user_parts.append(
            "Génère un contenu structuré et DÉTAILLÉ (1-2 pages) qui comble COMPLÈTEMENT cette lacune. "
            "Le contenu doit être suffisamment spécifique et exhaustif pour que l'exigence soit "
            "considérée comme couverte lors d'une prochaine analyse de conformité. "
            "Inclus des éléments concrets: modalités, procédures, engagements, délais, moyens."
        )

        content = await ai_service.generate(system_prompt, "\n\n".join(user_parts), max_tokens=6000)

        # Log AI usage for gap remediation
        async with task_session() as usage_db:
            await log_ai_usage_from_service(usage_db, project_id, "gap_remediation", ai_service)

        # ── Phase 3: Deanonymize + save (short DB session) ──
        _update("deanonymizing", 80, "Deanonymisation...",
                chapter_id=resolved_chapter_id, chapter_title=chapter_title)

        async with task_session() as db:
            content = await AnonymizationService.deanonymize_text(content, project_id, db)

            # Extract image references from the generated content
            image_refs = re.findall(r'\[INSERT_IMAGE:([^\]]+)\]', content)
            new_image_references = []
            if image_refs and available_images:
                img_lookup = {str(img["id"]): img for img in available_images}
                for ref_id in image_refs:
                    ref_id_clean = ref_id.strip()
                    if ref_id_clean in img_lookup:
                        img = img_lookup[ref_id_clean]
                        new_image_references.append({
                            "image_id": ref_id_clean,
                            "file_path": img.get("file_path", ""),
                            "description": img.get("description", ""),
                            "image_type": img.get("image_type", ""),
                        })

            if inject and resolved_chapter_id:
                _update("saving", 90, f"Injection dans '{chapter_title}'...",
                        chapter_id=resolved_chapter_id, chapter_title=chapter_title)
                ch_result = await db.execute(
                    select(Chapter).where(Chapter.id == uuid.UUID(resolved_chapter_id))
                )
                chapter = ch_result.scalar_one_or_none()
                if chapter:
                    separator = "\n\n---\n\n" if chapter.content else ""
                    chapter.content = (chapter.content or "") + separator + content
                    # Merge new image references with existing ones
                    if new_image_references:
                        existing_refs = chapter.image_references or []
                        existing_ids = {r.get("image_id") for r in existing_refs}
                        for ref in new_image_references:
                            if ref["image_id"] not in existing_ids:
                                existing_refs.append(ref)
                        chapter.image_references = existing_refs
                    await db.commit()

        set_progress(_NS_REC, task_id, {
            "status": "completed", "step": "done", "progress": 100,
            "message": f"Contenu integre dans '{chapter_title}'" if inject and chapter_title else "Contenu genere",
            "chapter_id": resolved_chapter_id,
            "chapter_title": chapter_title,
            "content": content,
        })

    except Exception as e:
        logger.exception("Recommendation generation failed for task %s", task_id)
        set_progress(_NS_REC, task_id, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
            "chapter_id": None, "chapter_title": None, "content": None,
        })


@router.get("/{project_id}/compliance-analysis/export-pdf")
async def export_compliance_pdf(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export the latest compliance analysis as a PDF document."""
    import fitz  # PyMuPDF

    result = await db.execute(
        select(ComplianceResult)
        .where(ComplianceResult.project_id == project_id)
        .order_by(ComplianceResult.created_at.desc())
        .limit(1)
    )
    cr = result.scalar_one_or_none()
    if not cr:
        raise HTTPException(status_code=404, detail="Aucune analyse de conformite disponible")

    # Also fetch project name for the header
    proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = proj_result.scalar_one_or_none()
    project_name = project.name if project else "Projet"

    # Build PDF with PyMuPDF
    doc = fitz.open()

    MARGIN = 50
    PAGE_W, PAGE_H = fitz.paper_size("a4")
    TEXT_W = PAGE_W - 2 * MARGIN
    Y_BOTTOM = PAGE_H - MARGIN

    # Colors as RGB tuples (0-1 range)
    COL_TITLE = (0.0, 0.0, 0.5)
    COL_GREEN = (0.13, 0.55, 0.13)
    COL_ORANGE = (0.85, 0.55, 0.0)
    COL_RED = (0.8, 0.1, 0.1)
    COL_GRAY = (0.4, 0.4, 0.4)
    COL_BLACK = (0.0, 0.0, 0.0)
    COL_LIGHTGRAY = (0.88, 0.88, 0.88)
    COL_STEEL = (0.27, 0.51, 0.71)

    coverage_colors = {"complete": COL_GREEN, "partial": COL_ORANGE, "missing": COL_RED}
    coverage_labels = {"complete": "Complet", "partial": "Partiel", "missing": "Manquant"}

    page = doc.new_page(width=PAGE_W, height=PAGE_H)
    y = MARGIN

    def _new_page():
        nonlocal page, y
        page = doc.new_page(width=PAGE_W, height=PAGE_H)
        y = MARGIN

    def _check_space(needed: float):
        nonlocal y
        if y + needed > Y_BOTTOM:
            _new_page()

    def _write(text: str, fontsize: float = 10, color=COL_BLACK, bold: bool = False, indent: float = 0, max_width: float = 0):
        nonlocal y
        fontname = "helv" if not bold else "hebo"
        w = max_width or (TEXT_W - indent)
        # Wrap long text
        lines = []
        for paragraph in text.split("\n"):
            if not paragraph.strip():
                lines.append("")
                continue
            words = paragraph.split()
            current_line = ""
            for word in words:
                test = f"{current_line} {word}".strip()
                tw = fitz.get_text_length(test, fontname=fontname, fontsize=fontsize)
                if tw > w and current_line:
                    lines.append(current_line)
                    current_line = word
                else:
                    current_line = test
            if current_line:
                lines.append(current_line)

        line_h = fontsize * 1.4
        for line in lines:
            _check_space(line_h)
            page.insert_text(
                fitz.Point(MARGIN + indent, y + fontsize),
                line, fontsize=fontsize, fontname=fontname, color=color,
            )
            y += line_h

    # ── Title ──
    _write(f"Analyse de Conformite", fontsize=18, color=COL_TITLE, bold=True)
    _write(f"{project_name}", fontsize=12, color=COL_GRAY)
    if cr.created_at:
        _write(f"Date: {cr.created_at.strftime('%d/%m/%Y %H:%M')}", fontsize=9, color=COL_GRAY)
    y += 10

    # ── Score ──
    score = cr.score or 0
    score_color = COL_GREEN if score >= 80 else COL_ORANGE if score >= 50 else COL_RED
    _check_space(40)
    page.insert_text(fitz.Point(MARGIN, y + 24), f"{score}/100", fontsize=24, fontname="hebo", color=score_color)
    page.insert_text(fitz.Point(MARGIN + 100, y + 14), "Score de conformite", fontsize=12, fontname="hebo", color=COL_TITLE)
    y += 35

    # Score bar
    _check_space(15)
    bar_w = TEXT_W
    bar_h = 8
    page.draw_rect(fitz.Rect(MARGIN, y, MARGIN + bar_w, y + bar_h), color=None, fill=COL_LIGHTGRAY)
    fill_w = bar_w * score / 100
    page.draw_rect(fitz.Rect(MARGIN, y, MARGIN + fill_w, y + bar_h), color=None, fill=score_color)
    y += bar_h + 10

    # Summary
    if cr.summary:
        _write(cr.summary, fontsize=10, color=COL_GRAY)
    y += 15

    # ── Covered requirements ──
    reqs = cr.covered_requirements or []
    if reqs:
        _write(f"Exigences couvertes ({len(reqs)})", fontsize=14, color=COL_TITLE, bold=True)
        y += 5
        for req in reqs:
            coverage = req.get("coverage", "missing")
            cov_label = coverage_labels.get(coverage, coverage)
            cov_color = coverage_colors.get(coverage, COL_BLACK)

            _check_space(50)
            # Requirement title with coverage badge
            _write(f"[{cov_label}] {req.get('requirement', '')}", fontsize=10, bold=True, indent=10, color=cov_color)

            # Comment
            comment = req.get("comment", "")
            if comment:
                _write(comment, fontsize=9, color=COL_GRAY, indent=20)

            # Sources
            src_rfp = req.get("source_rfp", "")
            src_resp = req.get("source_response", "")
            if src_rfp or src_resp:
                sources = []
                if src_rfp:
                    sources.append(f"AO: {src_rfp}")
                if src_resp:
                    sources.append(f"Reponse: {src_resp}")
                _write(" | ".join(sources), fontsize=8, color=COL_STEEL, indent=20)
            y += 5

    # ── Missing elements ──
    missing = cr.missing_elements or []
    if missing:
        y += 10
        _write(f"Elements manquants ({len(missing)})", fontsize=14, color=COL_RED, bold=True)
        y += 5
        for elem in missing:
            _check_space(40)
            _write(f"- {elem.get('requirement', '')}", fontsize=10, bold=True, indent=10, color=COL_RED)
            desc = elem.get("description", "")
            if desc:
                _write(desc, fontsize=9, color=COL_GRAY, indent=20)
            src_rfp = elem.get("source_rfp", "")
            if src_rfp:
                _write(f"AO: {src_rfp}", fontsize=8, color=COL_STEEL, indent=20)
            y += 5

    # ── Recommendations ──
    recs = cr.recommendations or []
    if recs:
        y += 10
        _write(f"Recommandations ({len(recs)})", fontsize=14, color=COL_TITLE, bold=True)
        y += 5
        for i, rec in enumerate(recs, 1):
            _check_space(30)
            _write(f"{i}. {rec}", fontsize=10, indent=10)
            y += 3

    # Save to bytes
    pdf_bytes = doc.tobytes()
    doc.close()

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="conformite_{project_name}.pdf"'},
    )


import json
from datetime import datetime as dt_datetime


def _parse_improvement_axes(raw: str) -> list[dict]:
    """Parse improvement_axes field. Supports JSON array format or legacy text format."""
    if not raw:
        return []
    raw = raw.strip()
    if raw.startswith("["):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            pass
    # Legacy text format: convert lines starting with "- " to structured items
    items = []
    for i, line in enumerate(raw.split("\n")):
        line = line.strip()
        if line.startswith("- "):
            line = line[2:]
        if not line:
            continue
        # Try to extract source info from "(Source: ...)" at end
        source = ""
        if " (Source: " in line and line.endswith(")"):
            idx = line.rfind(" (Source: ")
            source = line[idx + 10:-1]
            line = line[:idx]
        items.append({
            "id": str(uuid.uuid4()),
            "content": line,
            "source": source,
            "created_at": dt_datetime.now().isoformat(),
        })
    return items


def _serialize_improvement_axes(items: list[dict]) -> str:
    return json.dumps(items, ensure_ascii=False)


@router.get("/{project_id}/improvement-axes")
async def list_improvement_axes(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all improvement axes for a project."""
    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    raw = project.improvement_axes or ""
    items = _parse_improvement_axes(raw)
    # Persist JSON format if data was in legacy text format so IDs remain stable
    if items and not raw.strip().startswith("["):
        project.improvement_axes = _serialize_improvement_axes(items)
        await db.commit()
    return {"axes": items}


@router.post("/{project_id}/improvement-axes")
async def add_improvement_axis(
    project_id: uuid.UUID,
    request: ImprovementAxisRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add an improvement axis from client feedback."""
    moderation = moderate_prompt(request.content, "improvement_axis")
    if not moderation:
        raise HTTPException(status_code=422, detail=moderation.message)

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    items = _parse_improvement_axes(project.improvement_axes or "")
    new_item = {
        "id": str(uuid.uuid4()),
        "content": request.content,
        "source": request.source or "",
        "created_at": dt_datetime.now().isoformat(),
    }
    items.append(new_item)
    project.improvement_axes = _serialize_improvement_axes(items)

    await db.commit()
    return {"success": True, "message": "Axe d'amélioration ajouté", "axis": new_item}


@router.put("/{project_id}/improvement-axes/{axis_id}")
async def update_improvement_axis(
    project_id: uuid.UUID,
    axis_id: str,
    request: ImprovementAxisUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update an existing improvement axis."""
    moderation = moderate_prompt(request.content, "improvement_axis")
    if not moderation:
        raise HTTPException(status_code=422, detail=moderation.message)

    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    items = _parse_improvement_axes(project.improvement_axes or "")
    for item in items:
        if item.get("id") == axis_id:
            item["content"] = request.content
            item["source"] = request.source or ""
            project.improvement_axes = _serialize_improvement_axes(items)
            await db.commit()
            return {"success": True, "message": "Axe mis à jour", "axis": item}

    raise HTTPException(status_code=404, detail="Axe non trouvé")


@router.delete("/{project_id}/improvement-axes/{axis_id}")
async def delete_improvement_axis(
    project_id: uuid.UUID,
    axis_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete an improvement axis."""
    result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    items = _parse_improvement_axes(project.improvement_axes or "")
    new_items = [item for item in items if item.get("id") != axis_id]
    if len(new_items) == len(items):
        raise HTTPException(status_code=404, detail="Axe non trouvé")

    project.improvement_axes = _serialize_improvement_axes(new_items)
    await db.commit()
    return {"success": True, "message": "Axe supprimé"}


@router.get("/{project_id}/statistics", response_model=StatisticsOut)
async def get_statistics(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get project statistics."""
    # Count chapters by status
    chapters_result = await db.execute(
        select(Chapter).where(Chapter.project_id == project_id)
    )
    chapters = chapters_result.scalars().all()

    total_content = " ".join([c.content for c in chapters if c.content])
    total_words = len(total_content.split()) if total_content else 0
    total_chars = len(total_content) if total_content else 0
    total_pages = max(1, total_words // 300)  # ~300 words per page estimate

    completed = sum(1 for c in chapters if c.status == ChapterStatus.COMPLETED)
    in_progress = sum(1 for c in chapters if c.status == ChapterStatus.IN_PROGRESS)

    # Count documents
    doc_count = (await db.execute(
        select(func.count()).where(Document.project_id == project_id)
    )).scalar() or 0

    # Count anonymized entities
    anon_count = (await db.execute(
        select(func.count()).where(AnonymizationMapping.project_id == project_id)
    )).scalar() or 0

    # Count images
    from ..models.document import DocumentImage
    img_count_result = await db.execute(
        select(func.count())
        .select_from(DocumentImage)
        .join(Document, Document.id == DocumentImage.document_id)
        .where(Document.project_id == project_id)
    )
    img_count = img_count_result.scalar() or 0

    completion = (completed / len(chapters) * 100) if chapters else 0

    # Build chapters_by_status breakdown
    by_status: Dict[str, int] = {}
    for c in chapters:
        s = c.status.value if hasattr(c.status, 'value') else str(c.status)
        by_status[s] = by_status.get(s, 0) + 1

    return StatisticsOut(
        total_pages=total_pages,
        total_words=total_words,
        total_characters=total_chars,
        anonymized_entities=anon_count,
        chapters_completed=completed,
        chapters_total=len(chapters),
        chapters_in_progress=in_progress,
        documents_count=doc_count,
        images_count=img_count,
        completion_percentage=round(completion, 1),
        chapters_by_status=by_status,
    )


@router.get("/{project_id}/anonymization-mappings", response_model=list[AnonymizationMappingOut])
async def get_anonymization_mappings(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get all anonymization mappings for a project."""
    result = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.project_id == project_id)
        .order_by(AnonymizationMapping.entity_type)
    )
    mappings = result.scalars().all()

    return [
        AnonymizationMappingOut(
            id=str(m.id),
            entity_type=m.entity_type.value,
            original_value=m.original_value,
            anonymized_value=m.anonymized_value,
            is_active=m.is_active,
        )
        for m in mappings
    ]


@router.get("/{project_id}/anonymization-report", response_model=AnonymizationReportOut)
async def get_anonymization_report(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get a structured anonymization report with statistics and samples."""
    from ..models.project import EntityType
    from collections import defaultdict

    result = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.project_id == project_id)
        .order_by(AnonymizationMapping.entity_type, AnonymizationMapping.created_at)
    )
    all_mappings = result.scalars().all()

    # Group by entity type
    groups_dict = defaultdict(list)
    for m in all_mappings:
        groups_dict[m.entity_type].append(m)

    # Human-readable labels
    type_labels = {
        EntityType.COMPANY: "Entreprises / Organisations",
        EntityType.PERSON: "Personnes",
        EntityType.EMAIL: "Adresses email",
        EntityType.PHONE: "Numéros de téléphone",
        EntityType.ADDRESS: "Adresses postales",
        EntityType.PROJECT_CODE: "Codes projet",
        EntityType.RFP_CODE: "Codes AO",
        EntityType.SOLUTION_NAME: "Noms de solutions",
        EntityType.DATE: "Dates",
        EntityType.AMOUNT: "Montants",
        EntityType.OTHER: "Autres entités",
    }

    entity_groups = []
    for entity_type in EntityType:
        mappings_for_type = groups_dict.get(entity_type, [])
        if not mappings_for_type:
            continue
        entity_groups.append(AnonymizationEntityGroup(
            entity_type=entity_type.value,
            label=type_labels.get(entity_type, entity_type.value),
            count=len(mappings_for_type),
            mappings=[
                AnonymizationMappingOut(
                    id=str(m.id),
                    entity_type=m.entity_type.value,
                    original_value=m.original_value,
                    anonymized_value=m.anonymized_value,
                    is_active=m.is_active,
                )
                for m in mappings_for_type
            ],
        ))

    # Generate a sample before/after from a real document chunk
    sample_before = ""
    sample_after = ""
    chunk_result = await db.execute(
        select(DocumentChunk)
        .join(Document, Document.id == DocumentChunk.document_id)
        .where(Document.project_id == project_id)
        .where(DocumentChunk.anonymized_content != "")
        .where(DocumentChunk.anonymized_content != DocumentChunk.content)
        .limit(1)
    )
    sample_chunk = chunk_result.scalar_one_or_none()
    if sample_chunk:
        sample_before = sample_chunk.content[:500]
        sample_after = sample_chunk.anonymized_content[:500]

    active_count = sum(1 for m in all_mappings if m.is_active)

    return AnonymizationReportOut(
        total_entities=len(all_mappings),
        active_entities=active_count,
        entity_groups=entity_groups,
        sample_before=sample_before,
        sample_after=sample_after,
    )


# ── Anonymization Mapping CRUD ──────────────────────────────────────

@router.post("/{project_id}/anonymization-mappings", response_model=AnonymizationMappingOut, status_code=201)
async def create_anonymization_mapping(
    project_id: uuid.UUID,
    request: AnonymizationMappingCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a new anonymization mapping manually."""
    from ..services.anonymization_service import ENTITY_PREFIXES
    from collections import defaultdict

    # Validate entity type
    try:
        entity_type = EntityType(request.entity_type)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Type d'entité invalide: {request.entity_type}")

    # Check for duplicate original_value (done in Python since values are encrypted at rest)
    existing = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.project_id == project_id)
    )
    for m in existing.scalars().all():
        if m.original_value == request.original_value:
            raise HTTPException(status_code=409, detail="Cette valeur originale existe déjà dans les mappings")

    # Auto-generate placeholder if not provided
    anonymized_value = request.anonymized_value
    if not anonymized_value:
        # Count existing mappings of this type to generate next placeholder
        count_result = await db.execute(
            select(func.count())
            .where(AnonymizationMapping.project_id == project_id)
            .where(AnonymizationMapping.entity_type == entity_type)
        )
        count = count_result.scalar() or 0
        prefix = ENTITY_PREFIXES.get(entity_type, "ENTITE")
        anonymized_value = f"[{prefix}_{count + 1}]"

    mapping = AnonymizationMapping(
        project_id=project_id,
        entity_type=entity_type,
        original_value=request.original_value,
        anonymized_value=anonymized_value,
    )
    db.add(mapping)
    await db.commit()
    await db.refresh(mapping)

    return AnonymizationMappingOut(
        id=str(mapping.id),
        entity_type=mapping.entity_type.value,
        original_value=mapping.original_value,
        anonymized_value=mapping.anonymized_value,
        is_active=mapping.is_active,
    )


@router.put("/{project_id}/anonymization-mappings/{mapping_id}", response_model=AnonymizationMappingOut)
async def update_anonymization_mapping(
    project_id: uuid.UUID,
    mapping_id: uuid.UUID,
    request: AnonymizationMappingUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update an existing anonymization mapping."""
    result = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.id == mapping_id)
        .where(AnonymizationMapping.project_id == project_id)
    )
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping non trouvé")

    if request.original_value is not None:
        mapping.original_value = request.original_value
    if request.anonymized_value is not None:
        mapping.anonymized_value = request.anonymized_value
    if request.entity_type is not None:
        try:
            mapping.entity_type = EntityType(request.entity_type)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Type d'entité invalide: {request.entity_type}")
    if request.is_active is not None:
        mapping.is_active = request.is_active

    await db.commit()
    await db.refresh(mapping)

    return AnonymizationMappingOut(
        id=str(mapping.id),
        entity_type=mapping.entity_type.value,
        original_value=mapping.original_value,
        anonymized_value=mapping.anonymized_value,
        is_active=mapping.is_active,
    )


@router.delete("/{project_id}/anonymization-mappings/{mapping_id}", status_code=204)
async def delete_anonymization_mapping(
    project_id: uuid.UUID,
    mapping_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete an anonymization mapping."""
    result = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.id == mapping_id)
        .where(AnonymizationMapping.project_id == project_id)
    )
    mapping = result.scalar_one_or_none()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping non trouvé")

    await db.delete(mapping)
    await db.commit()


@router.post("/{project_id}/re-anonymize")
async def re_anonymize_project(
    project_id: uuid.UUID,

    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch re-anonymization as a background task (NER is slow on many chunks).

    Returns a task_id to poll progress via GET /{project_id}/re-anonymize-status.
    """
    pid = str(project_id)

    # Prevent double-launch
    existing = get_or_idle(_NS_REANON, pid)
    if existing and existing.get("status") == "running":
        return {"task_id": pid, "already_running": True}

    set_progress(_NS_REANON, pid, {
        "status": "running",
        "progress": 0,
        "current": 0,
        "total": 0,
        "phase": "init",
        "message": "Demarrage de la re-anonymisation...",
    })

    # Read workspace_id while we still have the request DB session
    project_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouve")

    from ..tasks.project_tasks import reanonymize_task
    reanonymize_task.delay(str(project_id))
    return {"task_id": pid}


@router.get("/{project_id}/re-anonymize-status")
async def get_reanonymize_status(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll re-anonymization progress."""
    pid = str(project_id)
    return get_or_idle(_NS_REANON, pid)


@router.get("/{project_id}/ner-diagnostic")
async def get_ner_diagnostic(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Test NER connectivity and optionally run a quick detection test.

    Returns diagnostic info: Ollama reachability, model status, and optionally
    results from a sample NER call so the user can verify it works.
    """
    from ..services.anonymization_service import AnonymizationService

    diag = AnonymizationService.get_ner_diagnostic()

    # Run a live connectivity check (don't use cache)
    AnonymizationService._ollama_available = None  # reset cache to force fresh check
    ollama_ok = await AnonymizationService._check_ollama()
    diag["ollama_reachable"] = ollama_ok
    diag["failure_reason"] = AnonymizationService._last_ner_failure_reason

    # If Ollama is reachable, do a quick test NER call with known entities
    test_result = None
    if ollama_ok:
        sample_text = (
            "La société SCC France, éditrice de la solution Atrium FinOps, "
            "propose une réponse complète. Le chef de projet est Jean Dupont "
            "(jean.dupont@scc.fr)."
        )
        try:
            entities = await AnonymizationService._detect_entities_llm(sample_text)
            test_result = {
                "sample_text": sample_text,
                "entities_found": [
                    {"text": e[0], "type": e[1]} for e in entities
                ],
                "count": len(entities),
                "status": "ok" if entities else "empty_response",
            }
        except Exception as e:
            test_result = {
                "sample_text": sample_text,
                "entities_found": [],
                "count": 0,
                "status": f"error: {e}",
            }

    diag["test_result"] = test_result
    return diag


async def _run_reanonymize(project_id: uuid.UUID):
    """Background task: re-anonymize all document chunks and chapter content."""
    from ..database import task_session
    from ..services.anonymization_service import AnonymizationService

    pid = str(project_id)

    def _update(phase: str, progress: int, message: str, current: int = 0, total: int = 0):
        set_progress(_NS_REANON, pid, {
            "status": "running",
            "progress": progress,
            "current": current,
            "total": total,
            "phase": phase,
            "message": message,
        })

    try:
        # ── Phase 1: NER detection on document chunks ──
        _update("loading", 5, "Chargement des chunks...")

        async with task_session() as db:
            chunks_result = await db.execute(
                select(DocumentChunk)
                .join(Document, Document.id == DocumentChunk.document_id)
                .where(Document.project_id == project_id)
            )
            chunks = chunks_result.scalars().all()
            chunk_texts = [chunk.content for chunk in chunks if chunk.content]
            chunk_ids = [chunk.id for chunk in chunks if chunk.content]

        total_chunks = len(chunk_texts)
        new_entities = 0

        if total_chunks > 0:
            _update("ner", 10, f"Detection NER sur {total_chunks} chunks...", 0, total_chunks)

            def on_ner_progress(current: int, total: int):
                pct = 10 + int(70 * current / total)
                _update("ner", pct, f"Detection NER : {current}/{total} chunks analyses", current, total)

            async with task_session() as db:
                # Count mappings before
                before_count_result = await db.execute(
                    select(func.count(AnonymizationMapping.id))
                    .where(AnonymizationMapping.project_id == project_id)
                )
                before_count = before_count_result.scalar() or 0

                # Run NER (this is the slow part)
                anonymized_texts = await AnonymizationService.anonymize_chunks_batch(
                    chunk_texts, project_id, db,
                    progress_callback=on_ner_progress,
                )
                await db.flush()

                # Count new entities
                after_count_result = await db.execute(
                    select(func.count(AnonymizationMapping.id))
                    .where(AnonymizationMapping.project_id == project_id)
                )
                after_count = after_count_result.scalar() or 0
                new_entities = after_count - before_count

                # Save anonymized content to chunks
                _update("saving_chunks", 82, "Enregistrement des chunks anonymises...")
                for i, anon_text in enumerate(anonymized_texts):
                    chunk_result = await db.execute(
                        select(DocumentChunk).where(DocumentChunk.id == chunk_ids[i])
                    )
                    chunk = chunk_result.scalar_one_or_none()
                    if chunk:
                        chunk.anonymized_content = anon_text

                await db.commit()

        # ── Phase 2: Apply all active mappings to chapters ──
        _update("chapters", 85, "Application des mappings aux chapitres...")

        async with task_session() as db:
            all_mappings_result = await db.execute(
                select(AnonymizationMapping)
                .where(AnonymizationMapping.project_id == project_id)
                .where(AnonymizationMapping.is_active == True)
            )
            all_mappings = all_mappings_result.scalars().all()
            active_with_value = sorted(
                [m for m in all_mappings if m.original_value],
                key=lambda m: len(m.original_value),
                reverse=True,
            )

            def apply_mappings(text: str) -> str:
                result_text = text
                for m in active_with_value:
                    pattern = re.compile(re.escape(m.original_value), re.IGNORECASE)
                    result_text = pattern.sub(m.anonymized_value, result_text)
                return result_text

            chapters_result = await db.execute(
                select(Chapter).where(Chapter.project_id == project_id)
            )
            chapters = chapters_result.scalars().all()
            updated_chapters = 0
            for ch in chapters:
                if ch.content:
                    new_anon = apply_mappings(ch.content)
                    if new_anon != ch.anonymized_content:
                        ch.anonymized_content = new_anon
                        updated_chapters += 1

            await db.commit()

        ner_available = AnonymizationService.is_ner_available()
        ner_diagnostic = AnonymizationService.get_ner_diagnostic()

        # Build a clear user-facing message
        if not ner_available:
            message = (
                f"Re-anonymisation terminée (regex uniquement — "
                f"NER indisponible: {ner_diagnostic.get('failure_reason', 'raison inconnue')}). "
                f"Seuls les emails et téléphones ont été détectés."
            )
        elif new_entities == 0 and ner_diagnostic.get("last_ner_produced_entities") is False:
            message = (
                f"Re-anonymisation terminée mais le modèle NER n'a retourné aucune entité. "
                f"Raison possible: {ner_diagnostic.get('failure_reason', 'réponse vide du modèle')}. "
                f"Vérifiez les logs du worker Celery."
            )
        elif new_entities == 0:
            message = "Re-anonymisation terminée — aucune nouvelle entité détectée."
        else:
            message = f"Re-anonymisation terminée — {new_entities} nouvelle(s) entité(s) détectée(s)."

        set_progress(_NS_REANON, pid, {
            "status": "done",
            "progress": 100,
            "current": total_chunks,
            "total": total_chunks,
            "phase": "done",
            "message": message,
            "updated_chunks": total_chunks,
            "updated_chapters": updated_chapters,
            "new_entities": new_entities,
            "ner_available": ner_available,
            "ner_diagnostic": ner_diagnostic,
        })

    except Exception as e:
        logger.error("Re-anonymize background task failed: %s", e, exc_info=True)
        set_progress(_NS_REANON, pid, {
            "status": "error",
            "progress": 0,
            "phase": "error",
            "message": f"Erreur: {str(e)}",
        })


@router.get("/{project_id}/chapters/{chapter_id}/anonymized-content")
async def get_chapter_anonymized_content(
    project_id: uuid.UUID,
    chapter_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get the anonymized version of a chapter's content (what the AI sees)."""
    result = await db.execute(
        select(Chapter)
        .where(Chapter.id == chapter_id)
        .where(Chapter.project_id == project_id)
    )
    chapter = result.scalar_one_or_none()
    if not chapter:
        raise HTTPException(status_code=404, detail="Chapitre non trouvé")

    if not chapter.content:
        return {"anonymized_content": ""}

    # Apply all active mappings to get the anonymized version
    mappings_result = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.project_id == project_id)
        .where(AnonymizationMapping.is_active == True)
    )
    mappings = mappings_result.scalars().all()
    mappings_sorted = sorted(mappings, key=lambda m: len(m.original_value), reverse=True)

    anonymized = chapter.content
    for m in mappings_sorted:
        anonymized = anonymized.replace(m.original_value, m.anonymized_value)

    return {"anonymized_content": anonymized}


@router.post("/{project_id}/resolve-orphans-ai")
async def resolve_orphans_with_ai(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Use AI to analyze context around orphan placeholders and suggest real values."""
    from ..services.anonymization_service import AnonymizationService

    # Need AI service for the project's workspace
    project_result = await db.execute(
        select(RFPProject).where(RFPProject.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    ai_service = await _get_ai_service(project.workspace_id, db)
    result = await AnonymizationService.resolve_orphans_with_ai(project_id, db, ai_service)

    # Log AI usage for orphan resolution
    await log_ai_usage_from_service(db, project_id, "resolve_orphans", ai_service)

    await db.commit()
    return result


@router.post("/{project_id}/consolidate-mappings")
async def consolidate_mappings(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Find and merge duplicate anonymization mappings for the same entity."""
    from ..services.anonymization_service import AnonymizationService

    result = await AnonymizationService.consolidate_mappings(project_id, db)
    await db.commit()
    return result


@router.post("/{project_id}/purge-anonymization")
async def purge_anonymization(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Remove ALL anonymization: deanonymize chapter content, clear anonymized fields, delete all mappings.

    This restores the project to its original non-anonymized state.
    Steps:
    1. Replace all [PLACEHOLDER_N] tokens in chapter.content with their original values
    2. Clear chapter.anonymized_content
    3. Clear document_chunk.anonymized_content
    4. Clear document.anonymized_full_text
    5. Delete all anonymization mappings
    """
    from ..services.anonymization_service import ENTITY_PREFIXES

    # Get all mappings (including inactive) to restore original values
    all_mappings_result = await db.execute(
        select(AnonymizationMapping)
        .where(AnonymizationMapping.project_id == project_id)
    )
    all_mappings = all_mappings_result.scalars().all()

    # Build placeholder → original_value map
    placeholder_to_original = {
        m.anonymized_value: m.original_value
        for m in all_mappings
        if m.original_value  # skip unresolved
    }

    # Also build a regex to catch any placeholder pattern (even orphans)
    prefix_pattern = '|'.join(re.escape(p) for p in ENTITY_PREFIXES.values())
    placeholder_re = re.compile(r'\[(?:' + prefix_pattern + r')_\d+\]')

    def deanonymize(text: str) -> str:
        """Replace known placeholders with original values, remove unknown ones."""
        if not text:
            return text
        result = text
        # Replace known placeholders (longest first to avoid partial matches)
        for placeholder, original in sorted(placeholder_to_original.items(), key=lambda x: len(x[0]), reverse=True):
            result = result.replace(placeholder, original)
        # Remove any remaining unknown placeholders (leave text clean)
        result = placeholder_re.sub('', result)
        # Clean up double spaces left by removed placeholders
        result = re.sub(r'  +', ' ', result)
        return result

    # 1. Deanonymize chapter content
    chapters_result = await db.execute(
        select(Chapter).where(Chapter.project_id == project_id)
    )
    chapters = chapters_result.scalars().all()
    restored_chapters = 0
    for ch in chapters:
        changed = False
        if ch.content and placeholder_re.search(ch.content):
            ch.content = deanonymize(ch.content)
            changed = True
        if ch.anonymized_content:
            ch.anonymized_content = ""
            changed = True
        if changed:
            restored_chapters += 1

    # 2. Clear anonymized fields on document chunks
    chunks_result = await db.execute(
        select(DocumentChunk)
        .join(Document, Document.id == DocumentChunk.document_id)
        .where(Document.project_id == project_id)
    )
    chunks = chunks_result.scalars().all()
    cleared_chunks = 0
    for chunk in chunks:
        if chunk.anonymized_content:
            chunk.anonymized_content = ""
            cleared_chunks += 1

    # 3. Clear anonymized_full_text on documents
    docs_result = await db.execute(
        select(Document).where(Document.project_id == project_id)
    )
    docs = docs_result.scalars().all()
    for doc in docs:
        if doc.anonymized_full_text:
            doc.anonymized_full_text = ""

    # 4. Delete all mappings
    deleted_mappings = len(all_mappings)
    for m in all_mappings:
        await db.delete(m)

    await db.commit()

    return {
        "restored_chapters": restored_chapters,
        "cleared_chunks": cleared_chunks,
        "deleted_mappings": deleted_mappings,
    }


# ── Fields to Complete (AI-invented placeholders) ──────────────────

@router.get("/{project_id}/fields-to-complete", response_model=FieldsToCompleteOut)
async def get_fields_to_complete(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Scan chapter content for AI-invented placeholders that need user input.

    When Mistral doesn't have certain information (year of creation, share
    capital, SIRET number, etc.), it inserts placeholder tokens like
    [ANNÉE_DE_CRÉATION] or [CAPITAL_SOCIAL]. This endpoint finds them all
    and returns them so the user can fill them in.
    """
    from ..services.anonymization_service import AnonymizationService

    # Get known anonymization placeholders so we can exclude them
    mappings = await AnonymizationService.get_mappings_by_placeholder(db, project_id)
    known_placeholders = set(mappings.keys())

    chapters_result = await db.execute(
        select(Chapter).where(Chapter.project_id == project_id)
    )
    chapters = chapters_result.scalars().all()

    # Scan all chapters for AI-invented fields
    placeholder_info: dict = {}  # placeholder -> {occurrences, chapters, chapter_details}
    for ch in chapters:
        if not ch.content:
            continue
        fields = AnonymizationService.find_ai_fields_to_complete(ch.content, known_placeholders)
        for field in fields:
            if field not in placeholder_info:
                placeholder_info[field] = {"occurrences": 0, "chapters": set(), "chapter_details": []}
            placeholder_info[field]["occurrences"] += ch.content.count(field)
            ch_title = ch.title or f"Chapitre {ch.numbering}"
            placeholder_info[field]["chapters"].add(ch_title)
            placeholder_info[field]["chapter_details"].append({
                "chapter_id": str(ch.id),
                "title": ch_title,
                "numbering": ch.numbering or "",
            })

    fields_out = []
    for placeholder, info in sorted(placeholder_info.items()):
        # Build a human-readable label from the placeholder:
        # [ANNÉE_DE_CRÉATION] → "Année de création"
        inner = placeholder.strip("[]")
        readable = inner.replace("_", " ").capitalize()
        fields_out.append(FieldToComplete(
            placeholder=placeholder,
            readable_label=readable,
            occurrences=info["occurrences"],
            chapters=sorted(info["chapters"]),
            chapter_details=info["chapter_details"],
        ))

    return FieldsToCompleteOut(total=len(fields_out), fields=fields_out)


@router.post("/{project_id}/fields-to-complete/replace")
async def replace_field_to_complete(
    project_id: uuid.UUID,
    request: FieldReplaceRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Replace an AI-invented placeholder with a real value across all chapters.

    For example, replace [ANNÉE_DE_CRÉATION] with "2014" everywhere.
    """
    if not request.placeholder or not request.value:
        raise HTTPException(status_code=400, detail="placeholder and value are required")

    chapters_result = await db.execute(
        select(Chapter).where(Chapter.project_id == project_id)
    )
    chapters = chapters_result.scalars().all()

    updated_chapters = 0
    total_replacements = 0
    for ch in chapters:
        if ch.content and request.placeholder in ch.content:
            count = ch.content.count(request.placeholder)
            ch.content = ch.content.replace(request.placeholder, request.value)
            updated_chapters += 1
            total_replacements += count

    await db.commit()

    return {
        "updated_chapters": updated_chapters,
        "total_replacements": total_replacements,
        "placeholder": request.placeholder,
        "value": request.value,
    }


# ── Content Reuse Statistics ────────────────────────────────────────

@router.get("/{project_id}/content-reuse-stats")
async def get_content_reuse_stats(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Compute content reuse statistics between old response and generated chapters.

    Uses semantic embedding similarity (via the project's ChromaDB index) to
    detect content reuse even when the AI has paraphrased or reformulated the
    original text.  Each chapter is split into ~300-word chunks, each chunk is
    compared to the OLD_RESPONSE embeddings, and the average of the best-match
    similarity scores gives the chapter-level reuse percentage.
    """

    _EMPTY_RESPONSE = {
        "has_old_response": False,
        "overall_reuse_percentage": 0,
        "chapters": [],
        "summary": {
            "total_chapters": 0,
            "chapters_with_reuse": 0,
            "avg_reuse_percentage": 0,
            "old_response_word_count": 0,
            "new_content_word_count": 0,
        },
    }

    # ------------------------------------------------------------------
    # 1. Check that OLD_RESPONSE documents exist and are indexed
    # ------------------------------------------------------------------
    old_docs_result = await db.execute(
        select(Document)
        .where(Document.project_id == project_id)
        .where(Document.category == DocumentCategory.OLD_RESPONSE)
        .where(Document.processing_status == ProcessingStatus.COMPLETED)
    )
    old_docs = old_docs_result.scalars().all()

    if not old_docs:
        return _EMPTY_RESPONSE

    # Count old response words for the summary
    old_word_count = 0
    for doc in old_docs:
        if doc.full_text:
            old_word_count += len(doc.full_text.split())
        else:
            chunks_result = await db.execute(
                select(DocumentChunk)
                .where(DocumentChunk.document_id == doc.id)
            )
            for chunk in chunks_result.scalars().all():
                if chunk.content:
                    old_word_count += len(chunk.content.split())

    if old_word_count == 0:
        _EMPTY_RESPONSE["has_old_response"] = True
        return _EMPTY_RESPONSE

    # ------------------------------------------------------------------
    # 2. Load chapters
    # ------------------------------------------------------------------
    chapters_result = await db.execute(
        select(Chapter)
        .where(Chapter.project_id == project_id)
        .order_by(Chapter.order)
    )
    chapters = chapters_result.scalars().all()

    if not chapters:
        return {
            "has_old_response": True,
            "overall_reuse_percentage": 0,
            "chapters": [],
            "summary": {
                "total_chapters": 0,
                "chapters_with_reuse": 0,
                "avg_reuse_percentage": 0,
                "old_response_word_count": old_word_count,
                "new_content_word_count": 0,
            },
        }

    # ------------------------------------------------------------------
    # 3. Split chapters into ~300-word chunks for embedding comparison
    # ------------------------------------------------------------------
    CHUNK_SIZE = 300   # words
    CHUNK_OVERLAP = 50  # words

    def _split_into_chunks(text: str) -> list[str]:
        """Split text into overlapping word-level chunks."""
        words = text.split()
        if len(words) <= CHUNK_SIZE:
            return [text] if words else []
        chunks = []
        start = 0
        while start < len(words):
            end = start + CHUNK_SIZE
            chunks.append(" ".join(words[start:end]))
            start += CHUNK_SIZE - CHUNK_OVERLAP
        return chunks

    # Collect all chunks across all chapters (with back-references)
    all_chunks: list[str] = []
    chunk_to_chapter: list[int] = []  # index -> chapter list position
    chapter_infos: list[dict] = []

    for idx, ch in enumerate(chapters):
        if not ch.content or not ch.content.strip():
            continue
        ch_word_count = len(ch.content.split())
        ch_chunks = _split_into_chunks(ch.content)
        chapter_infos.append({
            "chapter": ch,
            "word_count": ch_word_count,
            "chunk_start": len(all_chunks),
            "chunk_count": len(ch_chunks),
        })
        for chunk_text in ch_chunks:
            all_chunks.append(chunk_text)
            chunk_to_chapter.append(len(chapter_infos) - 1)

    if not all_chunks:
        return {
            "has_old_response": True,
            "overall_reuse_percentage": 0,
            "chapters": [],
            "summary": {
                "total_chapters": 0,
                "chapters_with_reuse": 0,
                "avg_reuse_percentage": 0,
                "old_response_word_count": old_word_count,
                "new_content_word_count": 0,
            },
        }

    # ------------------------------------------------------------------
    # 4. Embed all chapter chunks in one batch & query ChromaDB
    # ------------------------------------------------------------------
    project_str = str(project_id)

    # --- Auto-reindex if ChromaDB collection is empty but DB has chunks ---
    try:
        col_count = await asyncio.to_thread(
            VectorService.collection_count, project_str,
        )
    except Exception:
        col_count = -1  # ChromaDB unreachable — skip reindex, fall back to zeros

    if col_count == 0:
        # Collection exists but is empty — rebuild from DB chunks
        all_doc_chunks_result = await db.execute(
            select(DocumentChunk)
            .join(Document, DocumentChunk.document_id == Document.id)
            .where(Document.project_id == project_id)
            .where(Document.processing_status == ProcessingStatus.COMPLETED)
        )
        db_chunks = all_doc_chunks_result.scalars().all()

        if db_chunks:
            vector_chunks = []
            for chunk in db_chunks:
                meta = chunk.metadata_json or {}
                vector_chunks.append({
                    "id": str(chunk.id),
                    "content": chunk.content or "",
                    "document_id": str(chunk.document_id),
                    "document_name": meta.get("document_name", ""),
                    "category": meta.get("category", ""),
                    "page_number": chunk.page_number or 0,
                    "section_title": chunk.section_title or "",
                    "chunk_index": chunk.chunk_index or 0,
                })

            try:
                await asyncio.to_thread(
                    VectorService.index_chunks, project_str, vector_chunks,
                )
                logger.info(
                    "[reuse-stats] Re-indexed %d chunks for project %s",
                    len(vector_chunks), project_str,
                )
            except Exception as idx_err:
                logger.warning("[reuse-stats] Re-indexing failed: %s", idx_err)

    loop = asyncio.get_event_loop()

    def _compute_semantic_scores() -> list[float]:
        """Run embedding + ChromaDB queries (CPU-bound, run in executor)."""
        collection = VectorService.get_collection(project_str)
        embed_fn = VectorService.get_embedding_function()

        # Embed chapter chunks using the E5 "query:" prefix
        prefixed = [f"query: {c}" for c in all_chunks]
        embeddings = embed_fn(prefixed)

        # Query ChromaDB for each chunk's best match in OLD_RESPONSE
        scores: list[float] = []
        QUERY_BATCH = 32
        for start in range(0, len(embeddings), QUERY_BATCH):
            batch_embs = embeddings[start:start + QUERY_BATCH]
            try:
                results = collection.query(
                    query_embeddings=batch_embs,
                    n_results=3,
                    where={"category": DocumentCategory.OLD_RESPONSE.value},
                )
            except Exception as exc:
                logger.warning("ChromaDB query failed for reuse stats: %s", exc)
                scores.extend([0.0] * len(batch_embs))
                continue

            for i in range(len(batch_embs)):
                if (results and results["distances"]
                        and i < len(results["distances"])
                        and results["distances"][i]):
                    # ChromaDB returns cosine distances; similarity = 1 - distance
                    best_distance = min(results["distances"][i])
                    similarity = max(0.0, 1.0 - best_distance)
                    scores.append(similarity)
                else:
                    scores.append(0.0)

        return scores

    loop = asyncio.get_event_loop()
    try:
        chunk_scores = await loop.run_in_executor(None, _compute_semantic_scores)
    except Exception as exc:
        logger.warning("ChromaDB unavailable for reuse stats (possibly readonly): %s", exc)
        chunk_scores = [0.0] * len(all_chunks)

    # ------------------------------------------------------------------
    # 5. Convert similarity scores to reuse percentages per chapter
    # ------------------------------------------------------------------
    # Similarity calibration for E5 multilingual embeddings:
    #   >= 0.90  → near-identical text (copy-paste)
    #   0.75-0.90 → clearly paraphrased / reformulated same content
    #   0.60-0.75 → same topic, partial overlap
    #   < 0.60   → different content
    #
    # We map [LOW_THRESHOLD, HIGH_THRESHOLD] → [0%, 100%] linearly.
    LOW_THRESHOLD = 0.55   # below this = 0% reuse
    HIGH_THRESHOLD = 0.92  # above this = 100% reuse

    def _similarity_to_reuse(sim: float) -> float:
        """Map a cosine similarity score to a reuse percentage."""
        if sim <= LOW_THRESHOLD:
            return 0.0
        if sim >= HIGH_THRESHOLD:
            return 100.0
        return (sim - LOW_THRESHOLD) / (HIGH_THRESHOLD - LOW_THRESHOLD) * 100.0

    chapter_stats = []
    total_new_words = 0
    total_reused_words = 0

    for info in chapter_infos:
        ch = info["chapter"]
        ch_word_count = info["word_count"]
        total_new_words += ch_word_count

        # Get scores for this chapter's chunks
        start = info["chunk_start"]
        count = info["chunk_count"]
        ch_scores = chunk_scores[start:start + count]

        if not ch_scores:
            chapter_stats.append({
                "chapter_id": str(ch.id),
                "title": ch.title,
                "numbering": ch.numbering or "",
                "word_count": ch_word_count,
                "reuse_percentage": 0,
                "semantic_similarity": 0,
                "ngram_match": 0,
                "sequence_match": 0,
            })
            continue

        # Convert each chunk score to reuse % and average
        reuse_values = [_similarity_to_reuse(s) for s in ch_scores]
        avg_reuse = sum(reuse_values) / len(reuse_values)

        # Also report raw average similarity for transparency
        avg_similarity = sum(ch_scores) / len(ch_scores)

        final_pct = round(avg_reuse, 1)
        reused_words = int(ch_word_count * final_pct / 100)
        total_reused_words += reused_words

        chapter_stats.append({
            "chapter_id": str(ch.id),
            "title": ch.title,
            "numbering": ch.numbering or "",
            "word_count": ch_word_count,
            "reuse_percentage": final_pct,
            "semantic_similarity": round(avg_similarity * 100, 1),
            "ngram_match": final_pct,  # Keep for frontend compatibility
            "sequence_match": round(avg_similarity * 100, 1),  # Keep for frontend compat
        })

    overall_pct = round(total_reused_words / total_new_words * 100, 1) if total_new_words else 0
    chapters_with_reuse = sum(1 for c in chapter_stats if c["reuse_percentage"] > 10)
    avg_reuse = round(
        sum(c["reuse_percentage"] for c in chapter_stats) / len(chapter_stats), 1
    ) if chapter_stats else 0

    result_data = {
        "has_old_response": True,
        "overall_reuse_percentage": overall_pct,
        "chapters": chapter_stats,
        "summary": {
            "total_chapters": len(chapter_stats),
            "chapters_with_reuse": chapters_with_reuse,
            "avg_reuse_percentage": avg_reuse,
            "old_response_word_count": old_word_count,
            "new_content_word_count": total_new_words,
        },
    }

    # Persist the results to database
    try:
        reuse_result = ContentReuseResult(
            project_id=project_id,
            has_old_response=True,
            overall_reuse_percentage=overall_pct,
            chapters=chapter_stats,
            summary=result_data["summary"],
        )
        db.add(reuse_result)
        await db.commit()
        result_data["created_at"] = reuse_result.created_at.isoformat() if reuse_result.created_at else None
    except Exception as exc:
        logger.warning("Failed to persist content reuse results: %s", exc)
        await db.rollback()

    return result_data


@router.get("/{project_id}/content-reuse-stats/latest")
async def get_content_reuse_stats_latest(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return the most recently persisted content reuse analysis, or null."""
    result = await db.execute(
        select(ContentReuseResult)
        .where(ContentReuseResult.project_id == project_id)
        .order_by(ContentReuseResult.created_at.desc())
        .limit(1)
    )
    cr = result.scalar_one_or_none()
    if not cr:
        return {"result": None}
    return {
        "result": {
            "has_old_response": cr.has_old_response,
            "overall_reuse_percentage": cr.overall_reuse_percentage,
            "chapters": cr.chapters or [],
            "summary": cr.summary or {},
            "created_at": cr.created_at.isoformat() if cr.created_at else None,
        }
    }


# ── AI Cost Tracking ───────────────────────────────────────────────

@router.get("/{project_id}/ai-cost-tracking")
async def get_ai_cost_tracking(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get AI usage logs and cost summary for a project (admin only)."""
    from ..models.user import UserRole
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    # Load AI usage logs from the project
    from ..models.project import AIUsageLog
    result = await db.execute(
        select(AIUsageLog)
        .where(AIUsageLog.project_id == project_id)
        .order_by(AIUsageLog.created_at.desc())
    )
    logs = result.scalars().all()

    # Load pricing config
    from ..models.project import AIModelPricing
    pricing_result = await db.execute(select(AIModelPricing).order_by(AIModelPricing.provider, AIModelPricing.model_name))
    pricing_rows = pricing_result.scalars().all()
    pricing = [
        {
            "id": str(p.id),
            "provider": p.provider,
            "model_name": p.model_name,
            "price_per_1k_input": p.price_per_1k_input,
            "price_per_1k_output": p.price_per_1k_output,
            "currency": p.currency,
        }
        for p in pricing_rows
    ]

    # Build pricing lookup
    pricing_map = {}
    for p in pricing_rows:
        pricing_map[(p.provider, p.model_name)] = p

    # Build daily summary
    daily_summary: dict = {}
    total_input_tokens = 0
    total_output_tokens = 0
    total_cost = 0.0
    by_model: dict = {}

    for log in logs:
        day = log.created_at.strftime("%Y-%m-%d")
        if day not in daily_summary:
            daily_summary[day] = {"date": day, "input_tokens": 0, "output_tokens": 0, "cost": 0.0, "requests": 0}

        daily_summary[day]["input_tokens"] += log.input_tokens
        daily_summary[day]["output_tokens"] += log.output_tokens
        daily_summary[day]["requests"] += 1
        total_input_tokens += log.input_tokens
        total_output_tokens += log.output_tokens

        # Compute cost
        p = pricing_map.get((log.provider, log.model_name))
        cost = 0.0
        if p:
            cost = (log.input_tokens / 1000 * p.price_per_1k_input) + (log.output_tokens / 1000 * p.price_per_1k_output)
        daily_summary[day]["cost"] += cost
        total_cost += cost

        model_key = f"{log.provider}/{log.model_name}"
        if model_key not in by_model:
            by_model[model_key] = {"provider": log.provider, "model": log.model_name, "input_tokens": 0, "output_tokens": 0, "cost": 0.0, "requests": 0}
        by_model[model_key]["input_tokens"] += log.input_tokens
        by_model[model_key]["output_tokens"] += log.output_tokens
        by_model[model_key]["cost"] += cost
        by_model[model_key]["requests"] += 1

    # Round costs
    for v in daily_summary.values():
        v["cost"] = round(v["cost"], 4)
    for v in by_model.values():
        v["cost"] = round(v["cost"], 4)

    return {
        "total_input_tokens": total_input_tokens,
        "total_output_tokens": total_output_tokens,
        "total_cost": round(total_cost, 4),
        "total_requests": len(logs),
        "daily": sorted(daily_summary.values(), key=lambda x: x["date"]),
        "by_model": list(by_model.values()),
        "pricing": pricing,
        "recent_logs": [
            {
                "id": str(log.id),
                "operation": log.operation,
                "provider": log.provider,
                "model_name": log.model_name,
                "input_tokens": log.input_tokens,
                "output_tokens": log.output_tokens,
                "created_at": log.created_at.isoformat(),
            }
            for log in logs[:100]  # Last 100 logs
        ],
    }


@router.put("/{project_id}/ai-pricing")
async def update_ai_pricing(
    project_id: uuid.UUID,
    request: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update AI model pricing (admin only)."""
    from ..models.user import UserRole
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    from ..models.project import AIModelPricing

    pricing_list = request.get("pricing", [])
    for item in pricing_list:
        if item.get("id"):
            # Update existing
            result = await db.execute(select(AIModelPricing).where(AIModelPricing.id == uuid.UUID(item["id"])))
            row = result.scalar_one_or_none()
            if row:
                row.price_per_1k_input = item.get("price_per_1k_input", row.price_per_1k_input)
                row.price_per_1k_output = item.get("price_per_1k_output", row.price_per_1k_output)
        else:
            # Create new
            new_pricing = AIModelPricing(
                provider=item["provider"],
                model_name=item["model_name"],
                price_per_1k_input=item.get("price_per_1k_input", 0),
                price_per_1k_output=item.get("price_per_1k_output", 0),
                currency=item.get("currency", "EUR"),
            )
            db.add(new_pricing)

    await db.commit()
    return {"status": "ok"}


@router.delete("/{project_id}/ai-pricing/{pricing_id}")
async def delete_ai_pricing(
    project_id: uuid.UUID,
    pricing_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete an AI pricing entry (admin only)."""
    from ..models.user import UserRole
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    from ..models.project import AIModelPricing
    result = await db.execute(select(AIModelPricing).where(AIModelPricing.id == pricing_id))
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Pricing non trouvé")
    await db.delete(row)
    await db.commit()
    return {"status": "ok"}


# ── Public pricing catalog ─────────────────────────────────────────

# Comprehensive catalog of public AI model pricing (EUR per 1K tokens)
# Sources: official provider pricing pages, converted to EUR where needed (~0.92 EUR/USD)
PUBLIC_PRICING_CATALOG = [
    # ── Mistral AI (Cloud) ──
    {"provider": "mistral", "model_name": "mistral-large-latest", "price_per_1k_input": 0.0018, "price_per_1k_output": 0.0055},
    {"provider": "mistral", "model_name": "mistral-medium-latest", "price_per_1k_input": 0.0025, "price_per_1k_output": 0.0075},
    {"provider": "mistral", "model_name": "mistral-small-latest", "price_per_1k_input": 0.0002, "price_per_1k_output": 0.0006},
    {"provider": "mistral", "model_name": "open-mistral-nemo", "price_per_1k_input": 0.00015, "price_per_1k_output": 0.00015},
    {"provider": "mistral", "model_name": "codestral-latest", "price_per_1k_input": 0.0003, "price_per_1k_output": 0.0009},
    {"provider": "mistral", "model_name": "pixtral-large-latest", "price_per_1k_input": 0.0018, "price_per_1k_output": 0.0055},
    {"provider": "mistral", "model_name": "pixtral-12b-2409", "price_per_1k_input": 0.00015, "price_per_1k_output": 0.00015},
    # ── Ollama (Local) – coût 0 par défaut ──
    {"provider": "ollama", "model_name": "mistral:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "mistral-nemo:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "mixtral:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "llama3.1:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "llama3.1:70b", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "qwen2.5:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "qwen2.5:14b", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "qwen2.5:32b", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "gemma3:12b", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "deepseek-r1:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "command-r:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "llama3.2-vision:11b", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "llama3.2-vision:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "llava:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "llava:13b", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    {"provider": "ollama", "model_name": "bakllava:latest", "price_per_1k_input": 0.0, "price_per_1k_output": 0.0},
    # ── Scaleway (EU API) ──
    {"provider": "scaleway", "model_name": "mistral-large-3-675b-instruct-2512", "price_per_1k_input": 0.002, "price_per_1k_output": 0.006},
    {"provider": "scaleway", "model_name": "mistral-small-3.2-24b-instruct-2506", "price_per_1k_input": 0.0002, "price_per_1k_output": 0.0006},
    {"provider": "scaleway", "model_name": "mistral-small-3.1-24b-instruct-2503", "price_per_1k_input": 0.0002, "price_per_1k_output": 0.0006},
    {"provider": "scaleway", "model_name": "llama-3.3-70b-instruct", "price_per_1k_input": 0.00035, "price_per_1k_output": 0.0008},
    {"provider": "scaleway", "model_name": "qwen2.5-coder-32b-instruct", "price_per_1k_input": 0.0003, "price_per_1k_output": 0.0006},
    {"provider": "scaleway", "model_name": "pixtral-12b-2409", "price_per_1k_input": 0.00015, "price_per_1k_output": 0.00015},
    # ── OpenAI ──
    {"provider": "openai", "model_name": "gpt-4o", "price_per_1k_input": 0.0023, "price_per_1k_output": 0.0092},
    {"provider": "openai", "model_name": "gpt-4o-mini", "price_per_1k_input": 0.000138, "price_per_1k_output": 0.00055},
    {"provider": "openai", "model_name": "gpt-4-turbo", "price_per_1k_input": 0.0092, "price_per_1k_output": 0.0276},
    {"provider": "openai", "model_name": "gpt-4", "price_per_1k_input": 0.0276, "price_per_1k_output": 0.055},
    {"provider": "openai", "model_name": "gpt-3.5-turbo", "price_per_1k_input": 0.00046, "price_per_1k_output": 0.00138},
    {"provider": "openai", "model_name": "o1", "price_per_1k_input": 0.0138, "price_per_1k_output": 0.055},
    {"provider": "openai", "model_name": "o1-mini", "price_per_1k_input": 0.00276, "price_per_1k_output": 0.011},
    {"provider": "openai", "model_name": "o3-mini", "price_per_1k_input": 0.001, "price_per_1k_output": 0.004},
    # ── Anthropic ──
    {"provider": "anthropic", "model_name": "claude-opus-4", "price_per_1k_input": 0.0138, "price_per_1k_output": 0.069},
    {"provider": "anthropic", "model_name": "claude-sonnet-4", "price_per_1k_input": 0.00276, "price_per_1k_output": 0.0138},
    {"provider": "anthropic", "model_name": "claude-3.5-sonnet", "price_per_1k_input": 0.00276, "price_per_1k_output": 0.0138},
    {"provider": "anthropic", "model_name": "claude-3.5-haiku", "price_per_1k_input": 0.00074, "price_per_1k_output": 0.0037},
    {"provider": "anthropic", "model_name": "claude-3-opus", "price_per_1k_input": 0.0138, "price_per_1k_output": 0.069},
    {"provider": "anthropic", "model_name": "claude-3-haiku", "price_per_1k_input": 0.000230, "price_per_1k_output": 0.00115},
    # ── Google ──
    {"provider": "google", "model_name": "gemini-2.0-flash", "price_per_1k_input": 0.000069, "price_per_1k_output": 0.000368},
    {"provider": "google", "model_name": "gemini-2.0-flash-lite", "price_per_1k_input": 0.000069, "price_per_1k_output": 0.000276},
    {"provider": "google", "model_name": "gemini-1.5-pro", "price_per_1k_input": 0.00115, "price_per_1k_output": 0.0046},
    {"provider": "google", "model_name": "gemini-1.5-flash", "price_per_1k_input": 0.000069, "price_per_1k_output": 0.000276},
    # ── DeepSeek ──
    {"provider": "deepseek", "model_name": "deepseek-chat", "price_per_1k_input": 0.000253, "price_per_1k_output": 0.001104},
    {"provider": "deepseek", "model_name": "deepseek-reasoner", "price_per_1k_input": 0.000506, "price_per_1k_output": 0.002116},
    # ── Cohere ──
    {"provider": "cohere", "model_name": "command-r-plus", "price_per_1k_input": 0.00276, "price_per_1k_output": 0.0138},
    {"provider": "cohere", "model_name": "command-r", "price_per_1k_input": 0.000138, "price_per_1k_output": 0.000552},
]


@router.get("/{project_id}/ai-pricing/catalog")
async def get_public_pricing_catalog(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Return the public pricing catalog for all known AI models (admin only)."""
    from ..models.user import UserRole
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    return {"catalog": PUBLIC_PRICING_CATALOG}


@router.post("/{project_id}/ai-pricing/load-public")
async def load_public_pricing(
    project_id: uuid.UUID,
    request: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Load public pricing for selected models. Only adds models that don't already exist.
    Body: { "models": [ { "provider": "...", "model_name": "..." }, ... ] }
    If "models" is empty or absent, loads ALL catalog entries.
    """
    from ..models.user import UserRole
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    from ..models.project import AIModelPricing

    requested = request.get("models", [])

    # Build lookup of what's in the catalog
    catalog_map = {(e["provider"], e["model_name"]): e for e in PUBLIC_PRICING_CATALOG}

    # Determine which entries to load
    if requested:
        entries_to_load = []
        for m in requested:
            key = (m["provider"], m["model_name"])
            if key in catalog_map:
                entries_to_load.append(catalog_map[key])
    else:
        entries_to_load = list(PUBLIC_PRICING_CATALOG)

    # Load existing pricing to avoid duplicates
    existing_result = await db.execute(select(AIModelPricing))
    existing_rows = existing_result.scalars().all()
    existing_keys = {(r.provider, r.model_name) for r in existing_rows}

    added = 0
    for entry in entries_to_load:
        key = (entry["provider"], entry["model_name"])
        if key not in existing_keys:
            new_pricing = AIModelPricing(
                provider=entry["provider"],
                model_name=entry["model_name"],
                price_per_1k_input=entry["price_per_1k_input"],
                price_per_1k_output=entry["price_per_1k_output"],
                currency="EUR",
            )
            db.add(new_pricing)
            existing_keys.add(key)
            added += 1

    await db.commit()
    return {"status": "ok", "added": added, "total_catalog": len(entries_to_load)}


# ── Carbon estimation (ADEME methodology) ──────────────────────────

# Energy consumption per 1K tokens (Wh) — estimated by model class
# Based on GPU inference benchmarks (A100/H100) and published energy studies
_ENERGY_WH_PER_1K_TOKENS = {
    "small": 0.4,    # ≤13B params (Mistral Small, Haiku, GPT-3.5, Gemma, etc.)
    "medium": 1.0,   # 14–34B params (Codestral, Qwen 32B, Command R, etc.)
    "large": 2.5,    # 35–80B params (Llama 70B, Mistral Large, GPT-4o, etc.)
    "xlarge": 6.0,   # >80B params (GPT-4, Claude Opus, Mistral 675B, etc.)
}

# Carbon intensity of electricity grid by provider location (gCO2eq/kWh)
# Source: ADEME Base Carbone 2024, IEA 2023
_CARBON_INTENSITY = {
    "mistral": 56,      # France (nuclear-dominated grid)
    "scaleway": 56,     # France
    "ollama": 56,       # Local — default France
    "openai": 380,      # US average (Azure data centers)
    "anthropic": 380,   # US (AWS/GCP)
    "google": 120,      # Mixed global, ~80% renewable pledge
    "deepseek": 580,    # China average
    "cohere": 150,      # Canada/US mix
}

# PUE (Power Usage Effectiveness) — ADEME recommends 1.2 for modern DC
_PUE = 1.2

# Water usage per kWh of cooling (L/kWh) — ADEME/WRI estimates
_WATER_L_PER_KWH = 1.8


def _model_size_class(provider: str, model_name: str) -> str:
    """Classify a model into a size class for energy estimation."""
    name = model_name.lower()
    # Extra-large models (>80B)
    if any(k in name for k in ["gpt-4-turbo", "gpt-4", "opus", "675b", "mixtral"]):
        if "mini" in name or "small" in name:
            return "small"
        return "xlarge"
    # Large models (35-80B)
    if any(k in name for k in ["large", "70b", "gpt-4o", "o1", "pro", "command-r-plus"]):
        if "mini" in name:
            return "medium"
        return "large"
    # Medium models (14-34B)
    if any(k in name for k in ["medium", "codestral", "32b", "24b", "14b", "13b", "reasoner"]):
        return "medium"
    # Default: small
    return "small"


@router.get("/{project_id}/ai-carbon-tracking")
async def get_ai_carbon_tracking(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Estimate carbon footprint of AI usage based on ADEME methodology (admin only)."""
    from ..models.user import UserRole
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")

    from ..models.project import AIUsageLog
    result = await db.execute(
        select(AIUsageLog)
        .where(AIUsageLog.project_id == project_id)
        .order_by(AIUsageLog.created_at.desc())
    )
    logs = result.scalars().all()

    total_energy_wh = 0.0
    total_co2_g = 0.0
    total_water_l = 0.0
    by_provider: dict = {}
    by_model: dict = {}
    daily: dict = {}

    for log in logs:
        size_class = _model_size_class(log.provider, log.model_name)
        total_tokens_k = (log.input_tokens + log.output_tokens) / 1000.0
        energy_wh = total_tokens_k * _ENERGY_WH_PER_1K_TOKENS.get(size_class, 0.4) * _PUE
        carbon_intensity = _CARBON_INTENSITY.get(log.provider, 300)
        co2_g = energy_wh / 1000.0 * carbon_intensity  # convert Wh→kWh then multiply
        water_l = energy_wh / 1000.0 * _WATER_L_PER_KWH

        total_energy_wh += energy_wh
        total_co2_g += co2_g
        total_water_l += water_l

        # By provider
        if log.provider not in by_provider:
            by_provider[log.provider] = {"provider": log.provider, "energy_wh": 0, "co2_g": 0, "water_l": 0, "tokens": 0, "requests": 0}
        by_provider[log.provider]["energy_wh"] += energy_wh
        by_provider[log.provider]["co2_g"] += co2_g
        by_provider[log.provider]["water_l"] += water_l
        by_provider[log.provider]["tokens"] += log.input_tokens + log.output_tokens
        by_provider[log.provider]["requests"] += 1

        # By model
        model_key = f"{log.provider}/{log.model_name}"
        if model_key not in by_model:
            by_model[model_key] = {
                "provider": log.provider, "model": log.model_name, "size_class": size_class,
                "energy_wh": 0, "co2_g": 0, "water_l": 0, "tokens": 0, "requests": 0,
            }
        by_model[model_key]["energy_wh"] += energy_wh
        by_model[model_key]["co2_g"] += co2_g
        by_model[model_key]["water_l"] += water_l
        by_model[model_key]["tokens"] += log.input_tokens + log.output_tokens
        by_model[model_key]["requests"] += 1

        # Daily
        day = log.created_at.strftime("%Y-%m-%d")
        if day not in daily:
            daily[day] = {"date": day, "energy_wh": 0, "co2_g": 0, "water_l": 0, "tokens": 0, "requests": 0}
        daily[day]["energy_wh"] += energy_wh
        daily[day]["co2_g"] += co2_g
        daily[day]["water_l"] += water_l
        daily[day]["tokens"] += log.input_tokens + log.output_tokens
        daily[day]["requests"] += 1

    # Round values
    def _round_entry(e: dict) -> dict:
        e["energy_wh"] = round(e["energy_wh"], 2)
        e["co2_g"] = round(e["co2_g"], 2)
        e["water_l"] = round(e["water_l"], 3)
        return e

    # Equivalences ADEME pour vulgarisation
    co2_kg = total_co2_g / 1000.0
    equivalences = {
        "km_voiture": round(co2_kg / 0.218, 1),        # ADEME: 218 gCO2/km voiture moyenne
        "heures_streaming": round(co2_kg / 0.036, 1),   # ~36 gCO2/h streaming vidéo
        "emails": round(co2_kg / 0.004, 0),              # ~4 gCO2/email (ADEME)
        "charges_smartphone": round(co2_kg / 0.008, 0),  # ~8 gCO2/charge
        "litres_eau": round(total_water_l, 1),
    }

    return {
        "total_energy_wh": round(total_energy_wh, 2),
        "total_co2_g": round(total_co2_g, 2),
        "total_water_l": round(total_water_l, 3),
        "total_requests": len(logs),
        "total_tokens": sum(log.input_tokens + log.output_tokens for log in logs),
        "equivalences": equivalences,
        "by_provider": [_round_entry(v) for v in by_provider.values()],
        "by_model": sorted([_round_entry(v) for v in by_model.values()], key=lambda x: x["co2_g"], reverse=True),
        "daily": sorted([_round_entry(v) for v in daily.values()], key=lambda x: x["date"]),
        "methodology": {
            "source": "ADEME Base Carbone 2024 / IEA 2023",
            "pue": _PUE,
            "water_l_per_kwh": _WATER_L_PER_KWH,
            "carbon_intensities": _CARBON_INTENSITY,
            "energy_per_1k_tokens_wh": _ENERGY_WH_PER_1K_TOKENS,
        },
    }


# ── Project Members ─────────────────────────────────────────────────

@router.get("/{project_id}/members")
async def list_project_members(
    project_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List project members and workspace members with source indication."""
    from ..models.user import UserRole

    project_result = await db.execute(
        select(RFPProject).where(RFPProject.id == project_id)
    )
    project = project_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Check access: must be project member or admin
    await get_project_membership(project_id, current_user, db)

    # Get project-specific members
    result = await db.execute(
        select(ProjectMember, User)
        .join(User, User.id == ProjectMember.user_id)
        .where(ProjectMember.project_id == project_id)
    )
    project_rows = result.all()
    project_member_user_ids = {user.id for _, user in project_rows}

    # Determine if current user can manage members (owner or admin)
    is_admin = current_user.role == UserRole.ADMIN
    current_pm = await db.execute(
        select(ProjectMember)
        .where(ProjectMember.project_id == project_id)
        .where(ProjectMember.user_id == current_user.id)
    )
    current_membership = current_pm.scalar_one_or_none()
    can_manage = is_admin or (current_membership and current_membership.role == "owner")

    members = [
        {
            "id": str(pm.id),
            "user_id": str(user.id),
            "username": user.username,
            "email": user.email,
            "full_name": user.full_name,
            "role": pm.role,
            "joined_at": pm.joined_at.isoformat(),
            "source": "project",
        }
        for pm, user in project_rows
    ]

    # Also return workspace members (not already project members) for reference
    # Only if current user can manage members (so they can add them)
    if can_manage:
        ws_result = await db.execute(
            select(WorkspaceMember, User)
            .join(User, User.id == WorkspaceMember.user_id)
            .where(WorkspaceMember.workspace_id == project.workspace_id)
        )
        ws_rows = ws_result.all()
        for wm, user in ws_rows:
            if user.id not in project_member_user_ids:
                members.append({
                    "id": str(wm.id),
                    "user_id": str(user.id),
                    "username": user.username,
                    "email": user.email,
                    "full_name": user.full_name,
                    "role": wm.role.value,
                    "joined_at": wm.joined_at.isoformat(),
                    "source": "workspace",
                })

    return members


@router.post("/{project_id}/members", status_code=status.HTTP_201_CREATED)
async def add_project_member(
    project_id: uuid.UUID,
    request: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Add a member to a project (project owner or admin only). The user must be a workspace member first."""
    await require_project_owner_or_admin(project_id, current_user, db)

    user_id = request.get("user_id")
    role = request.get("role", "editor")

    # Check user exists
    user_result = await db.execute(select(User).where(User.id == uuid.UUID(user_id)))
    user = user_result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")

    # Get project to find workspace_id
    proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    proj = proj_result.scalar_one_or_none()
    if not proj:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    # Verify user is a workspace member
    ws_member = await db.execute(
        select(WorkspaceMember)
        .where(WorkspaceMember.workspace_id == proj.workspace_id)
        .where(WorkspaceMember.user_id == uuid.UUID(user_id))
    )
    if not ws_member.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="L'utilisateur doit d'abord être membre de l'espace de travail")

    # Check not already member
    existing = await db.execute(
        select(ProjectMember)
        .where(ProjectMember.project_id == project_id)
        .where(ProjectMember.user_id == uuid.UUID(user_id))
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Utilisateur déjà membre du projet")

    member = ProjectMember(
        project_id=project_id,
        user_id=uuid.UUID(user_id),
        role=role if role in ("owner", "editor", "viewer") else "editor",
    )
    db.add(member)
    await db.commit()
    return {"success": True, "message": "Membre ajouté au projet"}


@router.put("/{project_id}/members/{user_id}")
async def update_project_member_role(
    project_id: uuid.UUID,
    user_id: uuid.UUID,
    request: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update a project member's role (project owner or admin only)."""
    await require_project_owner_or_admin(project_id, current_user, db)

    result = await db.execute(
        select(ProjectMember)
        .where(ProjectMember.project_id == project_id)
        .where(ProjectMember.user_id == user_id)
    )
    member = result.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Membre non trouvé")

    role = request.get("role", "editor")
    if role not in ("owner", "editor", "viewer"):
        raise HTTPException(status_code=400, detail="Rôle invalide")

    member.role = role
    await db.commit()

    user_result = await db.execute(select(User).where(User.id == user_id))
    user = user_result.scalar_one()
    return {
        "id": str(member.id),
        "user_id": str(user.id),
        "username": user.username,
        "email": user.email,
        "full_name": user.full_name,
        "role": member.role,
        "joined_at": member.joined_at.isoformat(),
    }


@router.delete("/{project_id}/members/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
async def remove_project_member(
    project_id: uuid.UUID,
    user_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Remove a member from a project (project owner or admin only)."""
    await require_project_owner_or_admin(project_id, current_user, db)

    result = await db.execute(
        select(ProjectMember)
        .where(ProjectMember.project_id == project_id)
        .where(ProjectMember.user_id == user_id)
    )
    member = result.scalar_one_or_none()
    if not member:
        raise HTTPException(status_code=404, detail="Membre non trouvé")

    await db.delete(member)
    await db.commit()


# ── Source document matching (shared by fill-excel and fill-pdf) ────

def _match_source_document(resp_doc, candidates: list):
    """Find the best matching uploaded source document for a ResponseDocument.

    Scores every candidate and returns the best match (or None).
    This replaces the old "first keyword wins" approach that would pick
    the wrong file when multiple documents shared a generic keyword
    like "bordereau".
    """
    if not candidates:
        return None

    doc_title_lower = (resp_doc.title or "").lower()
    rfp_source_lower = (resp_doc.rfp_source or "").lower()

    # Extract meaningful words from the deliverable title (length > 2)
    title_words = [w for w in re.split(r'[\s\-_/(),.]+', doc_title_lower) if len(w) > 2]

    best_doc = None
    best_score = 0

    for doc in candidates:
        fname_lower = (doc.original_filename or "").lower()
        # Strip extension for matching
        fname_stem = re.sub(r'\.\w{2,4}$', '', fname_lower)
        score = 0

        # Strong signal: rfp_source substring match (bidirectional)
        if rfp_source_lower and (rfp_source_lower in fname_lower or fname_lower in rfp_source_lower):
            score += 100

        # Count how many title words appear in the filename
        word_matches = sum(1 for w in title_words if w in fname_stem)
        score += word_matches * 10

        # Bonus for discriminating words (lot numbers, specific terms)
        for w in title_words:
            if w in fname_stem and re.match(r'^(lot\d*|\d+|tjm|bpu|dqe|dpgf|qds|dc\d|attri\d?)$', w):
                score += 15  # discriminating identifiers get extra weight

        if score > best_score:
            best_score = score
            best_doc = doc

    if best_doc:
        logger.info(
            "Source match: '%s' -> '%s' (score=%d)",
            resp_doc.title, best_doc.original_filename, best_score,
        )

    return best_doc if best_score > 0 else None


# ── Fill Excel endpoint ─────────────────────────────────────────────

def _ensure_xlsx_path(file_path: str) -> str:
    """Ensure the file at file_path can be opened by openpyxl.

    Handles two cases:
    1. True .xls file (OLE2 format): convert to .xlsx using xlrd
    2. .xlsx content with .xls extension: create a copy with .xlsx extension
       (openpyxl rejects files based on extension before reading content)

    Returns a path to a valid .xlsx file.
    """
    with open(file_path, "rb") as f:
        magic = f.read(4)

    has_xls_extension = file_path.lower().endswith('.xls') and not file_path.lower().endswith('.xlsx')

    if magic == b'\xd0\xcf\x11\xe0':
        # True OLE2/.xls file — convert via xlrd
        import xlrd
        from openpyxl import Workbook

        xls_book = xlrd.open_workbook(file_path)
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_index in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_index)
            ws = wb.create_sheet(title=xls_sheet.name)
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                    cell_type = xls_sheet.cell_type(row_idx, col_idx)
                    if cell_type == xlrd.XL_CELL_DATE:
                        try:
                            date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                            from datetime import datetime
                            cell_value = datetime(*date_tuple)
                        except Exception:
                            pass
                    elif cell_type == xlrd.XL_CELL_BOOLEAN:
                        cell_value = bool(cell_value)
                    elif cell_type == xlrd.XL_CELL_EMPTY:
                        continue
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

        xlsx_path = file_path + "x"
        wb.save(xlsx_path)
        wb.close()
        return xlsx_path

    if has_xls_extension:
        # File has .xls extension but is actually xlsx content (ZIP/PK format).
        # openpyxl validates the extension, so copy/link with .xlsx extension.
        xlsx_path = file_path + "x"
        if not os.path.exists(xlsx_path):
            shutil.copy2(file_path, xlsx_path)
        return xlsx_path

    return file_path


def _extract_excel_search_queries(excel_structure: str, doc_title: str) -> list[str]:
    """Extract targeted search queries from Excel structure for multi-query RAG.

    Parses the Excel text representation to find question cells and groups them
    into thematic search queries so that KPIs (turn over, taux de formation,
    handicap, etc.) are individually searched for in the vector store.
    """
    import re as _re

    # Always include a broad query with the document title
    queries = [doc_title]

    # Thematic groups: map keywords found in Excel questions → targeted search query
    thematic_queries = {
        "environnement": (
            {"environnement", "carbone", "co2", "émission", "énergie", "déchet",
             "recyclage", "iso 14001", "iso 50001", "bilan carbone", "empreinte"},
            "bilan carbone émissions CO2 énergie déchets recyclage ISO 14001 ISO 50001 environnement"
        ),
        "social": (
            {"formation", "turn over", "turnover", "absentéisme", "handicap",
             "diversité", "égalité", "qvt", "qualité de vie", "salari",
             "iso 45001", "effectif", "masse salariale"},
            "taux formation professionnelle turn over absentéisme handicapés effectif masse salariale heures formation"
        ),
        "social_kpi": (
            {"taux", "pourcentage", "index", "nombre"},
            "taux turn over absentéisme formation handicap index égalité hommes femmes pourcentage salariés"
        ),
        "ethique": (
            {"éthique", "corruption", "anti-corruption", "lanceur d'alerte",
             "sapin", "devoir de vigilance", "droits humains"},
            "éthique anti-corruption lanceur alerte Sapin II devoir vigilance droits humains"
        ),
        "rse_general": (
            {"rse", "responsabilité sociétale", "développement durable",
             "parties prenantes", "charte", "global compact", "pacte mondial"},
            "RSE charte responsabilité sociétale Pacte Mondial Global Compact parties prenantes développement durable"
        ),
        "certifications": (
            {"certif", "label", "norme", "iso"},
            "certifications labels ISO 14001 ISO 50001 ISO 45001 ISO 27001 normes"
        ),
        "donnees_chiffrees": (
            {"chiffre d'affaires", "ca ", "effectif", "collaborateur"},
            "chiffre affaires effectif collaborateurs nombre salariés résultat"
        ),
    }

    struct_lower = excel_structure.lower()
    for _name, (keywords, query) in thematic_queries.items():
        if any(kw in struct_lower for kw in keywords):
            queries.append(query)

    # Always add a [TABLEAU] search to find structured table data with KPIs
    queries.append("[TABLEAU] turn over handicap formation effectif")

    # Deduplicate while preserving order
    seen: set[str] = set()
    unique: list[str] = []
    for q in queries:
        if q not in seen:
            seen.add(q)
            unique.append(q)

    logger.info("Excel fill: generated %d search queries from structure", len(unique))
    return unique


def _read_excel_structure(file_path: str) -> str:
    """Read an Excel file and return a textual representation of its structure with cell references.
    Skips fully empty rows and only marks empty cells adjacent to filled cells to reduce noise."""
    from openpyxl import load_workbook
    file_path = _ensure_xlsx_path(file_path)
    wb = load_workbook(file_path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"\n=== Onglet: {sheet_name} ===")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_cells = []
            has_value = False
            for cell in row:
                coord = cell.coordinate
                val = cell.value
                if val is not None:
                    has_value = True
                    row_cells.append(f"{coord}={val}")
                else:
                    row_cells.append(f"{coord}=(vide)")
            # Only include rows that have at least one non-empty cell
            if has_value:
                parts.append(" | ".join(row_cells))
    wb.close()
    return "\n".join(parts)


def _fill_excel_with_data(file_path: str, fill_data: list) -> bytes:
    """Open an Excel file, fill cells from AI-generated data, return modified bytes."""
    from openpyxl import load_workbook
    import re as _re
    file_path = _ensure_xlsx_path(file_path)
    wb = load_workbook(file_path)

    for entry in fill_data:
        sheet_name = entry.get("sheet", "")
        cell_ref = entry.get("cell", "")
        value = entry.get("value")

        if not sheet_name or not cell_ref or value is None:
            continue
        if isinstance(value, str) and value.strip() == "[A COMPLÉTER]":
            continue  # skip placeholder values

        # Find the sheet (try exact match, then case-insensitive)
        ws = None
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            for sn in wb.sheetnames:
                if sn.lower() == sheet_name.lower():
                    ws = wb[sn]
                    break
        if ws is None:
            continue

        # Validate cell reference format
        if not _re.match(r'^[A-Z]{1,3}[0-9]+$', cell_ref.upper()):
            continue

        try:
            ws[cell_ref.upper()] = value
        except Exception:
            continue

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return output.read()


@router.post("/{project_id}/fill-excel/{doc_id}")
async def fill_excel_document(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch async Excel fill as a background task (returns immediately)."""
    result = await db.execute(
        select(ResponseDocument)
        .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
    )
    resp_doc = result.scalar_one_or_none()
    if not resp_doc:
        raise HTTPException(status_code=404, detail="Document livrable non trouvé")

    proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = proj_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    task_key = str(doc_id)
    existing = get_or_idle(_NS_FILL_EXCEL, task_key)
    if existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Remplissage deja en cours pour ce document")

    set_progress(_NS_FILL_EXCEL, task_key, {
        "status": "running", "step": "queued", "progress": 0,
        "message": f"En file d'attente: {resp_doc.title}",
        "doc_title": resp_doc.title,
    })

    from ..tasks.project_tasks import fill_excel_task
    fill_excel_task.apply_async(
        args=(str(project_id), str(doc_id), str(project.workspace_id)),
        priority=1,
    )
    return {"success": True, "message": "Remplissage Excel lance en arriere-plan"}


@router.get("/{project_id}/fill-excel-status/{doc_id}")
async def get_fill_excel_status(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll progress of Excel fill task."""
    return get_or_idle(_NS_FILL_EXCEL, str(doc_id))


@router.get("/{project_id}/fill-excel-download/{doc_id}")
async def download_filled_excel(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Download the filled Excel file once the task is complete."""
    status = get_or_idle(_NS_FILL_EXCEL, str(doc_id))
    if status.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Fichier pas encore pret")

    file_path = status.get("file_path", "")
    filename = status.get("filename", "document_rempli.xlsx")
    if not file_path or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="Fichier genere introuvable")

    from fastapi.responses import FileResponse
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _run_fill_excel(project_id: uuid.UUID, doc_id: uuid.UUID, workspace_id: uuid.UUID):
    """Background task: generate a filled Excel file for a completion-type document."""
    from ..database import task_session
    task_key = str(doc_id)

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_FILL_EXCEL, task_key, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
        })

    try:
        # ── Phase 1: Load data (short DB session) ──
        async with task_session() as db:
            result = await db.execute(
                select(ResponseDocument)
                .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
            )
            resp_doc = result.scalar_one_or_none()
            if not resp_doc:
                set_progress(_NS_FILL_EXCEL, task_key, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Document livrable non trouve",
                })
                return

            doc_title = resp_doc.title or ""
            doc_source_ids = resp_doc.source_document_ids or []
            doc_source_cats = resp_doc.source_categories or []
            doc_include_generated = resp_doc.include_generated_content or False
            doc_custom_notes = resp_doc.custom_notes or ""
            has_custom_sources = bool(doc_source_ids or doc_source_cats or doc_include_generated)
            _update("loading", 5, f"Chargement: {doc_title}")

            docs_result = await db.execute(
                select(Document)
                .where(Document.project_id == project_id)
                .where(Document.category == DocumentCategory.NEW_RFP)
            )
            all_dce_docs = docs_result.scalars().all()

            excel_candidates = [
                doc for doc in all_dce_docs
                if doc.file_type.value in ("xlsx", "xls")
            ]
            excel_doc = _match_source_document(resp_doc, excel_candidates)
            if not excel_doc and excel_candidates:
                excel_doc = excel_candidates[0]

            if not excel_doc or not os.path.isfile(excel_doc.file_path):
                set_progress(_NS_FILL_EXCEL, task_key, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Aucun fichier Excel source trouve dans le DCE.",
                })
                return

            excel_file_path = excel_doc.file_path
            excel_original_filename = excel_doc.original_filename

        _update("reading", 10, "Lecture de la structure Excel...")
        excel_structure = _read_excel_structure(excel_file_path)

        # ── Phase 2: Load context (short DB session) ──
        _update("loading_context", 15, "Chargement du contexte AO + ancienne reponse...")
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            # If user selected specific sources (categories, docs, or generated content), load only those
            if has_custom_sources:
                parts = []
                if doc_source_cats:
                    parts.append(await _get_chunks_anonymized_by_categories(
                        db, project_id, doc_source_cats
                    ))
                if doc_source_ids:
                    parts.append(await _get_chunks_anonymized_by_document_ids(
                        db, project_id, doc_source_ids
                    ))
                if doc_include_generated:
                    parts.append(await _get_generated_chapters_context(db, project_id))
                anon_context = "\n\n".join(p for p in parts if p)
                anon_new_rfp = await _get_all_chunks_anonymized_by_category(
                    db, project_id, DocumentCategory.NEW_RFP
                )
                anon_old_response = ""
            else:
                # Load ALL source content (old_response + inspiration) — not just old_response
                anon_old_response_part, anon_inspiration_part, anon_new_rfp = await asyncio.gather(
                    _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.OLD_RESPONSE),
                    _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.INSPIRATION),
                    _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.NEW_RFP),
                )
                anon_old_response = "\n\n".join(p for p in [anon_old_response_part, anon_inspiration_part] if p)
                anon_context = None

        # ── Vector search: ALWAYS run to find the most relevant chunks ──
        # Even when user selected custom sources, we need vector search to
        # prioritize KPI-containing chunks that would otherwise be truncated
        # (e.g. table data on page 82 of a 90-page document gets cut at 50K chars).
        _update("searching", 25, "Recherche de contenu pertinent...")
        title_lower = doc_title.lower()
        conformity_keywords = ["rgpd", "conformit", "gdpr", "protection des données",
                               "questionnaire", "grille", "annexe", "déclaration",
                               "engagement", "certification", "audit", "sécurité",
                               "environnement", "rse", "social", "qualité"]
        is_conformity_doc = any(kw in title_lower for kw in conformity_keywords)

        from ..tasks.chapter_tasks import _hybrid_search

        # No category filter — search ALL source documents (old_response + inspiration)
        # This avoids the N queries × M categories explosion that was causing slowness.

        if is_conformity_doc:
            search_queries = _extract_excel_search_queries(excel_structure, doc_title)
            # Limit to 6 most relevant queries to keep search fast
            search_queries = search_queries[:8]
            seen_chunk_ids: set[str] = set()
            all_chunks: list[dict] = []
            for sq in search_queries:
                chunks = _hybrid_search(str(project_id), sq, top_k=10)
                for c in chunks:
                    cid = c.get("chunk_id", c.get("content", "")[:80])
                    if cid not in seen_chunk_ids:
                        seen_chunk_ids.add(cid)
                        all_chunks.append(c)
            all_chunks.sort(key=lambda c: c.get("score", 0), reverse=True)
            relevant_chunks = all_chunks[:40]
        else:
            search_query = (
                f"prix unitaire tarif {doc_title} BPU bordereau montant "
                "coût forfait taux journalier"
            )
            relevant_chunks = _hybrid_search(str(project_id), search_query, top_k=25)

        relevant_context = "\n\n".join([
            f"[{c['document_name']} p.{c['page_number']}] {c['content']}"
            for c in relevant_chunks
        ])

        # Log what vector search found for debugging KPI retrieval
        logger.info(
            "Excel fill: vector search returned %d unique chunks, relevant_context=%d chars",
            len(relevant_chunks), len(relevant_context),
        )
        for i, c in enumerate(relevant_chunks[:5]):
            preview = c['content'][:150].replace('\n', ' ')
            logger.info(
                "  chunk[%d] score=%.3f doc=%s p.%s: %s...",
                i, c.get('score', 0), c['document_name'], c['page_number'], preview,
            )

        # Build context: vector search results (most relevant) FIRST,
        # then full document content (may be truncated but that's OK —
        # the important data is already in the vector search results above).
        base_content = anon_context if anon_context is not None else anon_old_response
        if relevant_context:
            label = "EXTRAITS PERTINENTS" if is_conformity_doc else "EXTRAITS PERTINENTS SUR LES PRIX"
            old_response_with_context = (
                f"=== {label} (PRIORITÉ — CONTIENT LES DONNÉES CHIFFRÉES) ===\n{relevant_context}\n\n"
                f"=== CONTENU COMPLET DES DOCUMENTS SOURCE ===\n{base_content}"
            )
        else:
            old_response_with_context = base_content

        # ── Phase 3: AI generation (NO DB held) ──
        _update("generating", 30, f"Generation IA du contenu pour {doc_title}...")
        logger.info(
            "fill-excel %s: excel_structure=%d chars, old_response=%d chars, new_rfp=%d chars",
            doc_title, len(excel_structure), len(old_response_with_context), len(anon_new_rfp),
        )
        fill_data = await ai_service.generate_excel_fill_data(
            document_title=doc_title,
            excel_structure=excel_structure,
            new_rfp_content=anon_new_rfp,
            old_response_content=old_response_with_context,
            custom_notes=doc_custom_notes,
        )
        logger.info("fill-excel %s: AI returned %d cell entries", doc_title, len(fill_data))

        # Log AI usage
        async with task_session() as usage_db:
            await log_ai_usage_from_service(usage_db, project_id, "fill_excel", ai_service)

        # ── Phase 4: Fill Excel and save to disk ──
        _update("filling", 85, "Remplissage du fichier Excel...")
        filled_bytes = _fill_excel_with_data(excel_file_path, fill_data)

        base_name = os.path.splitext(excel_original_filename)[0]
        output_filename = f"{base_name}_rempli.xlsx"
        from ..config import settings
        output_path = os.path.join(settings.export_dir, f"{doc_id}_{output_filename}")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(filled_bytes)

        set_progress(_NS_FILL_EXCEL, task_key, {
            "status": "completed", "step": "done", "progress": 100,
            "message": f"Excel rempli: {output_filename} ({len(fill_data)} cellules)",
            "file_path": output_path,
            "filename": output_filename,
            "cell_count": len(fill_data),
        })

    except Exception as e:
        logger.exception("Fill Excel failed for doc %s", doc_id)
        set_progress(_NS_FILL_EXCEL, task_key, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })


# ── Fill PDF endpoint ──────────────────────────────────────────────

def _extract_pdf_zones(file_path: str) -> dict:
    """Extract PDF layout and identify fillable zones with real coordinates.

    Returns a dict with:
      - "has_form_fields": bool
      - "form_fields": list of form field dicts (for form PDFs)
      - "zones": list of fillable zone dicts (for non-form PDFs)
      - "text_for_ai": formatted text describing zones for AI consumption
    """
    import fitz
    import re
    doc = fitz.open(file_path)
    result = {"has_form_fields": False, "form_fields": [], "zones": [], "text_for_ai": ""}

    # ── 1. Check for interactive form fields ──
    for page_num in range(len(doc)):
        page = doc[page_num]
        for w in page.widgets() or []:
            result["has_form_fields"] = True
            result["form_fields"].append({
                "page": page_num + 1,
                "field_name": w.field_name or "",
                "field_type": w.field_type_string or "",
                "field_value": w.field_value or "",
            })

    if result["has_form_fields"]:
        # For form PDFs, we just list the fields for the AI
        parts = ["=== CHAMPS DE FORMULAIRE PDF ==="]
        for f in result["form_fields"]:
            parts.append(
                f"  Page {f['page']}: Champ '{f['field_name']}' "
                f"(type={f['field_type']}, valeur actuelle='{f['field_value']}')"
            )
        # Also add text context per page
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text("text")
            if text.strip():
                parts.append(f"\n=== TEXTE PAGE {page_num + 1} ===")
                parts.append(text[:5000])
        result["text_for_ai"] = "\n".join(parts)
        doc.close()
        return result

    # ── 2. For non-form PDFs: extract layout with zone detection ──
    # Pattern matching for filler characters (dots, underscores, etc.)
    _FILLER_RE = re.compile(r'[.…_]{3,}|(\.\s){3,}')

    zones = []
    ai_parts = []
    zone_id = 0

    # Track page-level right margin (the farthest content boundary)
    page_margins = {}

    for page_num in range(len(doc)):
        page = doc[page_num]
        page_width = page.rect.width
        page_height = page.rect.height
        text_dict = page.get_text("dict")
        page_zones = []

        # First pass: compute the effective right margin for the page
        max_right = 0
        for block in text_dict.get("blocks", []):
            if block.get("type") != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    txt = span.get("text", "").strip()
                    if txt and not _FILLER_RE.fullmatch(txt):
                        max_right = max(max_right, span["bbox"][2])
        # Use page width minus a small margin as fallback
        effective_right = min(page_width - 20, max(max_right + 30, page_width * 0.85))
        page_margins[page_num] = effective_right

        for block in text_dict.get("blocks", []):
            if block.get("type") != 0:  # text blocks only
                continue

            for line in block.get("lines", []):
                spans = line.get("spans", [])
                if not spans:
                    continue

                line_text = "".join(s["text"] for s in spans).strip()
                if not line_text:
                    continue

                line_bbox = line["bbox"]  # [x0, y0, x1, y1]
                first_span = spans[0]
                last_span = spans[-1]
                last_span_bbox = last_span["bbox"]
                first_span_bbox = first_span["bbox"]
                font_size = last_span.get("size", 10)

                # Compute the line height (distance from top to bottom of the line)
                line_height = line_bbox[3] - line_bbox[1]

                # ── Zone type B: line with dots/underscores (fill-in line) ──
                # Check this FIRST because these lines might also end with ":"
                has_filler = bool(_FILLER_RE.search(line_text))

                if has_filler:
                    # Find where the filler starts to position text AFTER the label part
                    # Walk through spans to find the first span containing filler chars
                    label_end_x = first_span_bbox[0]
                    fill_start_x = first_span_bbox[0]
                    filler_found = False
                    label_text = ""

                    for s in spans:
                        s_text = s.get("text", "")
                        if _FILLER_RE.search(s_text) and not filler_found:
                            filler_found = True
                            # Check if span has label text before the filler
                            match = _FILLER_RE.search(s_text)
                            if match and match.start() > 0:
                                # There's label text before the filler in this span
                                label_prefix = s_text[:match.start()].rstrip()
                                if label_prefix:
                                    label_text += label_prefix
                                    # Estimate X position where filler starts
                                    span_width = s["bbox"][2] - s["bbox"][0]
                                    char_width = span_width / max(len(s_text), 1)
                                    fill_start_x = s["bbox"][0] + len(label_prefix) * char_width
                                else:
                                    fill_start_x = s["bbox"][0]
                            else:
                                fill_start_x = s["bbox"][0]
                        elif not filler_found:
                            label_text += s_text
                            label_end_x = s["bbox"][2]

                    # If no label found before filler, use the line start
                    if not label_text.strip():
                        fill_start_x = first_span_bbox[0]
                    else:
                        # Position fill right after the label text
                        fill_start_x = max(fill_start_x, label_end_x + 2)

                    fill_end_x = last_span_bbox[2]

                    # Build a clean label for the AI (strip filler chars)
                    clean_label = _FILLER_RE.sub("", line_text).strip()
                    if clean_label.endswith(":"):
                        clean_label = clean_label[:-1].strip()

                    z = {
                        "id": f"z{zone_id}",
                        "page": page_num + 1,
                        "type": "fill_line",
                        "label": clean_label[:100] if clean_label else line_text[:100],
                        "x": fill_start_x,
                        "y": line_bbox[1],  # top of line (we'll compute baseline at fill time)
                        "y_bottom": line_bbox[3],
                        "max_width": fill_end_x - fill_start_x,
                        "line_height": line_height,
                        "font_size": min(font_size, 9),
                        # Store the full line bbox for clearing
                        "clear_rect": [fill_start_x - 1, line_bbox[1], fill_end_x + 1, line_bbox[3]],
                    }
                    zones.append(z)
                    page_zones.append(z)
                    zone_id += 1

                # ── Zone type A: label ending with ":" with space after ──
                elif line_text.rstrip().endswith(":"):
                    space_after = effective_right - last_span_bbox[2]
                    if space_after > 40:  # enough room to fill after the colon
                        z = {
                            "id": f"z{zone_id}",
                            "page": page_num + 1,
                            "type": "after_label",
                            "label": line_text.rstrip(": ").strip(),
                            "x": last_span_bbox[2] + 4,
                            "y": line_bbox[1],
                            "y_bottom": line_bbox[3],
                            "max_width": space_after - 10,
                            "line_height": line_height,
                            "font_size": min(font_size, 9),
                            "clear_rect": None,  # nothing to clear for after-label
                        }
                        zones.append(z)
                        page_zones.append(z)
                        zone_id += 1

                # ── Zone type C: line ending with empty space (> 50% of page) ──
                elif last_span_bbox[2] < page_width * 0.5 and len(line_text) < 40:
                    # Short label with lots of space → likely a field
                    space_after = effective_right - last_span_bbox[2]
                    if space_after > 150:
                        z = {
                            "id": f"z{zone_id}",
                            "page": page_num + 1,
                            "type": "after_short_label",
                            "label": line_text,
                            "x": last_span_bbox[2] + 8,
                            "y": line_bbox[1],
                            "y_bottom": line_bbox[3],
                            "max_width": space_after - 15,
                            "line_height": line_height,
                            "font_size": min(font_size, 9),
                            "clear_rect": None,
                        }
                        zones.append(z)
                        page_zones.append(z)
                        zone_id += 1

        # Build AI-friendly text for this page
        if page_zones:
            ai_parts.append(f"\n=== PAGE {page_num + 1} — ZONES REMPLISSABLES ===")
            # Also include full page text for context
            page_text = page.get_text("text")
            if page_text.strip():
                ai_parts.append(f"[Contenu de la page pour contexte:]\n{page_text[:3000]}")
            ai_parts.append("")
            for z in page_zones:
                ai_parts.append(f'  {z["id"]}: ({z["type"]}) Label: "{z["label"]}"')
        else:
            # Include page text even without zones for context
            page_text = page.get_text("text")
            if page_text.strip():
                ai_parts.append(f"\n=== PAGE {page_num + 1} (pas de zones détectées) ===")
                ai_parts.append(page_text[:3000])

    result["zones"] = zones
    result["text_for_ai"] = "\n".join(ai_parts)
    doc.close()
    return result


def _fill_pdf_with_zones(file_path: str, fill_data: list, zone_map: dict) -> bytes:
    """Fill a PDF using pre-computed zone coordinates.

    fill_data: [{"zone_id": "z0", "value": "text"}, ...] or [{"field": ..., "value": ...}, ...]
    zone_map: dict mapping zone_id → zone dict with x, y, page, font_size, max_width, clear_rect
    """
    import fitz
    doc = fitz.open(file_path)

    # ── Handle form field fills ──
    field_fills = [e for e in fill_data if e.get("field")]
    if field_fills:
        for page_num in range(len(doc)):
            page = doc[page_num]
            for widget in page.widgets() or []:
                field_name = widget.field_name or ""
                for entry in field_fills:
                    if entry["field"].strip().lower() == field_name.strip().lower():
                        value = entry.get("value")
                        if value is not None and str(value).strip() != "[A COMPLÉTER]":
                            widget.field_value = str(value)
                            widget.update()
                        break

    # ── Handle zone-based fills ──
    zone_fills = [e for e in fill_data if e.get("zone_id")]

    # Step 1: Collect all clear_rect areas per page, then apply redactions to erase
    # dots/underscores before writing new text
    page_rects: Dict[int, list] = {}
    zone_fill_entries = []
    for entry in zone_fills:
        zone_id = entry.get("zone_id", "")
        value = str(entry.get("value", "")).strip()

        if not value or value == "[A COMPLÉTER]":
            continue
        if zone_id not in zone_map:
            logger.warning("fill-pdf: zone_id '%s' not found in zone_map, skipping", zone_id)
            continue

        zone = zone_map[zone_id]
        page_idx = int(zone["page"]) - 1
        if page_idx < 0 or page_idx >= len(doc):
            continue

        zone_fill_entries.append((zone_id, value, zone, page_idx))

        # Collect clear rectangles
        clear_rect = zone.get("clear_rect")
        if clear_rect:
            page_rects.setdefault(page_idx, []).append(clear_rect)

    # Apply redaction annotations to erase filler content (dots, underscores)
    for page_idx, rects in page_rects.items():
        page = doc[page_idx]
        for rect_coords in rects:
            rect = fitz.Rect(rect_coords)
            # Add a redaction annotation that will white-out the area
            page.add_redact_annot(rect, fill=(1, 1, 1))  # white fill
        page.apply_redactions()  # apply all redactions for this page at once

    # Step 2: Insert the fill text in the cleared areas
    filled_count = 0
    for zone_id, value, zone, page_idx in zone_fill_entries:
        page = doc[page_idx]
        x = float(zone["x"])
        y_top = float(zone["y"])
        y_bottom = float(zone.get("y_bottom", y_top + 12))
        font_size = float(zone.get("font_size", 9))
        max_width = float(zone.get("max_width", 300))
        line_height = float(zone.get("line_height", y_bottom - y_top))

        # Ensure font_size fits within the line height (leave 1pt padding)
        if font_size > line_height - 1:
            font_size = max(line_height - 1, 5)

        # Compute baseline: position text so it sits within the line box
        # PyMuPDF insert_text uses the baseline as the y coordinate.
        # baseline ≈ top of line + ascent ≈ top + font_size * 0.85
        baseline_y = y_top + font_size * 0.85
        # Make sure baseline doesn't go below the line bottom
        if baseline_y > y_bottom - 1:
            baseline_y = y_bottom - 1

        # Calculate how many characters fit on one line
        # Use a more accurate estimate: average char width ≈ font_size * 0.52
        char_width_est = font_size * 0.52
        max_chars = int(max_width / char_width_est) if char_width_est > 0 else 50

        # Truncate value to fit
        if len(value) > max_chars:
            value = value[:max(max_chars - 1, 1)] + "…"

        point = fitz.Point(x, baseline_y)
        try:
            page.insert_text(
                point,
                value,
                fontname="helv",  # Helvetica — clean sans-serif
                fontsize=font_size,
                color=(0.05, 0.05, 0.35),  # Dark blue to distinguish filled text
            )
            filled_count += 1
        except Exception as exc:
            logger.warning("fill-pdf: failed to insert text at zone '%s': %s", zone_id, exc)

    logger.info("fill-pdf: filled %d zones out of %d entries", filled_count, len(zone_fills))
    output = doc.tobytes()
    doc.close()
    return output


@router.post("/{project_id}/fill-pdf/{doc_id}")
async def fill_pdf_document(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Launch async PDF fill as a background task (returns immediately)."""
    result = await db.execute(
        select(ResponseDocument)
        .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
    )
    resp_doc = result.scalar_one_or_none()
    if not resp_doc:
        raise HTTPException(status_code=404, detail="Document livrable non trouvé")

    proj_result = await db.execute(select(RFPProject).where(RFPProject.id == project_id))
    project = proj_result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Projet non trouvé")

    task_key = str(doc_id)
    existing = get_or_idle(_NS_FILL_PDF, task_key)
    if existing.get("status") == "running":
        raise HTTPException(status_code=409, detail="Remplissage deja en cours pour ce document")

    set_progress(_NS_FILL_PDF, task_key, {
        "status": "running", "step": "queued", "progress": 0,
        "message": f"En file d'attente: {resp_doc.title}",
        "doc_title": resp_doc.title,
    })

    from ..tasks.project_tasks import fill_pdf_task
    fill_pdf_task.apply_async(
        args=(str(project_id), str(doc_id), str(project.workspace_id)),
        priority=1,
    )
    return {"success": True, "message": "Remplissage PDF lance en arriere-plan"}


@router.get("/{project_id}/fill-pdf-status/{doc_id}")
async def get_fill_pdf_status(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Poll progress of PDF fill task."""
    return get_or_idle(_NS_FILL_PDF, str(doc_id))


@router.get("/{project_id}/fill-pdf-download/{doc_id}")
async def download_filled_pdf(
    project_id: uuid.UUID,
    doc_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
):
    """Download the filled PDF file once the task is complete."""
    status = get_or_idle(_NS_FILL_PDF, str(doc_id))
    if status.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Fichier pas encore pret")

    file_path = status.get("file_path", "")
    filename = status.get("filename", "document_rempli.pdf")
    if not file_path or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="Fichier genere introuvable")

    from fastapi.responses import FileResponse
    return FileResponse(
        file_path,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _run_fill_pdf(project_id: uuid.UUID, doc_id: uuid.UUID, workspace_id: uuid.UUID):
    """Background task: generate a filled PDF file for a completion-type document."""
    from ..database import task_session
    task_key = str(doc_id)

    def _update(step: str, progress: int, message: str):
        set_progress(_NS_FILL_PDF, task_key, {
            "status": "running", "step": step,
            "progress": progress, "message": message,
        })

    try:
        # ── Phase 1: Load data (short DB session) ──
        async with task_session() as db:
            result = await db.execute(
                select(ResponseDocument)
                .where(ResponseDocument.id == doc_id, ResponseDocument.project_id == project_id)
            )
            resp_doc = result.scalar_one_or_none()
            if not resp_doc:
                set_progress(_NS_FILL_PDF, task_key, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Document livrable non trouve",
                })
                return

            doc_title = resp_doc.title or ""
            doc_source_ids = resp_doc.source_document_ids or []
            doc_source_cats = resp_doc.source_categories or []
            doc_include_generated = resp_doc.include_generated_content or False
            doc_custom_notes = resp_doc.custom_notes or ""
            has_custom_sources = bool(doc_source_ids or doc_source_cats or doc_include_generated)
            _update("loading", 5, f"Chargement: {doc_title}")

            docs_result = await db.execute(
                select(Document)
                .where(Document.project_id == project_id)
                .where(Document.category == DocumentCategory.NEW_RFP)
            )
            all_dce_docs = docs_result.scalars().all()

            all_pdfs = [doc for doc in all_dce_docs if doc.file_type.value == "pdf"]
            pdf_doc = _match_source_document(resp_doc, all_pdfs)
            if not pdf_doc and all_pdfs:
                pdf_doc = all_pdfs[0]

            if not pdf_doc or not os.path.isfile(pdf_doc.file_path):
                set_progress(_NS_FILL_PDF, task_key, {
                    "status": "error", "step": "error", "progress": 0,
                    "message": "Aucun fichier PDF source trouve dans le DCE.",
                })
                return

            pdf_file_path = pdf_doc.file_path
            pdf_original_filename = pdf_doc.original_filename

        # ── Phase 2: Extract PDF structure ──
        _update("reading", 10, "Analyse de la structure du PDF...")
        pdf_zones = _extract_pdf_zones(pdf_file_path)
        logger.info(
            "fill-pdf '%s': has_form_fields=%s, detected %d zones",
            doc_title, pdf_zones["has_form_fields"], len(pdf_zones["zones"]),
        )

        # ── Phase 3: Load context (short DB session) ──
        _update("loading_context", 15, "Chargement du contexte AO + ancienne reponse...")
        async with task_session() as db:
            ai_service = await _get_ai_service(workspace_id, db)

            # If user selected specific source categories/documents, load only those
            if has_custom_sources:
                parts = []
                if doc_source_cats:
                    parts.append(await _get_chunks_anonymized_by_categories(
                        db, project_id, doc_source_cats
                    ))
                if doc_source_ids:
                    parts.append(await _get_chunks_anonymized_by_document_ids(
                        db, project_id, doc_source_ids
                    ))
                if doc_include_generated:
                    parts.append(await _get_generated_chapters_context(db, project_id))
                anon_context = "\n\n".join(p for p in parts if p)
                anon_new_rfp = await _get_all_chunks_anonymized_by_category(
                    db, project_id, DocumentCategory.NEW_RFP
                )
            else:
                anon_new_rfp, anon_old_response = await asyncio.gather(
                    _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.NEW_RFP),
                    _get_all_chunks_anonymized_by_category(db, project_id, DocumentCategory.OLD_RESPONSE),
                )
                anon_context = None

        if anon_context is not None:
            old_response_with_context = anon_context
        else:
            # Vector search (no DB needed)
            _update("searching", 25, "Recherche de contenu pertinent...")
            search_query = f"{doc_title} formulaire informations candidat société entreprise"
            from ..tasks.chapter_tasks import _vector_search
            relevant_chunks = _vector_search(str(project_id), search_query, top_k=15, category_filter="old_response")
            relevant_context = "\n\n".join([
                f"[{c['document_name']} p.{c['page_number']}] {c['content']}"
                for c in relevant_chunks
            ])

            old_response_with_context = anon_old_response
            if relevant_context:
                old_response_with_context = (
                    f"=== EXTRAITS PERTINENTS DE L'ANCIENNE RÉPONSE ===\n{relevant_context}\n\n"
                    f"=== CONTENU COMPLET ANCIENNE RÉPONSE ===\n{anon_old_response}"
                )

        # ── Phase 4: AI generation (NO DB held) ──
        _update("generating", 30, f"Generation IA du contenu pour {doc_title}...")
        fill_data = await ai_service.generate_pdf_fill_data(
            document_title=doc_title,
            pdf_structure=pdf_zones["text_for_ai"],
            new_rfp_content=anon_new_rfp,
            old_response_content=old_response_with_context,
            has_form_fields=pdf_zones["has_form_fields"],
            custom_notes=doc_custom_notes,
        )
        logger.info("fill-pdf %s: AI returned %d fill entries", doc_title, len(fill_data))

        # Log AI usage
        async with task_session() as usage_db:
            await log_ai_usage_from_service(usage_db, project_id, "fill_pdf", ai_service)

        # ── Phase 5: Fill PDF and save to disk ──
        _update("filling", 85, "Remplissage du fichier PDF...")
        zone_map = {z["id"]: z for z in pdf_zones["zones"]}
        filled_bytes = _fill_pdf_with_zones(pdf_file_path, fill_data, zone_map)

        base_name = os.path.splitext(pdf_original_filename)[0]
        output_filename = f"{base_name}_rempli.pdf"
        from ..config import settings
        output_path = os.path.join(settings.export_dir, f"{doc_id}_{output_filename}")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(filled_bytes)

        set_progress(_NS_FILL_PDF, task_key, {
            "status": "completed", "step": "done", "progress": 100,
            "message": f"PDF rempli: {output_filename} ({len(fill_data)} champs)",
            "file_path": output_path,
            "filename": output_filename,
            "field_count": len(fill_data),
        })

    except Exception as e:
        logger.exception("Fill PDF failed for doc %s", doc_id)
        set_progress(_NS_FILL_PDF, task_key, {
            "status": "error", "step": "error", "progress": 0,
            "message": f"Erreur: {str(e)[:200]}",
        })
